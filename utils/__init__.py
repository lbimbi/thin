"""Core utilities and mathematical functions for musical intonation systems.

This module provides essential utility functions, mathematical operations, and
system integration capabilities that support all aspects of THIN and SIM
applications, from basic mathematical operations to complex audio processing.

Mathematical Operations:
- Exact rational arithmetic using fractions.Fraction for precision
- Ratio-to-cents and cents-to-ratio conversions with Ellis factor
- Octave reduction and normalization algorithms
- Harmonic and subharmonic series generation
- Geometric progression calculations with overflow protection

Musical Note Processing:
- Comprehensive note name parsing with microtonal support (+ - ! .)
- MIDI note number conversion and validation
- Frequency calculation with custom diapason settings
- Support for various naming conventions and accidental systems

Type Conversion and Validation:
- Intelligent parsing of numeric values (int, float, Fraction)
- Safe string-to-number conversion with error handling
- Argument validation for command-line interfaces
- Range checking and bounds enforcement

Audio Processing Support:
- Integration with librosa, scipy, and other audio libraries
- Conditional import system for optional dependencies
- Warning suppression for cleaner console output
- Memory-efficient processing for large audio files

File System Operations:
- Cross-platform file path handling and validation
- Atomic file operations with backup creation
- Directory management and cleanup utilities
- Safe file existence checking and creation

System Integration:
- Environment detection and configuration
- Terminal capability detection and formatting
- Multilingual error message support (Italian/English)
- Progress reporting and user feedback systems

Caching and Performance:
- LRU caching for expensive mathematical operations
- Memoization of frequently-used calculations
- Efficient algorithms for large-scale computations
- Memory usage optimization for batch processing

Error Handling:
- Comprehensive exception handling with context preservation
- Graceful degradation when optional features are unavailable
- Detailed error reporting with multilingual support
- Recovery mechanisms for common failure scenarios

Dependencies Management:
- Dynamic import system for optional packages
- Version compatibility checking
- Feature detection and capability reporting
- Fallback implementations when dependencies are missing

This module serves as the foundation layer that all other modules depend on,
ensuring consistent behavior and reliable operation across the entire application.
"""

import argparse
import functools
import logging
import math
import os
import sys
import warnings
from fractions import Fraction
from typing import List, Tuple, Optional, Iterable, Union, Any, Dict

import consts

# --- Warning suppression and logging system ---

def setup_warning_logging(log_file: str = "warnings.log") -> None:
    """Setup warning suppression and logging system for external modules."""
    # Configure logging for warnings
    logging.basicConfig(
        level=logging.WARNING,
        format='%(asctime)s - %(name)s - %(levelname)s - %(message)s',
        handlers=[
            logging.FileHandler(log_file, encoding='utf-8'),
        ]
    )

    # Capture warnings and redirect to logger
    def warning_handler(message, category, filename, lineno):
        logger = logging.getLogger('warnings')
        logger.warning(f"{category.__name__}: {message} ({filename}:{lineno})")

    warnings.showwarning = warning_handler

    # Suppress specific warnings from external modules
    warnings.filterwarnings('ignore', category=FutureWarning, module='tensorflow.*')
    warnings.filterwarnings('ignore', category=DeprecationWarning, module='tensorflow.*')
    warnings.filterwarnings('ignore', category=UserWarning, module='tensorflow.*')
    warnings.filterwarnings('ignore', category=FutureWarning, module='librosa.*')
    warnings.filterwarnings('ignore', category=DeprecationWarning, module='librosa.*')
    warnings.filterwarnings('ignore', category=UserWarning, module='numpy.*')
    warnings.filterwarnings('ignore', category=RuntimeWarning, module='numpy.*')
    warnings.filterwarnings('ignore', category=FutureWarning, module='scipy.*')
    warnings.filterwarnings('ignore', category=DeprecationWarning, module='scipy.*')

def suppress_tensorflow_warnings() -> None:
    """Suppress TensorFlow warnings specifically."""
    os.environ['TF_CPP_MIN_LOG_LEVEL'] = '3'  # Suppress TF logs (0=all, 1=info+, 2=warn+, 3=error+)

    try:
        import tensorflow as tf
        tf.get_logger().setLevel('ERROR')
        tf.autograph.set_verbosity(0)
    except ImportError:
        pass

# --- Common scale generation utilities ---

@functools.lru_cache(maxsize=1)
def generate_pythagorean_12() -> List[float]:
    """Generate normalized Pythagorean 12-tone scale ratios. Cached."""
    try:
        from fractions import Fraction as _Fr
        # Use geometric ratios with 3/2 generator, 12 steps, reduce to octave
        ratios = generate_geometric_ratios(
            generator=_Fr(3, 2),
            steps=12,
            interval=2.0,
            reduce_to_interval=True
        )
        return normalize_ratios(ratios, reduce_octave=True)
    except (ImportError, ValueError, TypeError):
        # Fallback to float arithmetic
        ratios = generate_geometric_ratios(
            generator=1.5,
            steps=12,
            interval=2.0,
            reduce_to_interval=True
        )
        return normalize_ratios(ratios, reduce_octave=True)

@functools.lru_cache(maxsize=1)
def generate_pythagorean_7() -> List[float]:
    """Generate normalized Pythagorean 7-tone diatonic scale ratios. Cached."""
    try:
        from fractions import Fraction as _Fr
        # The diatonic scale uses specific powers of 3/2: [0, 2, 4, -1, 1, 3, 5]
        # Generate each ratio using generate_geometric_ratios for consistency
        powers = [0, 2, 4, -1, 1, 3, 5]  # Powers for C, D, E, F, G, A, B
        ratios = []

        for power in powers:
            if power >= 0:
                # Generate positive powers: (3/2)^power
                if power == 0:
                    ratios.append(1.0)
                else:
                    series = generate_geometric_ratios(
                        generator=_Fr(3, 2),
                        steps=power + 1,
                        interval=2.0,
                        reduce_to_interval=True
                    )
                    ratios.append(series[power])  # Take the power-th element
            else:
                # For negative powers: 1 / (3/2)^|power| = (2/3)^|power|
                abs_power = abs(power)
                series = generate_geometric_ratios(
                    generator=_Fr(2, 3),  # Inverse of 3/2
                    steps=abs_power + 1,
                    interval=2.0,
                    reduce_to_interval=True
                )
                ratios.append(series[abs_power])

        return sorted(ratios)  # Sort to get ascending order
    except (ImportError, ValueError, TypeError):
        # Fallback to hardcoded values
        return [1.0, 9/8, 81/64, 4/3, 3/2, 27/16, 243/128]

@functools.lru_cache(maxsize=32)
def generate_tet_ratios(n: int = 12) -> List[float]:
    """Generate n-TET scale ratios. Cached for performance."""
    if n <= 0:
        n = 12
    return [2.0 ** (k / float(n)) for k in range(n)]

def generate_geometric_ratios(generator: Union[float, Fraction], steps: int,
                            interval: Union[float, Fraction] = 2.0,
                            reduce_to_interval: bool = True) -> List[float]:
    """Generate geometric progression ratios.

    Args:
        generator: The generator ratio (e.g., 3/2 for perfect fifth)
        steps: Number of steps to generate
        interval: Interval for reduction (default: octave = 2.0)
        reduce_to_interval: Whether to reduce ratios to the specified interval

    Returns:
        List of ratio values as floats
    """
    if steps <= 0:
        return [1.0]

    ratios = []
    for i in range(steps):
        # Calculate ratio using appropriate arithmetic
        if isinstance(generator, Fraction):
            r = pow_fraction(generator, i)
        else:
            r = float(generator) ** i

        # Reduce to interval if requested
        if reduce_to_interval and i > 0:
            r = reduce_to_interval_value(r, interval)

        ratios.append(float(r))

    return ratios

def reduce_to_interval_value(value: Union[float, Fraction], interval: Union[float, Fraction]) -> Union[float, Fraction]:
    """Reduce a value to the specified interval [1, interval).
    This is a wrapper for reduce_to_interval with better naming.
    """
    return reduce_to_interval(value, interval)

# --- Audio analysis utilities (extracted from duplicated code) ---

def safe_import(module_name: str):
    """Safely import a module. Returns module or None."""
    try:
        return __import__(module_name)
    except ImportError:
        return None

def ensure_numpy_import():
    """Ensure numpy is imported with alias _np."""
    return safe_import('numpy')

def ensure_librosa_import():
    """Ensure librosa is imported with alias _lb."""
    return safe_import('librosa')

def ensure_warnings_import():
    """Ensure warnings is imported with alias _warnings."""
    return safe_import('warnings')

def ratio_to_cents(ratio: float) -> float:
    """Convert ratio to cents using logarithmic formula."""
    try:
        return 1200.0 * math.log2(float(ratio))
    except (ValueError, ZeroDivisionError):
        return 0.0

def find_closest_circular_index(target_cents: float, degree_list: list, modulus: float = 1200.0) -> tuple:
    """Find closest index in degree list using circular distance. Returns (index, error)."""
    if not degree_list:
        return 0, 0.0
    idx = min(range(len(degree_list)), key=lambda k: circular_difference(target_cents, degree_list[k], modulus))
    err = circular_difference(target_cents, degree_list[idx], modulus)
    return idx, err

def accumulate_weighted_error(error: float, count: int, total_werr: float = 0.0, total_w: float = 0.0) -> tuple:
    """Accumulate weighted error with count. Returns (new_total_werr, new_total_w)."""
    weight = max(1, int(count))
    return total_werr + float(error) * weight, total_w + weight

def calculate_f0_params(sr: float, frame_size: int, fmin: float, fmax: float) -> tuple:
    """Calculate effective F0 min/max parameters based on frame size. Returns (fmin_eff, fmax_eff)."""
    required_min_fmin = (2.0 * sr) / float(frame_size) if frame_size and sr else fmin
    fmin_eff = fmin
    try:
        fmin_eff = max(float(fmin), float(required_min_fmin) + 1e-6)
    except ValueError:
        fmin_eff = fmin
    if fmin_eff >= fmax:
        fmin_eff = 0.5 * fmax
    return fmin_eff, fmax

def fold_frequency_to_octave(f: float, low: float, high: float) -> float:
    """Fold frequency into specified octave range [low, high)."""
    if f <= 0 or low <= 0 or high <= low:
        return f
    while f >= high:
        f *= 0.5
    while f < low:
        f *= 2.0
    return f

def circular_difference(a: float, b: float, mod: float = 1200.0) -> float:
    """Calculate circular distance between two values with given modulus."""
    d = abs(a - b) % mod
    return d if d <= (mod * 0.5) else (mod - d)

def load_mono_audio(audio_path: str, verbose: bool = False):
    """Load audio file as mono, handling stereo conversion.
    Returns (y, sr) on success, or None on error.
    Set verbose=True to print informative messages.
    """
    try:
        import librosa as _lb
        # Load audio preserving original channel structure, native sampling rate
        y, sr = _lb.load(audio_path, sr=None, mono=False)

        # Convert to mono if stereo by summing channels
        if hasattr(y, 'ndim') and y.ndim > 1:
            if verbose:
                print(f"\nStereo audio detected ({getattr(y, 'shape', ['?'])[0]} channels), converting to mono for analysis")
            y = _lb.to_mono(y)
        return y, sr
    except (ImportError, FileNotFoundError, ValueError, RuntimeError) as e:
        if verbose:
            print(f"Audio loading error: {e}")
        return None

def extract_f0_with_pyin_yin_fallback(y, sr, fmin: float, fmax: float, frame_size: int, hop_size: int):
    """Extract F0 using pYIN with YIN fallback. Returns filtered numpy array."""
    try:
        import numpy as _np
        import librosa as _lb
        import warnings as _warnings

        with _warnings.catch_warnings():
            _warnings.simplefilter("ignore")
            try:
                f0_series, _, _ = _lb.pyin(y, fmin=fmin, fmax=fmax, frame_length=frame_size,
                                           hop_length=hop_size, sr=sr)
                f0_vals = _np.asarray(
                    [v for v in _np.nan_to_num(f0_series, nan=_np.nan) if _np.isfinite(v) and v > 0])
            except (ValueError, RuntimeError, TypeError):
                f0_track = _lb.yin(y, fmin=fmin, fmax=fmax, frame_length=frame_size, hop_length=hop_size, sr=sr)
                f0_vals = _np.asarray([v for v in f0_track if _np.isfinite(v) and v > 0])
        return f0_vals
    except ImportError:
        import numpy as _np
        return _np.array([], dtype=float)


def write_wav_with_fallback(out_path: str, audio_data, sr: int) -> bool:
    """Write WAV file with soundfile/scipy fallback. Returns success bool."""
    try:
        import numpy as _np

        # Validate inputs
        if audio_data is None:
            print("WAV write error: audio_data is None")
            return False

        # Ensure audio_data is a numpy array or can be converted to one
        try:
            audio_data = _np.asarray(audio_data)
            if audio_data.size == 0:
                print("WAV write error: audio_data is empty")
                return False
        except (ValueError, TypeError) as e:
            print(f"WAV write error: cannot convert audio_data to array: {e}")
            return False

        # Ensure output directory exists
        try:
            _dir = os.path.dirname(out_path)
            if _dir:
                os.makedirs(_dir, exist_ok=True)
        except (OSError, PermissionError):
            pass
        try:
            import soundfile as _sf
            _sf.write(out_path, audio_data, sr)
            return True
        except (ImportError, OSError, RuntimeError, ValueError):
            # Fall back to scipy if soundfile is not available or cannot write
            try:
                from scipy.io import wavfile as _wio
                # Safely convert to int16 format for scipy
                audio_int16 = (audio_data * 32767.0).astype(_np.int16)
                _wio.write(out_path, int(sr), audio_int16)
                return True
            except (ImportError, ValueError, TypeError, AttributeError) as e:
                print(f"WAV write error (scipy fallback): {e}")
                return False
    except (FileNotFoundError, OSError, PermissionError) as e:
        print(f"WAV write error: {e}")
        return False

def cluster_ratios_by_proximity(ratios_reduced: list, tolerance: float = 0.02) -> list:
    """Cluster ratios by relative proximity. Returns [(center, count), ...]."""
    if not ratios_reduced:
        return []

    rr_sorted = sorted(ratios_reduced)
    centers = []
    cur_center = rr_sorted[0]
    cur_sum = rr_sorted[0]
    cur_count = 1

    for val in rr_sorted[1:]:
        if abs(val - cur_center) <= tolerance * cur_center:
            cur_sum += val
            cur_count += 1
            cur_center = cur_sum / cur_count
        else:
            centers.append((cur_center, cur_count))
            cur_center = val
            cur_sum = val
            cur_count = 1
    centers.append((cur_center, cur_count))
    return centers


def cluster_ratios_weighted(ratios_reduced: List[float], weights: List[float], tolerance: float = 0.02) -> List[Tuple[float, int]]:
    """Clustering con pesi per i ratios (versione pesata).

    Args:
        ratios_reduced: Lista dei ratios ridotti all'ottava
        weights: Pesi corrispondenti (es. durata, ampiezza, confidence)
        tolerance: Tolleranza relativa per il clustering

    Returns:
        Lista di (centro_pesato, count_totale)
    """
    if not ratios_reduced:
        return []

    if len(weights) != len(ratios_reduced):
        # Fallback to simple clustering if weights don't match
        return cluster_ratios_by_proximity(ratios_reduced, tolerance)

    # Crea lista di (ratio, peso) e ordina per ratio
    ratio_weight_pairs = list(zip(ratios_reduced, weights))
    ratio_weight_pairs.sort(key=lambda x: x[0])

    centers = []
    cur_weighted_sum = ratio_weight_pairs[0][0] * ratio_weight_pairs[0][1]
    cur_weight_sum = ratio_weight_pairs[0][1]
    cur_count = 1
    cur_center = ratio_weight_pairs[0][0]

    for ratio, weight in ratio_weight_pairs[1:]:
        if abs(ratio - cur_center) <= tolerance * cur_center:
            cur_weighted_sum += ratio * weight
            cur_weight_sum += weight
            cur_count += 1
            cur_center = cur_weighted_sum / cur_weight_sum if cur_weight_sum > 0 else ratio
        else:
            centers.append((cur_center, cur_count))
            cur_weighted_sum = ratio * weight
            cur_weight_sum = weight
            cur_count = 1
            cur_center = ratio

    centers.append((cur_center, cur_count))
    return centers


def cluster_ratios_adaptive(ratios_reduced: List[float], base_tolerance: float = 0.02,
                          scale_size_hint: Optional[int] = None) -> List[Tuple[float, int]]:
    """Clustering adattivo con tolleranza variabile basata sulla densità locale.

    Args:
        ratios_reduced: Lista dei ratios ridotti all'ottava
        base_tolerance: Tolleranza base
        scale_size_hint: Numero atteso di gradi della scala (opzionale)

    Returns:
        Lista di (centro, count)
    """
    if not ratios_reduced:
        return []

    if len(ratios_reduced) < 3:
        # Fallback per pochi campioni
        return cluster_ratios_by_proximity(ratios_reduced, base_tolerance)

    rr_sorted = sorted(ratios_reduced)

    # Calcola densità locale per ogni punto
    local_densities = []
    for i, ratio in enumerate(rr_sorted):
        # Conta vicini entro base_tolerance
        neighbors = 0
        for other_ratio in rr_sorted:
            if abs(other_ratio - ratio) <= base_tolerance * ratio and other_ratio != ratio:
                neighbors += 1
        local_densities.append(neighbors)

    # Adatta tolleranza basata su densità
    centers = []
    i = 0

    while i < len(rr_sorted):
        cur_ratio = rr_sorted[i]
        cur_density = local_densities[i]

        # Tolleranza adattiva: minore in zone dense, maggiore in zone sparse
        adaptive_tolerance = base_tolerance
        if cur_density > 2:  # Zona densa
            adaptive_tolerance *= 0.7  # Più stretta
        elif cur_density == 0:  # Zona isolata
            adaptive_tolerance *= 1.5  # Più ampia

        # Trova tutti i ratio nel cluster corrente
        cluster_ratios = [cur_ratio]
        j = i + 1

        while j < len(rr_sorted):
            if abs(rr_sorted[j] - cur_ratio) <= adaptive_tolerance * cur_ratio:
                cluster_ratios.append(rr_sorted[j])
                j += 1
            else:
                break

        # Centro del cluster come media
        center = sum(cluster_ratios) / len(cluster_ratios)
        centers.append((center, len(cluster_ratios)))

        i = j

    # Se abbiamo troppi cluster rispetto al hint, unisci i più vicini
    if scale_size_hint and len(centers) > scale_size_hint * 1.5:
        centers = _merge_closest_clusters(centers, scale_size_hint)

    return centers


def _merge_closest_clusters(centers: List[Tuple[float, int]], target_count: int) -> List[Tuple[float, int]]:
    """Unisce i cluster più vicini fino al numero target."""
    clusters = list(centers)

    while len(clusters) > target_count:
        # Trova la coppia di cluster più vicina
        min_distance = float('inf')
        merge_idx = 0

        for i in range(len(clusters) - 1):
            center1, count1 = clusters[i]
            center2, count2 = clusters[i + 1]
            distance = abs(center2 - center1) / center1  # Distanza relativa

            if distance < min_distance:
                min_distance = distance
                merge_idx = i

        # Unisci i due cluster più vicini
        center1, count1 = clusters[merge_idx]
        center2, count2 = clusters[merge_idx + 1]

        # Nuovo centro pesato per conteggio
        total_count = count1 + count2
        new_center = (center1 * count1 + center2 * count2) / total_count

        # Sostituisci i due cluster con quello unito
        clusters[merge_idx] = (new_center, total_count)
        clusters.pop(merge_idx + 1)

    return clusters


def apply_harmonic_bias(centers: List[Tuple[float, int]], bias_strength: float = 0.1) -> List[Tuple[float, int]]:
    """Applica bias verso rapporti armonici semplici.

    Args:
        centers: Lista di (centro, count) da biasare
        bias_strength: Forza del bias (0.0 = nessun bias, 1.0 = bias massimo)

    Returns:
        Lista di centri aggiustati verso armonici semplici
    """
    # Rapporti armonici semplici comuni
    simple_harmonics = [
        1.0,         # Unisono
        9/8,         # Seconda maggiore
        6/5,         # Terza minore
        5/4,         # Terza maggiore
        4/3,         # Quarta giusta
        7/5,         # Tritono settimale
        3/2,         # Quinta giusta
        8/5,         # Sesta minore
        5/3,         # Sesta maggiore
        16/9,        # Settima minore pitagorica
        15/8,        # Settima maggiore
        2.0          # Ottava (se presente)
    ]

    biased_centers = []

    for center, count in centers:
        # Trova l'armonico più vicino
        closest_harmonic = min(simple_harmonics, key=lambda h: abs(h - center))
        distance_to_harmonic = abs(closest_harmonic - center) / center

        # Applica bias solo se abbastanza vicino (entro 10 cent ~ 0.6%)
        if distance_to_harmonic < 0.006:
            # Interpola verso l'armonico in base alla forza del bias
            biased_center = center * (1 - bias_strength) + closest_harmonic * bias_strength
            biased_centers.append((biased_center, count))
        else:
            # Mantieni il centro originale se troppo lontano
            biased_centers.append((center, count))

    return biased_centers

def normalize_mapping_with_zero_step(mapping: List[Tuple[int, float, int]]) -> List[Tuple[int, float, int]]:
    """Normalize mapping to ensure index 0 with ratio 1.0 is present."""
    try:
        has_zero = any(int(s) == 0 for (s, _, __) in mapping)
    except ValueError:
        has_zero = False

    steps = list(mapping)
    if has_zero:
        steps = [(0 if int(s) == 0 else int(s), (1.0 if int(s) == 0 else float(r)), int(c)) for (s, r, c) in steps]
    else:
        steps = [(0, 1.0, 0)] + steps
    return sorted(steps, key=lambda t: (int(t[0]), float(t[1])))

def clear_screen() -> None:
    """Clear system screen."""
    try:
        os.system("cls" if os.name == "nt" else "clear")
    except OSError:
        pass

def _clear_line(cols: int) -> None:
    print("\r" + (" " * int(cols)) + "\r", end="")



def print_banner() -> None:
    """Stampa sempre le info di programma: nome, versione, data, autore, licenza."""
    # Etichette localizzate
    lbl_ver = "Version"
    lbl_date = "Release"
    lbl_auth = "Author"
    lbl_lic = "License"
    title = consts.__program_name__
    info_line = f"{title}  |  {lbl_ver}: {consts.__version__}  |  {lbl_date}: {consts.__date__}  |  {lbl_auth}: {consts.__author__}  |  {lbl_lic}: {consts.__license__}"
    print(info_line)
    # Copyright / License line
    copyright_line = "Copyright (c) 2025 Luca Bimbi. Distributed under the MIT License. \nSee the LICENSE file for details."
    print(copyright_line)

def file_exists(file_path: str) -> bool:
    """Verifica esistenza file."""
    exists = os.path.exists(file_path)
    en_word = "exists" if exists else "not found"
    print(f"File {en_word}: {file_path}")
    return exists

def apply_cents(freq_hz: float, cents: float) -> float:
    """Applica offset in cents a una frequenza / Apply cents offset to a frequency."""
    return float(freq_hz) * (2.0 ** (float(cents) / 1200.0))

def is_fraction_type(value: Any) -> bool:
    """Verifica se value è una frazione."""
    return isinstance(value, Fraction)


def fraction_to_cents(ratio: Fraction) -> int:
    """Converte un rapporto razionale in cents."""
    if ratio.denominator == 0:
        raise ValueError("Zero denominator in fraction")
    decimal = math.log(ratio.numerator / ratio.denominator)
    return round(decimal * consts.ELLIS_CONVERSION_FACTOR)


def cents_to_fraction(cents: float) -> Fraction:
    """Converte cents in rapporto razionale approssimato."""
    return Fraction(math.exp(cents /consts.ELLIS_CONVERSION_FACTOR)).limit_denominator(10000)

def reduce_to_octave(value: consts.Numeric) -> consts.Numeric:
    """Riduce un rapporto nell'ambito di un'ottava [1, 2)."""
    if isinstance(value, Fraction):
        two = Fraction(2, 1)
        while value >= two:
            value /= two
        while value < 1:
            value *= two
    else:
        value = float(value)
        while value >= 2.0:
            value /= 2.0
        while value < 1.0:
            value *= 2.0
    return value


def reduce_to_interval(value: consts.Numeric, interval: consts.Numeric) -> consts.Numeric:
    """Riduce un rapporto nell'intervallo [1, interval)."""
    try:
        interval_num = float(interval)
    except (TypeError, ValueError):
        return value

    if not math.isfinite(interval_num) or interval_num <= 1.0:
        return value

    if isinstance(value, Fraction) and isinstance(interval, Fraction):
        one = Fraction(1, 1)
        while value >= interval:
            value /= interval
        while value < one:
            value *= interval
    else:
        v = float(value)
        i = float(interval)
        while v >= i:
            v /= i
        while v < 1.0:
            v *= i
        return v

    return value


def normalize_ratios(ratios: Iterable[consts.Numeric], reduce_octave: bool = True) -> List[float]:
    """
    Normalizza una lista di rapporti:
    - Opzionalmente riduce nell'ottava [1, 2)
    - Elimina duplicati con tolleranza
    - Ordina in modo crescente
    """
    processed = []
    seen = []

    for r in ratios:
        v = reduce_to_octave(r) if reduce_octave else r
        v_float = float(v)

        # Check duplicates with tolerance
        if not any(abs(v_float - s) <= consts.RATIO_EPS for s in seen):
            seen.append(v_float)
            processed.append(v_float)

    return sorted(processed)


def repeat_ratios(ratios: List[float], span: int, interval_factor: float) -> List[float]:
    """Ripete la serie di rapporti su un certo ambitus."""
    try:
        s = int(span)
    except (TypeError, ValueError):
        s = 1

    if s <= 1:
        return list(ratios)

    if not isinstance(interval_factor, (int, float)) or interval_factor <= 0:
        return list(ratios)

    out = []
    for k in range(s):
        factor_k = interval_factor ** k
        out.extend(float(r) * factor_k for r in ratios)

    return out


def pow_fraction(fr: consts.Numeric, k: int) -> consts.Numeric:
    """Eleva una Fraction/float ad un esponente intero."""
    k = int(k)
    return fr ** k if isinstance(fr, Fraction) else float(fr) ** k

def int_or_fraction(value: str) -> Union[int, Fraction]:
    """Parser per interi o frazioni."""
    try:
        return int(value)
    except ValueError:
        try:
            return Fraction(value)
        except ValueError:
            raise argparse.ArgumentTypeError(
                f"'{value}' non è un intero o frazione valida. / '{value}' is not a valid integer or fraction."
            )

def note_name_or_frequency(value: str) -> Union[float, str]:
    """Parser per nome nota o frequenza."""
    try:
        return float(value)
    except (ValueError, TypeError):
        return str(value)


def convert_note_name_to_midi(note_name: str) -> int:
    """Convert note name to MIDI value. Supports # and b/B."""
    note_map = {'C': 0, 'D': 2, 'E': 4, 'F': 5, 'G': 7, 'A': 9, 'B': 11}
    s = note_name.strip()
    if not s:
        raise ValueError("Empty note name")

    s_up = s.upper()
    note = s_up[0]
    if note not in note_map:
        raise ValueError(f"Invalid note name: {note_name}")

    alteration = 0
    idx = 1

    # gestisci # e b/B / handle # and b
    if idx < len(s_up):
        if s_up[idx] == '#':
            alteration += 1
            idx += 1
        elif s_up[idx] in ('B', 'b'):
            alteration -= 1
            idx += 1

    try:
        octave = int(s_up[idx:])
    except (ValueError, IndexError):
        raise ValueError(f"Formato ottava non valido in: {note_name}",
                           f"Invalid octave format in: {note_name}")

    midi_value = (octave + 1) * consts.SEMITONES_PER_OCTAVE + note_map[note] + alteration

    if not (consts.MIDI_MIN <= midi_value <= consts.MIDI_MAX):
        raise ValueError(f"Nota MIDI fuori range: {midi_value}",
                           f"MIDI note out of range: {midi_value}")

    return midi_value

def convert_midi_to_hz(midi_value: int, diapason_hz: float = consts.DEFAULT_DIAPASON) -> float:
    """Converte valore MIDI in frequenza Hz."""
    return diapason_hz * (2 ** ((float(midi_value) - consts.MIDI_A4) / consts.SEMITONES_PER_OCTAVE))


def midi_to_note_name_12tet(midi_value: int) -> str:
    """Converte un valore MIDI in nome nota 12-TET (con #) con ottava, es. A4, C#3."""
    try:
        m = int(midi_value)
    except (ValueError, TypeError, ArithmeticError):
        return ""
    if m < consts.MIDI_MIN or m > consts.MIDI_MAX:
        m = max(consts.MIDI_MIN, min(consts.MIDI_MAX, m))
    names = ["C", "C#", "D", "D#", "E", "F", "F#", "G", "G#", "A", "A#", "B"]
    name = names[m % consts.SEMITONES_PER_OCTAVE]
    octave = (m // consts.SEMITONES_PER_OCTAVE) - 1
    return f"{name}{octave}"

def ensure_midi_fit(ratios: List[float], basekey: int,
                    prefer_truncate: bool) -> Tuple[List[float], int]:
    """Ensures that ratios fit within MIDI range."""
    n = len(ratios)

    try:
        bk = int(basekey)
    except (TypeError, ValueError):
        bk = 0

    if n > 128:
        if not prefer_truncate:
            print(f"WARNING: number of steps ({n}) exceeds 128. Only the first 128 will be kept.")
            prefer_truncate = True
        n = 128
        ratios = ratios[:n]

    if prefer_truncate:
        bk = max(consts.MIDI_MIN, min(consts.MIDI_MAX, bk))
        max_len = 128 - bk
        if len(ratios) > max_len:
            print(f"WARNING: series exceeds MIDI limit. Truncated {len(ratios) - max_len} steps.")
            ratios = ratios[:max_len]
        return ratios, bk
    else:
        allowed_min = consts.MIDI_MIN
        allowed_max = consts.MIDI_MAX - (n - 1)
        if allowed_max < allowed_min:
            allowed_max = allowed_min

        bk_eff = bk
        if bk < allowed_min or bk > allowed_max:
            bk_eff = max(allowed_min, min(allowed_max, bk))
            print(f"WARNING: basekey adjusted from {bk} to {bk_eff} to include all steps.")

        return ratios, bk_eff

def page_lines(lines: List[str], rows_per_page: Optional[int] = None) -> None:
    """Simple pager: prints lines with a --More-- prompt, Enter continues, q quits.
    Caps the page height at 80 rows by requirement.
    """
    # If not running in an interactive TTY, just print and return (avoid blocking on input)
    try:
        if not (getattr(sys.stdin, "isatty", lambda: False)() and getattr(sys.stdout, "isatty", lambda: False)()):
            for ln in lines:
                print(ln)
            return
    except (ValueError, TypeError, ImportError):
        # On any detection error, fall back to non-interactive behavior
        for ln in lines:
            print(ln)
        return

    # Determine terminal rows
    if rows_per_page is None:
        cols, term_rows = (80, 24)
        rows = term_rows - 1
    else:
        rows = int(rows_per_page)
        cols = 80

    # Cap at 80 rows and enforce minimum
    rows = max(5, min(80, rows))

    i = 0
    n = len(lines)
    while i < n:
        # print up to rows lines
        end = min(n, i + rows)
        for j in range(i, end):
            print(lines[j])
        i = end
        if i < n:
            prompt = "--More-- (invio=continua, q=esci)"
            try:
                print(prompt, end="", flush=True)
            except (ValueError, TypeError, OSError):
                pass
            # Wait for Enter to continue or 'q' to quit
            while True:
                try:
                    ch = sys.stdin.read(1)
                except (OSError, ValueError, AttributeError):
                    try:
                        ch = input()
                    except EOFError:
                        ch = "q"
                    ch = ch[0] if ch else "\n"
                if ch.lower() == 'q' or ch in ('\n', '\r'):
                    break
            # clear prompt line according to terminal width
            try:
                _clear_line(cols)
            except (ValueError, TypeError, OSError):
                pass
            if ch.lower() == 'q':
                break


def print_step_hz_table(ratios: List[float], basenote_hz: float) -> None:
    """Prints multi-column table with Step/Hz."""
    pairs = [(str(i + 1), f"{basenote_hz * float(r):.2f}")
             for i, r in enumerate(ratios)]

    if not pairs:
        page_lines(["Step Hz"])
        return

    # Calcola larghezze
    step_w = max(len("Step"), max(len(p[0]) for p in pairs))
    hz_w = max(len("Hz"), max(len(p[1]) for p in pairs))
    cell_w = step_w + 1 + hz_w
    gap = 3

    # Larghezza terminale
    term_w = 80

    # Calcola colonne
    cols = max(1, (term_w + gap) // (cell_w + gap)) if cell_w < term_w else 1
    rows_count = math.ceil(len(pairs) / cols)

    # Compose lines instead of printing directly
    lines_out: List[str] = []
    header_cell = "Step".ljust(step_w) + " " + "Hz".ljust(hz_w)
    lines_out.append((" " * gap).join([header_cell] * cols))

    for r in range(rows_count):
        line_cells = []
        for c in range(cols):
            idx = c * rows_count + r
            if idx < len(pairs):
                s, h = pairs[idx]
                cell = s.rjust(step_w) + " " + h.rjust(hz_w)
                line_cells.append(cell)
        if line_cells:
            lines_out.append((" " * gap).join(line_cells))

    page_lines(lines_out)

def parse_note_with_microtones(note_name: str) -> Tuple[int, float]:
    """
    Parsing nota con microtoni per basenote e fondamentali.
    Simboli: # (diesis), b/B (bemolle), + (+50c), - (-50c), ! (+25c), . (-25c).
    Restituisce (midi_int, cents_offset).
    """
    note_map = {'C': 0, 'D': 2, 'E': 4, 'F': 5, 'G': 7, 'A': 9, 'B': 11}
    s = str(note_name).strip()
    if not s:
        raise ValueError("Nome nota vuoto")
    s_up = s.upper()
    note = s_up[0]
    if note not in note_map:
        raise ValueError(f"Nome nota non valida: {note_name}")
    idx = 1
    semitones = 0
    cents = 0.0

    # Accidenti base e simboli microtonali
    while idx < len(s_up) and s_up[idx] in ('#', 'B', 'b', '+', '-', '!', '.'):
        ch = s_up[idx]
        if ch == '#':
            semitones += 1
        elif ch in ('B', 'b'):
            semitones -= 1
        elif ch == '+':
            cents += 50.0
        elif ch == '-':
            cents -= 50.0
        elif ch == '!':
            cents += 25.0
        elif ch == '.':
            cents -= 25.0
        idx += 1

    try:
        octave = int(s_up[idx:])
    except (ValueError, IndexError):
        raise ValueError(f"Formato ottava non valido in: {note_name}")

    midi_value = (octave + 1) * consts.SEMITONES_PER_OCTAVE + note_map[note] + semitones
    if not (consts.MIDI_MIN <= midi_value <= consts.MIDI_MAX):
        raise ValueError(f"Nota MIDI fuori range: {midi_value}")

    return midi_value, cents


def parse_fundamental_frequency(value: Union[str, float], basenote_hz: float, diapason_hz: float, default: float) -> float:
    """Parse fundamental frequency from note name or return default."""
    if isinstance(value, str) and value.lower() == "basenote":
        return basenote_hz
    elif isinstance(value, float):
        return value
    else:
        try:
            midi_f, cents_f = parse_note_with_microtones(str(value))
            return apply_cents(convert_midi_to_hz(midi_f, diapason_hz), cents_f)
        except ValueError:
            return default


def process_cluster_centers_to_cents(cluster_centers: List[Tuple[float, int]]) -> List[Tuple[float, int]]:
    """Convert cluster centers to cents [0,1200) format.
    Auto-detects if input is already in cents or in ratio format.
    """
    centers = []
    for item in (cluster_centers or []):
        try:
            r, c = item
            if r and r > 0:
                # Auto-detect format: if value is likely already in cents, use as-is
                # Otherwise convert from ratio to cents
                if r >= 1200.0:
                    # Definitely in cents (over one octave), reduce to [0,1200)
                    cents = float(r % 1200.0)
                elif 1.0 < r < 10.0:
                    # Likely a ratio (between 1.0 and 10.0), convert to cents
                    cents = ratio_to_cents(r)
                elif r <= 1200.0:
                    # Likely already in cents (0-1200 range), use as-is
                    cents = float(r % 1200.0)
                else:
                    # Fallback: treat as ratio
                    cents = ratio_to_cents(r)

                centers.append((cents, int(c)))
        except (ValueError, TypeError):
            continue
    return centers


def score_scale_against_centers(scale_degrees: List[float], centers_c: List[Tuple[float, int]]) -> Tuple[float, List[Tuple[int, float, int]]]:
    """Score a scale against cluster centers and return (avg_error, mapping)."""
    if not scale_degrees or not centers_c:
        return float('inf'), []

    total_werr = 0.0
    total_w = 0
    mapping: List[Tuple[int, float, int]] = []

    for (ci, cnt) in centers_c:
        try:
            idx = min(range(len(scale_degrees)), key=lambda k: circular_difference(ci, scale_degrees[k], 1200.0))
            err = circular_difference(ci, scale_degrees[idx], 1200.0)
            total_werr += float(err) * max(1, int(cnt))
            total_w += max(1, int(cnt))
            ratio = 2.0 ** ((float(scale_degrees[idx]) % 1200.0) / 1200.0)
            mapping.append((int(idx), float(ratio), int(cnt)))
        except (ValueError, IndexError):
            continue

    if total_w == 0:
        return float('inf'), []

    avg = total_werr / float(total_w)
    return avg, mapping


def extract_cents_from_analysis(analysis_data: Dict, cents_delta: Optional[float] = None) -> List[float]:
    """Extract cents list from analysis data (ratio clusters or ratios_reduced)."""

    cents_from_analysis = []

    # Try ratio clusters first
    centers = analysis_data.get('ratio_clusters') if isinstance(analysis_data, dict) else None
    if centers:
        tmp = []
        for (ctr, _cnt) in centers:
            try:
                c = (1200.0 * math.log(float(ctr), 2.0)) % 1200.0
                tmp.append(float(c))
            except (ValueError, TypeError, ZeroDivisionError):
                pass
        cents_from_analysis = sorted(tmp)

    # Fallback to ratios_reduced
    if not cents_from_analysis:
        rr = analysis_data.get('ratios_reduced') if isinstance(analysis_data, dict) else None
        if rr:
            tmp = []
            for r in rr:
                try:
                    c = (1200.0 * math.log(float(r), 2.0)) % 1200.0
                    tmp.append(float(c))
                except (ValueError, TypeError, ZeroDivisionError):
                    pass
            cents_from_analysis = sorted(tmp)

    # Apply cents-delta if specified
    if cents_delta is not None:
        try:
            dd = float(cents_delta)
            cents_from_analysis = sorted(((x + dd) % 1200.0) for x in (cents_from_analysis or []))
        except (ValueError, TypeError):
            pass

    return cents_from_analysis


def conditional_print(condition: bool, *args, **kwargs) -> None:
    """Print only if condition is True."""
    if condition:
        print(*args, **kwargs)


def print_formatted_lines_if(condition: bool, lines: List[str]) -> None:
    """Print formatted lines only if condition is True."""
    if condition:
        for line in lines:
            print(line)


def process_scale_items_to_rows(items: List[Dict], start_index: int = 1) -> List[List[str]]:
    """Process scale items into formatted rows for table display."""
    rows = []
    for i, item in enumerate(items, start=start_index):
        nm = safe_string_conversion(item.get('name', ''))
        fl = safe_string_conversion(item.get('file', ''))
        er = f"{safe_float_conversion(item.get('avg_cents_error', 0.0)):.2f}"
        rows.append([str(i), nm, fl, er])
    return rows




def format_scala_match_line(pos_num: int, item: Dict, name_width: int = 35) -> str:
    """Format a single scala match line for display."""
    nm = str(item.get('name',''))[:name_width]
    er = f"{safe_float_conversion(item.get('avg_cents_error', 0.0)):.2f}"
    fl = safe_string_conversion(item.get('file',''))
    return f"{str(pos_num).rjust(3)}  {nm.ljust(name_width)}  {er.rjust(7)}    {fl}"




def deduplicate_near_equal_cents(steps: List[float], tolerance: float = 1e-6) -> List[float]:
    """Remove duplicate steps that are within tolerance of each other."""
    steps_u = []
    for s in sorted(steps):
        if not steps_u or abs(s - steps_u[-1]) > tolerance:
            steps_u.append(s)
    return steps_u






def avg_error_rotation(a_list: List[float], b_list: List[float]) -> float:
    """Calculate average error between two lists of cents values with rotation optimization.

    Args:
        a_list: First list of cents values
        b_list: Second list of cents values

    Returns:
        Average error in cents (float('inf') if lists are empty)
    """
    if not a_list or not b_list:
        return float('inf')

    a = sorted([float(x) % 1200.0 for x in a_list])
    b = sorted([float(x) % 1200.0 for x in b_list])
    n = min(len(a), len(b))
    a = a[:n]
    b = b[:n]

    best = float('inf')
    for sh in range(n):
        s = 0.0
        for i in range(n):
            dv = abs(a[i] - b[(i + sh) % n])
            if dv > 600.0:
                dv = 1200.0 - dv
            s += dv
        best = min(best, s / n)
    return best






class SpinnerAnimation:
    """A thread-safe spinner animation for long-running operations."""

    def __init__(self, message: str = "Processing", cols: int = 80):
        self.message = message
        self.cols = cols
        self._stop_event = None
        self._thread = None

    def start(self):
        """Start the spinner animation."""
        import threading
        import time
        import sys

        if self._thread and self._thread.is_alive():
            return

        self._stop_event = threading.Event()

        def _spinner_worker():
            frames = ['|', '/', '-', '\\']
            frame_idx = 0
            last = time.monotonic()

            while not self._stop_event.is_set():
                now = time.monotonic()
                if now - last >= 0.1:
                    last = now
                    ch = frames[frame_idx % len(frames)]
                    frame_idx += 1
                    try:
                        if hasattr(sys, 'stdout') and sys.stdout and sys.stdout.writable():
                            spinner_msg = f"{self.message} {ch}"
                            if len(spinner_msg) > self.cols - 1:
                                spinner_msg = spinner_msg[:max(0, self.cols - 1)]
                            padding = max(0, self.cols - len(spinner_msg) - 1)
                            print(f"\r{spinner_msg}{' ' * padding}", end="", flush=True)
                    except (BrokenPipeError, OSError, ValueError, AttributeError):
                        print("In progress...\n", flush=True)
                        break
                try:
                    time.sleep(0.02)
                except (InterruptedError, OSError):
                    break

        # Print initial line
        try:
            if hasattr(sys, 'stdout') and sys.stdout and sys.stdout.writable():
                print(self.message + " ", end="", flush=True)
        except (BrokenPipeError, OSError, ValueError, AttributeError):
            pass

        self._thread = threading.Thread(target=_spinner_worker, daemon=True)
        self._thread.start()

    def stop(self):
        """Stop the spinner animation and clean up."""
        if self._stop_event:
            self._stop_event.set()

        if self._thread and self._thread.is_alive():
            try:
                self._thread.join(timeout=1.0)
            except RuntimeError:
                pass

        # Clear spinner line
        try:
            print(f"\r{' ' * self.cols}\r", end="")
        except (BrokenPipeError, OSError):
            pass
        print("")  # newline after spinner

    def __enter__(self):
        self.start()
        return self

    def __exit__(self, exc_type, exc_val, exc_tb):
        self.stop()


def format_aligned_table(headers: List[str], rows: List[List[str]]) -> List[str]:
    """Format a table with aligned columns.

    Args:
        headers: List of column headers
        rows: List of rows, where each row is a list of strings

    Returns:
        List of formatted lines ready for printing
    """
    if not headers or not rows:
        return []

    # Calculate column widths
    widths = [len(h) for h in headers]
    for row in rows:
        for col_idx, cell in enumerate(row):
            if col_idx < len(widths):
                if len(cell) > widths[col_idx]:
                    widths[col_idx] = len(cell)

    # Format function
    def format_row(row_data):
        return "  ".join(str(row_data[i]).ljust(widths[i]) for i in range(min(len(row_data), len(widths))))

    # Build output
    lines = [format_row(headers)]
    for row in rows:
        lines.append(format_row(row))

    return lines


def safe_float_conversion(value, default: float = 0.0) -> float:
    """Safely convert a value to float with a default fallback."""
    try:
        return float(value)
    except (ValueError, TypeError, AttributeError):
        return default


def safe_string_conversion(value, default: str = "") -> str:
    """Safely convert a value to string with a default fallback."""
    try:
        return str(value).strip()
    except (ValueError, TypeError, AttributeError):
        return default


def safe_numeric_value(value, default: Union[str, float] = "") -> Union[str, float]:
    """Safely convert value to numeric for Excel, with NaN and None handling.

    Returns float if conversion succeeds, default otherwise.
    Handles NaN values properly for Excel compatibility.
    """
    if value is None:
        return default
    if isinstance(value, (int, float)) and not math.isnan(value):
        return float(value)
    return default


def safe_delta_calculation(val1, val2) -> Union[str, float]:
    """Safely calculate absolute difference between two values.

    Returns float difference if both values are valid numbers, empty string otherwise.
    """
    if (val1 and val2 and
        isinstance(val1, (int, float)) and not math.isnan(val1) and
        isinstance(val2, (int, float)) and not math.isnan(val2)):
        return abs(val1 - val2)
    return ""


def extract_peaks_from_spectrum(mag_spectrum, freqs_axis_vals, max_peaks=10, min_freq=150.0, max_freq_offset=100.0, sr=None):
    """Extract peaks from magnitude spectrum for formant analysis.
    Args:
        mag_spectrum: Magnitude spectrum array
        freqs_axis_vals: Frequency values corresponding to spectrum bins
        max_peaks: Maximum number of peaks to extract
        min_freq: Minimum frequency to consider (Hz)
        max_freq_offset: Offset from Nyquist frequency (Hz)
        sr: Sample rate (used to determine max frequency)
    Returns:
        Tuple of (formant_freqs, formant_amps) lists
    """
    import numpy as _np

    # Validate inputs
    if mag_spectrum is None or not hasattr(mag_spectrum, 'shape'):
        return [], []
    if freqs_axis_vals is None or not hasattr(freqs_axis_vals, '__len__'):
        return [], []

    try:
        from scipy.signal import find_peaks as _find_peaks
    except (ImportError, ModuleNotFoundError):
        _find_peaks = None

    try:
        if _find_peaks is not None:
            # Use scipy peak finding with distance constraint
            sr_eff = sr if sr else 44100.0  # fallback sample rate
            distance = max(1, int((200.0 / (sr_eff / 2.0)) * mag_spectrum.shape[0]))
            peak_indices, _ = _find_peaks(mag_spectrum, distance=distance)
            if peak_indices.size > max_peaks:
                # Take the highest amplitude peaks
                sort_idx = _np.argsort(mag_spectrum[peak_indices])[-max_peaks:]
                peak_indices = peak_indices[sort_idx]
        else:
            # Fallback: use simple argpartition
            peak_indices = _np.argpartition(mag_spectrum, -max_peaks)[-max_peaks:]
    except (AttributeError, IndexError, ValueError, TypeError):
        # If any array operation fails, return empty lists
        return [], []

    try:
        # Get amplitudes and normalize
        peak_amps = mag_spectrum[peak_indices]
        if peak_amps.size > 0:
            peak_amps = peak_amps / float(peak_amps.max())

        # Filter frequencies
        freq_selection = freqs_axis_vals[peak_indices]
        max_freq = (sr / 2.0 - max_freq_offset) if sr else float('inf')
        freq_filter = (freq_selection >= min_freq) & (freq_selection <= max_freq)

        formant_freqs = list(map(float, freq_selection[freq_filter]))
        formant_amps = list(map(float, peak_amps[freq_filter]))

        return formant_freqs, formant_amps
    except (AttributeError, IndexError, ValueError, TypeError):
        # If any array operation fails, return empty lists
        return [], []

# --- File export utilities ---

def log_export_success(file_path: str) -> None:
    """Log successful file export."""
    print(f"Exported: {file_path}")

def log_export_error(file_path: str, error: Exception) -> None:
    """Log file export error."""
    print(f"Write error {file_path}: {error}")

def safe_file_write(file_path: str, content: str, encoding: str = "utf-8") -> bool:
    """Safely write content to file with error handling."""
    try:
        with open(file_path, "w", encoding=encoding) as f:
            f.write(content)
        log_export_success(file_path)
        return True
    except IOError as e:
        log_export_error(file_path, e)
        return False

def lazy_import_openpyxl():
    """Lazy import openpyxl with error handling."""
    try:
        import importlib
        openpyxl = importlib.import_module('openpyxl')
        styles = importlib.import_module('openpyxl.styles')
        return openpyxl, styles
    except (ImportError, AttributeError):
        return None, None

def lazy_import_audio_libs():
    """Lazy import audio processing libraries."""
    libs = {}
    try:
        import librosa as _lb
        libs['librosa'] = _lb
    except ImportError:
        _lb = None
        libs['librosa'] = None

    try:
        import numpy as _np
        libs['numpy'] = _np
    except ImportError:
        _np = None
        libs['numpy'] = None

    try:
        import scipy.signal as _sp
        libs['scipy_signal'] = _sp
    except ImportError:
        _sp = None
        libs['scipy_signal'] = None

    return libs

# --- Table generation utilities ---

def generate_diapason_text_fallback(output_base: str, diapason_hz: float, basenote_hz: float, analysis_result: Optional[Dict], digits: Optional[int] = None) -> None:
    """Generate rich Diapason text report (legacy format) for --diapason-analysis.
    This mirrors the classic <output>_diapason.txt shown in the issue description.
    """
    try:
        txt_diap_path = f"{output_base}_diapason.txt"

        # Helper: 12-TET mapping and cents delta
        def _freq_to_12tet_name_midi_delta(freq: float, a4: float) -> Tuple[str, Optional[int], Optional[float]]:
            try:
                if not (isinstance(freq, (int, float)) and isinstance(a4, (int, float))):
                    return "", None, None
                f = float(freq)
                a = float(a4)
                if not (f > 0 and a > 0):
                    return "", None, None
                midi = int(round(consts.MIDI_A4 + consts.SEMITONES_PER_OCTAVE * math.log2(f / a)))
                midi = max(consts.MIDI_MIN, min(consts.MIDI_MAX, midi))
                # exact 12-TET frequency for that MIDI
                f_et = a * (2.0 ** ((midi - consts.MIDI_A4) / consts.SEMITONES_PER_OCTAVE))
                delta_c = 1200.0 * math.log2(f / f_et) if f_et > 0 else 0.0
                try:
                    name = midi_to_note_name_12tet(midi)
                except (ValueError, TypeError):
                    name = ""
                return name, midi, delta_c
            except (ValueError, TypeError, ArithmeticError):
                return "", None, None

        # Helper function for Hz formatting
        def _format_hz(value, digits_param):
            if digits_param is not None and isinstance(value, (int, float)):
                return f"{round(float(value), digits_param):.{digits_param}f}"
            else:
                return f"{value:.6f}"

        # Extract values from analysis_result
        # Determine user A4: if user explicitly set --diapason use it; otherwise use 440 Hz default
        user_set = bool(analysis_result.get('user_diapason_set', False)) if isinstance(analysis_result, dict) else False


        if user_set:
            a4_user = float(diapason_hz)  # Use actual diapason from user
            bn_user = float(basenote_hz)  # Use actual basenote from user's diapason
        else:
            # For --diapason-analysis, user A4 should be 440 Hz (default) unless explicitly specified
            a4_user = 440.0  # Use standard A4=440 Hz as reference
            # Calculate basenote that corresponds to A4=440 Hz
            try:
                bn_user = float(basenote_hz) * (440.0 / float(diapason_hz))
            except (ValueError, TypeError, ZeroDivisionError):
                bn_user = float(basenote_hz)

        try:
            a4_est = float(analysis_result.get('diapason_est')) if isinstance(analysis_result, dict) and analysis_result.get('diapason_est') is not None else None
        except (ValueError, TypeError, AttributeError):
            a4_est = None

        # Extract A4 statistical information
        try:
            a4_std = float(analysis_result.get('diapason_confidence_std')) if isinstance(analysis_result, dict) and analysis_result.get('diapason_confidence_std') is not None else None
        except (ValueError, TypeError, AttributeError):
            a4_std = None

        try:
            a4_ci = analysis_result.get('diapason_confidence_interval') if isinstance(analysis_result, dict) else None
        except (ValueError, TypeError, AttributeError):
            a4_ci = None

        try:
            a4_samples = int(analysis_result.get('diapason_sample_count')) if isinstance(analysis_result, dict) and analysis_result.get('diapason_sample_count') is not None else None
        except (ValueError, TypeError, AttributeError):
            a4_samples = None

        try:
            a4_delta_hz = float(analysis_result.get('diapason_delta_440_hz')) if isinstance(analysis_result, dict) and analysis_result.get('diapason_delta_440_hz') is not None else None
        except (ValueError, TypeError, AttributeError):
            a4_delta_hz = None

        try:
            a4_delta_percent = float(analysis_result.get('diapason_delta_440_percent')) if isinstance(analysis_result, dict) and analysis_result.get('diapason_delta_440_percent') is not None else None
        except (ValueError, TypeError, AttributeError):
            a4_delta_percent = None

        try:
            bn_est_hz = float(analysis_result.get('basenote_est_hz')) if isinstance(analysis_result, dict) and analysis_result.get('basenote_est_hz') is not None else None
        except (ValueError, TypeError, AttributeError):
            bn_est_hz = None

        if (not isinstance(bn_est_hz, (int, float))) or not (bn_est_hz and bn_est_hz > 0):
            # derive from proportion between A4 est and user A4
            try:
                if isinstance(a4_est, (int, float)) and a4_est and a4_user:
                    bn_est_hz = bn_user * (a4_est / a4_user)
                else:
                    bn_est_hz = None
            except (ValueError, TypeError, ZeroDivisionError):
                bn_est_hz = None

        # Basenote 12-TET (estimated)
        bn12_name = ""
        bn12_midi = None
        if isinstance(bn_est_hz, (int, float)) and bn_est_hz and isinstance(a4_est, (int, float)) and a4_est:
            nm, md, _dc = _freq_to_12tet_name_midi_delta(bn_est_hz, a4_est)
            bn12_name, bn12_midi = nm, md

        lines: List[str] = [
            "DIAPASON - Reference systems",
            "",
            "=== ANALYSIS LEGEND ===",
            "• A4_user/A4_estimated: User reference A4 vs. estimated A4 from audio analysis",
            "• Basenote_Hz: Base frequency for scale construction (user vs. estimated)",
            "• Clustering: Algorithm parameters used to group detected frequencies into scale degrees",
            "• Scala_match: Best matching scale from Scala database (if available)",
            "• Tuning_inferred/comparative: Detected tuning systems based on audio analysis",
            "• Scale steps: Detected intervals with occurrence counts from audio",
            "• Fragment analysis: Coverage of theoretical scale degrees in detected audio",
            "",
            f"A4_user (Hz): {_format_hz(a4_user, digits)}",
            f"A4_estimated (Hz): {_format_hz(a4_est, digits)}" if isinstance(a4_est, (int, float)) else "A4_estimated (Hz): ",
        ]

        # Add A4 statistical information if available
        if isinstance(a4_est, (int, float)):
            if a4_delta_hz is not None and a4_delta_percent is not None:
                sign = "+" if a4_delta_hz >= 0 else ""
                lines.append(f"A4_delta_from_440Hz: {sign}{a4_delta_hz:.3f} Hz ({sign}{a4_delta_percent:.2f}%)")

            if a4_samples is not None and a4_samples > 0:
                lines.append(f"A4_sample_count: {a4_samples}")

            if a4_std is not None:
                lines.append(f"A4_confidence_std: ±{a4_std:.3f} Hz")

            if a4_ci is not None and len(a4_ci) == 2:
                lines.append(f"A4_confidence_95CI: [{a4_ci[0]:.3f}, {a4_ci[1]:.3f}] Hz")

        lines.extend([
            f"Basenote_Hz (user): {_format_hz(bn_user, digits)}",
            f"Basenote_Hz (estimated): {_format_hz(bn_est_hz, digits)}" if isinstance(bn_est_hz, (int, float)) else "Basenote_Hz (estimated): "
        ])
        if bn12_name and bn12_midi is not None:
            lines.append(f"Basenote_12TET (estimated): {bn12_name} (MIDI {bn12_midi})")
        else:
            lines.append("Basenote_12TET (estimated): ")
        lines.append("")

        # Clustering tolerance information
        try:
            tolerance_percent = float(analysis_result.get('clustering_tolerance_percent', 2.0)) if isinstance(analysis_result, dict) else 2.0
            tolerance_cents = float(analysis_result.get('clustering_tolerance_cents_approx', tolerance_percent * 17.3)) if isinstance(analysis_result, dict) else tolerance_percent * 17.3
            clustering_method = safe_string_conversion(analysis_result.get('clustering_method', 'simple')) if isinstance(analysis_result, dict) else 'simple'
            scale_size_hint = analysis_result.get('scale_size_hint') if isinstance(analysis_result, dict) else None

            lines.append("Clustering Parameters:")
            lines.append(f"  Tolerance: {tolerance_percent:.1f}% (≈ {tolerance_cents:.1f} cents)")
            lines.append(f"  Method: {clustering_method}")
            if scale_size_hint:
                lines.append(f"  Scale size hint: {scale_size_hint} degrees")
            lines.append("")
        except (ValueError, TypeError, AttributeError):
            pass

        # Scala best match and top-5
        sc_info = analysis_result.get('scala_match_info') if isinstance(analysis_result, dict) else None
        sc_top = analysis_result.get('scala_top_matches') if isinstance(analysis_result, dict) else None
        sc_steps = analysis_result.get('scala_match_steps') if isinstance(analysis_result, dict) else None

        if sc_info and isinstance(sc_info, dict):
            nm = safe_string_conversion(sc_info.get('name', ''))
            fl = safe_string_conversion(sc_info.get('file', ''))
            try:
                er = float(sc_info.get('avg_cents_error', 0.0))
            except (ValueError, TypeError):
                er = 0.0
            lines.append("Scala_match:")
            lines.append(f"  name: {nm}")
            lines.append(f"  file: {fl}")
            lines.append(f"  avg_error_cents: {er:.2f}")
            lines.append("")


        # Helper function to format inferred system fragment
        def _format_inferred_fragment(fragment: dict, title_prefix: str, extra_info: str = "") -> None:
            system_name = safe_string_conversion(fragment.get('system_name', 'Unknown'))
            total_degrees = fragment.get('total_degrees', 0)
            detected_count = fragment.get('detected_count', 0)
            coverage_percent = fragment.get('coverage_percent', 0.0)

            lines.append(f"{title_prefix} ({system_name}):")
            if extra_info:
                lines.append(extra_info)
            lines.append(f"Coverage: {detected_count}/{total_degrees} degrees ({coverage_percent:.1f}%)")
            lines.append("")

            # Detected degrees
            detected = fragment.get('detected_degrees', [])
            if detected:
                lines.append("Detected degrees:")
                lines.append("Deg  Theoretical  Detected    Error   Count")
                lines.append("---  -----------  --------   ------  -----")
                for deg in detected[:25]:  # Limit to 25 entries for text output
                    deg_num = deg.get('degree', 0)
                    theo_cents = deg.get('theoretical_cents', 0.0)
                    det_cents = deg.get('detected_cents', 0.0)
                    cents_error = deg.get('cents_error', 0.0)
                    count = deg.get('count', 0)
                    lines.append(f"{deg_num:3d}  {theo_cents:9.1f}c  {det_cents:7.1f}c  {cents_error:+6.1f}c  {count:5d}")
                if len(detected) > 25:
                    lines.append(f"... and {len(detected) - 25} more detected degrees")
                lines.append("")

            # Missing degrees (limited output)
            missing = fragment.get('missing_degrees', [])
            if missing:
                missing_list = [str(deg.get('degree', 0)) for deg in missing[:20]]
                lines.append(f"Missing degrees: {', '.join(missing_list)}")
                if len(missing) > 20:
                    lines.append(f"... and {len(missing) - 20} more missing degrees")
                lines.append("")

        # Helper function to format scale map entries
        def _format_scale_entry(scale_idx: int, scale_ratio: float, scale_cnt: int) -> str:
            try:
                rr_val = float(scale_ratio)
                hz_user_val = bn_user * rr_val
                hz_est_val = (bn_est_hz * rr_val) if isinstance(bn_est_hz, (int, float)) else None
                cents_val = ratio_to_cents(rr_val) % 1200.0
                name_u_val, _m_u_val, dc_u_val = _freq_to_12tet_name_midi_delta(hz_user_val, a4_user)
                if isinstance(hz_est_val, (int, float)) and isinstance(a4_est, (int, float)):
                    name_e_val, _m_e_val, dc_e_val = _freq_to_12tet_name_midi_delta(hz_est_val, a4_est)
                else:
                    name_e_val, dc_e_val = "", None
                # Format Hz values using digits parameter
                hz_user_str = _format_hz(hz_user_val, digits).rjust(10)
                hz_est_str = _format_hz(hz_est_val if isinstance(hz_est_val,(int,float)) else 0.0, digits).rjust(10)

                return (
                    f"{str(scale_idx).ljust(4)}{rr_val:0.10f}  "
                    f"{hz_user_str}  "
                    f"{hz_est_str}  "
                    f"{cents_val:7.2f}  "
                    f"{int(scale_cnt):5d}  "
                    f"{name_u_val.ljust(11)}  "
                    f"{(dc_u_val if isinstance(dc_u_val,(int,float)) else 0.0):14.2f}  "
                    f"{name_e_val.ljust(10)}  "
                    f"{(dc_e_val if isinstance(dc_e_val,(int,float)) else 0.0):14.2f}"
                )
            except (ValueError, TypeError, AttributeError):
                return ""

        if sc_steps:
            # Filter out theoretical steps with count = 0 from display
            detected_sc_steps = [(idx, ratio, cnt) for (idx, ratio, cnt) in sc_steps if cnt > 0]

            if detected_sc_steps:
                # Normalize ratios relative to first detected step (traditional scale approach)
                first_ratio = detected_sc_steps[0][1]  # ratio of first detected step
                normalized_sc_steps = []
                for (idx, ratio, cnt) in detected_sc_steps:
                    normalized_ratio = ratio / first_ratio  # Normalize to first step
                    normalized_sc_steps.append((idx, normalized_ratio, cnt))

                lines.append("Scala_map (index, Ratio, Hz_from_Base(user), Hz_from_Base(estimated), Cents, Count, 12TET_user, DeltaCents_user, 12TET_est, DeltaCents_est):")
                lines.append("Idx  Ratio         Hz_user     Hz_est      Cents    Count  12TET_user  DeltaCents_user  12TET_est  DeltaCents_est")
                for (idx, normalized_ratio, cnt) in normalized_sc_steps:
                    entry = _format_scale_entry(idx, normalized_ratio, cnt)
                    if entry:
                        lines.append(entry)
                lines.append("")

        # Inferred tuning (primary)
        tinf = analysis_result.get('tuning_inferred') if isinstance(analysis_result, dict) else None
        steps_inf = analysis_result.get('scale_steps') if isinstance(analysis_result, dict) else None
        if tinf and isinstance(tinf, dict):
            nm = safe_string_conversion(tinf.get('name', ''))
            try:
                er = float(tinf.get('avg_cents_error', 0.0))
            except (ValueError, TypeError):
                er = 0.0
            lines.append("Tuning_inferred:")
            lines.append(f"  name: {nm}")
            lines.append(f"  avg_error_cents: {er:.2f}")
            lines.append("")
        if steps_inf and isinstance(steps_inf, list) and steps_inf:
            # Filter out theoretical steps with count = 0 from display
            detected_steps_inf = [(idx, ratio, cnt) for (idx, ratio, cnt) in steps_inf if cnt > 0]

            if detected_steps_inf:
                # Normalize ratios relative to first detected step (traditional scale approach)
                first_ratio_inf = detected_steps_inf[0][1]  # ratio of first detected step
                normalized_steps_inf = []
                for (idx, ratio, cnt) in detected_steps_inf:
                    normalized_ratio = ratio / first_ratio_inf  # Normalize to first step
                    normalized_steps_inf.append((idx, normalized_ratio, cnt))

                lines.append("Inferred steps (index, Ratio, Hz_from_Base(user), Hz_from_Base(estimated), Cents, Count, 12TET_user, DeltaCents_user, 12TET_est, DeltaCents_est):")
                lines.append("Idx  Ratio         Hz_user     Hz_est      Cents    Count  12TET_user  DeltaCents_user  12TET_est  DeltaCents_est")
                for (idx, normalized_ratio, cnt) in normalized_steps_inf:
                    entry = _format_scale_entry(idx, normalized_ratio, cnt)
                    if entry:
                        lines.append(entry)
                lines.append("")

        # Comparative tuning
        tcmp = analysis_result.get('tuning_comparative') if isinstance(analysis_result, dict) else None
        steps_cmp = analysis_result.get('scale_steps_comp') if isinstance(analysis_result, dict) else None
        if tcmp and isinstance(tcmp, dict):
            nm = safe_string_conversion(tcmp.get('name', ''))
            try:
                er = float(tcmp.get('avg_cents_error', 0.0))
            except (ValueError, TypeError):
                er = 0.0
            lines.append("Tuning_comparative:")
            lines.append(f"  name: {nm}")
            lines.append(f"  avg_error_cents: {er:.2f}")
            lines.append("")
        if steps_cmp and isinstance(steps_cmp, list) and steps_cmp:
            # Filter out theoretical steps with count = 0 from display
            detected_steps_cmp = [(idx, ratio, cnt) for (idx, ratio, cnt) in steps_cmp if cnt > 0]

            if detected_steps_cmp:
                # Normalize ratios relative to first detected step (traditional scale approach)
                first_ratio = detected_steps_cmp[0][1]  # ratio of first detected step
                normalized_steps_cmp = []
                for (idx, ratio, cnt) in detected_steps_cmp:
                    normalized_ratio = ratio / first_ratio  # Normalize to first step
                    normalized_steps_cmp.append((idx, normalized_ratio, cnt))

                lines.append("Comparative steps (index, Ratio, Hz_from_Base(user), Hz_from_Base(estimated), Cents, Count, 12TET_user, DeltaCents_user, 12TET_est, DeltaCents_est):")
                lines.append("Idx  Ratio         Hz_user     Hz_est      Cents    Count  12TET_user  DeltaCents_user  12TET_est  DeltaCents_est")
                for (idx, normalized_ratio, cnt) in normalized_steps_cmp:
                    entry = _format_scale_entry(idx, normalized_ratio, cnt)
                    if entry:
                        lines.append(entry)
                lines.append("")

        # Base note selection info
        try:
            base_selected_info = analysis_result.get('base_note_selection_info') if isinstance(analysis_result, dict) else None
            if base_selected_info:
                lines.append(safe_string_conversion(base_selected_info))
                lines.append("")
        except (ValueError, TypeError, AttributeError):
            pass

        # Refined matching info
        try:
            refined_info = analysis_result.get('refined_matching_info') if isinstance(analysis_result, dict) else None
            if refined_info:
                lines.append(safe_string_conversion(refined_info))
                lines.append("")
        except (ValueError, TypeError, AttributeError):
            pass

        # NEW: Fragment analysis section
        fragment_analysis = analysis_result.get('fragment_analysis') if isinstance(analysis_result, dict) else None
        if fragment_analysis and isinstance(fragment_analysis, dict):
            lines.append("=== SCALE FRAGMENT ANALYSIS ===")
            lines.append("")

            # Inferred system fragment
            inferred_frag = fragment_analysis.get('inferred_system')
            if inferred_frag and isinstance(inferred_frag, dict):
                _format_inferred_fragment(inferred_frag, "INFERRED SYSTEM FRAGMENT")

            # Scala fragment
            scala_frag = fragment_analysis.get('scala_match')
            if scala_frag and isinstance(scala_frag, dict):
                scala_file = safe_string_conversion(scala_frag.get('scala_file', 'Unknown'))
                # Modify fragment dict to use system_name field for compatibility
                modified_frag = dict(scala_frag)
                if 'scala_name' in modified_frag:
                    modified_frag['system_name'] = modified_frag['scala_name']
                _format_inferred_fragment(modified_frag, "SCALA MATCH FRAGMENT", f"File: {scala_file}")

        # Save file
        safe_file_write(txt_diap_path, "\n".join(lines) + "\n", encoding="utf-8")
        print(f"Exported: {txt_diap_path}")

    except (OSError, IOError, PermissionError, UnicodeError) as e:
        print(f"Diapason text fallback error: {e}")

def generate_excel_table_headers(has_analysis: bool = False) -> List[str]:
    """Generate standard Excel table headers."""
    headers = ["Step", "Ratio", "Hz", "Cents"]
    if has_analysis:
        headers.extend(["DeltaCents_F0", "DeltaCents_Formant"])
    return headers

def get_standard_system_headers() -> List[str]:
    """Get standard system table headers."""
    return ["Step", "MIDI", "Ratio", "Hz"]

def get_reference_system_headers() -> List[str]:
    """Get reference system headers for diapason analysis."""
    return ["Step", "Ratio", "Hz", "Cents"]

def get_comparison_table_headers() -> List[str]:
    """Get headers for comparison tables."""
    return [
        "Step", "MIDI", "Ratio", "Custom_Hz", "Harmonic_Hz", "DeltaHz_Harm",
        "Subharm_Hz", "DeltaHz_Sub", "TET_Hz", "TET_Note", "DeltaHz_TET",
        "AudioF0_Hz", "AudioFormant_Hz", "Formant_RelAmp", "DeltaHz_F0", "DeltaHz_Formant"
    ]

def get_comparison_table_headers_excel() -> List[str]:
    """Get Excel headers for comparison tables with absolute deltas."""
    return [
        "Step", "MIDI", "Ratio", "Custom_Hz", "Harmonic_Hz", "|DeltaHz_Harm|",
        "Subharm_Hz", "|DeltaHz_Sub|", "TET_Hz", "TET_Note", "|DeltaHz_TET|",
        "AudioF0_Hz", "AudioFormant_Hz", "Formant_RelAmp", "|DeltaHz_F0|", "|DeltaHz_Formant|"
    ]

def build_reference_system_rows(base_hz: float, tet_divisions: int = 12) -> Dict[str, List]:
    """Build reference system rows for TET, Pythagorean 12, and Pythagorean 7."""
    return {
        "tet": [[k, 2.0 ** (k / tet_divisions), base_hz * (2.0 ** (k / tet_divisions)), (1200.0 * k / tet_divisions)] for k in range(tet_divisions)],
        "py12": [[i, r, base_hz * r, ratio_to_cents(r)]
                 for i, r in enumerate(generate_pythagorean_12())],
        "py7": [[i, r, base_hz * r, ratio_to_cents(r)]
                for i, r in enumerate(sorted(generate_pythagorean_7()))]
    }

def get_excel_color_fills():
    """Get standard Excel color fills for different data series."""
    try:
        from openpyxl.styles import PatternFill
        return {
            'custom': PatternFill(start_color="FFFFCCCC", end_color="FFFFCCCC", fill_type="solid"),
            'harmonic': PatternFill(start_color="FFCCFFCC", end_color="FFCCFFCC", fill_type="solid"),
            'subharmonic': PatternFill(start_color="FFFFFFCC", end_color="FFFFFFCC", fill_type="solid"),
            'tet': PatternFill(start_color="FFCCE5FF", end_color="FFCCE5FF", fill_type="solid"),
            'header': PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid"),
            'diapason_header': PatternFill(start_color="FFEEEEEE", end_color="FFEEEEEE", fill_type="solid")
        }
    except ImportError:
        return {}

def get_excel_borders():
    """Get standard Excel border styles."""
    try:
        from openpyxl.styles import Border, Side
        thin_side = Side(style="thin", color="FFB0B0B0")
        return {
            'thin': Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
        }
    except ImportError:
        return {}

def setup_excel_worksheet_formatting(ws, headers: List[str]) -> None:
    """Setup Excel worksheet with standard formatting."""
    try:
        from openpyxl.styles import Font

        # Add headers
        ws.append(headers)

        # Header formatting
        fills = get_excel_color_fills()
        header_font = Font(bold=True)
        header_fill = fills.get('header')

        for cell in ws[1]:
            cell.font = header_font
            if header_fill:
                cell.fill = header_fill

    except (ImportError, AttributeError):
        # Fallback: just add headers without formatting
        ws.append(headers)


def safe_median_from_floats(values):
    """
    Calculate median from a list/array of numeric values with numpy.
    Returns float median or 0.0 if calculation fails.
    """
    try:
        _np = ensure_numpy_import()
        if _np is None:
            return 0.0
        if not values:
            return 0.0
        arr = _np.array([float(x) for x in values], dtype=float)
        return float(_np.median(arr))
    except (ValueError, TypeError, ImportError, AttributeError):
        return 0.0


def empty_float_array():
    """
    Returns an empty numpy array with dtype=float.
    Falls back to empty list if numpy not available.
    """
    try:
        _np = ensure_numpy_import()
        if _np is None:
            return []
        return _np.array([], dtype=float)
    except (ImportError, AttributeError):
        return []


def extract_audio_params(args):
    """Extract audio analysis parameters from args into a dictionary.

    This centralizes parameter extraction to reduce code duplication.
    """
    return {
        'diapason_analysis': getattr(args, 'diapason_analysis', False),
        'f0_only': getattr(args, 'f0_only', False),
        'cents_delta': getattr(args, 'cents_delta', None),
        'render_enabled': getattr(args, 'render', False),
        'cqt_bins': getattr(args, 'cqt_bins', None),
        'cqt_bpo': getattr(args, 'cqt_bpo', None),
        'cqt_fmin': getattr(args, 'cqt_fmin', None),
        'lpc_order': getattr(args, 'lpc_order', None),
        'lpc_preemph': getattr(args, 'lpc_preemph', 0.97),
        'lpc_window_ms': getattr(args, 'lpc_window_ms', None),
        'scala_cent': getattr(args, 'scala_cent', None),
        'delta_threshold_hz': getattr(args, 'delta_threshold_hz', 0.0),
        'digits': getattr(args, 'digits', None)
    }


# =============================================================================
# UTILITY FUNCTIONS (moved from audio_analysis)
# =============================================================================

def deduplicate_steps(steps: List[float], tolerance: float = 1e-6) -> List[float]:
    """Remove duplicate steps within tolerance."""
    if not steps:
        return []
    steps_u = []
    for s in sorted(steps):
        if not steps_u or abs(s - steps_u[-1]) > tolerance:
            steps_u.append(s)
    return steps_u


def safe_float_positive(value, default: float = None) -> Optional[float]:
    """Safely convert value to positive float, return default if invalid."""
    try:
        if isinstance(value, (int, float)) and float(value) > 0:
            return float(value)
        return default
    except (ValueError, TypeError):
        return default


def apply_digits_rounding(value, digits: Optional[int]):
    """Apply rounding to a value if digits is specified."""
    if digits is not None and isinstance(value, (int, float)):
        return round(float(value), digits)
    return value



def get_tet_scale(n: int) -> List[float]:
    """Cached TET scale generation."""
    return [k * (1200.0 / float(n)) for k in range(n)]


@functools.lru_cache(maxsize=32)
def get_pythagorean_scales() -> Tuple[List[float], List[float]]:
    """Cached Pythagorean scale generation."""
    p12 = [ratio_to_cents(r) for r in generate_pythagorean_12()]
    p7 = sorted([ratio_to_cents(r) for r in generate_pythagorean_7()])
    return p12, p7
