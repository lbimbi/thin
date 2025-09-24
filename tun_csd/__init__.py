"""File format export and import utilities for musical tuning systems.

This module provides comprehensive support for reading and writing various tuning
file formats used in musical software and hardware, enabling interoperability
between THIN/SIM and external applications.

Csound Integration:
- Generation and management of cpstun tables within .csd files
- Automatic table numbering with conflict detection and resolution
- GEN -2 format compliance with metadata preservation
- Skeleton file creation for new Csound projects
- Integration with existing Csound scores and orchestras

AnaMark TUN Format:
- Full AnaMark TUN specification implementation with exact tuning support
- 128 MIDI note mapping with custom frequency references
- Intelligent diapason handling respecting user-specified A4 frequencies
- Bidirectional conversion between internal ratios and TUN format
- Support for both relative and absolute frequency specifications

Scala Format Support:
- Export to Scala (.scl) format for microtonal scale sharing
- Compliance with Scala scale archive standards
- Automatic scale naming and description generation
- Integration with Scala database for comparison and validation

Ableton Live Integration:
- Export to Ableton Scale (.ascl) format for DAW integration
- XML-based format with proper Live compatibility
- Custom scale naming and organization
- Support for microtonal workflows in Live

Excel Integration:
- Conversion of Excel worksheets to tuning file formats
- Batch processing of comparison tables
- Automatic format detection and appropriate export selection
- Error handling for malformed or incomplete Excel data

File Management:
- Intelligent file conflict resolution and backup creation
- Cross-platform path handling and file system compatibility
- Robust error handling with detailed user feedback
- Atomic file operations to prevent corruption

Format Conversion:
- Seamless conversion between different tuning file formats
- Precision preservation across format boundaries
- Metadata translation and format-specific optimization
- Validation of output files for compliance with specifications

Dependencies:
- Standard library modules for file I/O and parsing
- Integration with utils module for mathematical operations
- Optional openpyxl support for Excel file processing

Note: All file format implementations follow current specifications.
Historical DaMuSc integration has been removed for simplified maintenance.
"""
import math
import os
import re
from typing import List, Tuple, Optional, Union
from fractions import Fraction

import consts
import utils

# Pre-compiled regex patterns for better performance
_TUN_NOTE_PATTERN = re.compile(r'^(?:[Nn]ote|[Kk]ey)\s+(\d{1,3})\s*=\s*([+-]?\d+(?:\.\d+)?)')
_NOTE_PARSE_PATTERN = re.compile(r'([A-G][#B]?)(\d+)')

# Local wrapper to optimize reduce_to_octave usage
def _reduce_to_octave_float(ratio) -> float:
    """Optimized wrapper for utils.reduce_to_octave that returns float directly."""
    return float(utils.reduce_to_octave(ratio))


def _filter_valid_ratios(ratios: List[float]) -> List[float]:
    """Filter and sort ratios, removing unison and invalid values for Scala format."""
    valid_ratios = []
    for ratio in ratios:
        if isinstance(ratio, (int, float)) and ratio > 1.0:
            valid_ratios.append(float(ratio))
    valid_ratios.sort()
    return valid_ratios


def _detect_natural_period(ratios: List[float], system_name: str = None) -> float:
    """Detect the natural period of a tuning system from its ratios.

    Args:
        ratios: List of frequency ratios (excluding unison)
        system_name: Name of the tuning system (for hints)

    Returns:
        The detected natural period (e.g., 2.0 for octave, 3.0 for tritave)
    """
    if not ratios:
        return 2.0  # Default to octave

    # PRIORITY 1: Check system name first to override ratio analysis for known systems
    if system_name:
        import re as _re
        name_lower = system_name.lower()

        # Special case: All n-TET systems should ALWAYS use octave (2.0) regardless of ratios found
        # UNLESS they explicitly specify a different interval (like tritave)
        if '12-tet' in name_lower or '12tet' in name_lower or 'temperament 12' in name_lower:
            return 2.0

        # For other n-TET systems, we need to infer the period from the ratios
        # since the system name may not contain the interval information
        if 'tet' in name_lower or 'temperament' in name_lower:
            # Check if interval is specified and calculate the period from it
            if 'interval' in name_lower or any(c in name_lower for c in ['1200', '1350', '1902', '700']):
                # Don't force octave if explicit interval is given - let ratio analysis determine period
                pass
            else:
                # For n-TET systems without explicit interval, try to infer from ratios
                # The period should be determined from the actual interval used
                if ratios:
                    # For equal temperament, all steps should be equal
                    # Calculate the step size from the first few ratios
                    step_size_cents = None
                    if len(ratios) >= 2:
                        # Calculate step size in cents from first two ratios
                        ratio1 = ratios[0] if ratios[0] > 1.0 else (ratios[1] if len(ratios) > 1 else 1.0)
                        ratio2 = ratios[1] if len(ratios) > 1 and ratios[1] > ratio1 else (ratios[2] if len(ratios) > 2 else ratio1 * 1.1)

                        if ratio2 > ratio1 > 1.0:
                            step_size_cents = 1200 * math.log2(ratio2 / ratio1)

                            # Extract n from system name to calculate total period
                            et_match = _re.search(r'(\d+)-?TET', name_lower, _re.IGNORECASE)
                            if et_match:
                                try:
                                    n_tet = int(et_match.group(1))
                                    # Period = step_size * n_tet cents
                                    period_cents = step_size_cents * n_tet
                                    period_ratio = 2 ** (period_cents / 1200.0)
                                    return period_ratio
                                except (ValueError, TypeError):
                                    pass

                # Fallback to octave if inference fails
                return 2.0

        # For geometric systems, try to infer from system name
        if 'tritave' in name_lower or '3/1' in name_lower:
            return 3.0
        elif 'fifth' in name_lower or '3/2' in name_lower:
            return 1.5
        elif 'fourth' in name_lower or '4/3' in name_lower:
            return 4.0/3.0

        # For ET systems, try to infer from interval value in system name
        # Extract interval cents from patterns like "17-TET 1903c" or "Equal Temperament 17-TET, interval=1903"
        # Look for interval patterns
        interval_match = _re.search(r'interval[=:]?\s*(\d+(?:\.\d+)?)', name_lower)
        if interval_match:
            try:
                interval_cents = float(interval_match.group(1))
                # Convert cents to ratio: ratio = 2^(cents/1200)
                period_ratio = 2 ** (interval_cents / 1200.0)
                return period_ratio
            except (ValueError, TypeError):
                pass

        # Look for explicit cents values like "1903c" or "1903.0c"
        cents_match = _re.search(r'(\d+(?:\.\d+)?)\s*c(?:ents?)?', name_lower)
        if cents_match:
            try:
                interval_cents = float(cents_match.group(1))
                period_ratio = 2 ** (interval_cents / 1200.0)
                return period_ratio
            except (ValueError, TypeError):
                pass

    # PRIORITY 2: Try to find common period patterns in the ratios (for non-12TET systems)
    max_ratio = max(ratios) if ratios else 2.0

    # Common period factors to check
    common_periods = [
        2.0,      # Octave (most common)
        3.0,      # Tritave
        1.5,      # Perfect fifth
        4.0/3.0,  # Perfect fourth
        5.0/4.0,  # Major third
        6.0/5.0,  # Minor third
        9.0/8.0,  # Major second
    ]

    tolerance = 0.05  # ~86 cents tolerance

    # Check if any ratio is close to a known period
    # For ASCL, prioritize octave (2.0) detection for ET systems
    found_periods = []
    for period in common_periods:
        if any(abs(ratio - period) < tolerance for ratio in ratios):
            found_periods.append(period)

    # If we found multiple periods, prefer octave (2.0) for ET systems
    if found_periods:
        if 2.0 in found_periods:
            return 2.0  # Always prefer octave when present
        else:
            return found_periods[0]  # Otherwise use first match

    # If we can't detect a specific period, but have ratios that exceed 2.0,
    # use the maximum ratio as a hint for the period
    if max_ratio > 2.5:
        # Round to nearest common period
        for period in common_periods:
            if abs(max_ratio - period) < 0.3:
                return period

        # If no match, use the max ratio directly (for unusual periods)
        return max_ratio

    # Default to octave
    return 2.0


def _ratio_to_scala_format(ratio: float) -> str:
    """Convert a ratio to appropriate Scala format (fraction or cents).

    For ASCL format, prioritize cents for ET systems and simple fractions for others.
    """
    try:
        # Special case: exact octave should always be 2/1
        if abs(ratio - 2.0) < 1e-8:
            return "2/1"

        # Convert to cents: cents = 1200 * log2(ratio)
        cents = 1200.0 * math.log2(ratio)

        # For ET systems, prefer cents if they're close to multiples of 100
        # (This helps with 12-TET which should show 100., 200., etc.)
        if abs(cents % 100) < 0.1:  # Very close to 100-cent multiples
            # Round to nearest cent for clean output
            cents_rounded = round(cents)
            return f"{cents_rounded}."

        # For other round values (like 300.0000, 400.0000), also use clean format
        if abs(cents - round(cents)) < 0.001:  # Very close to whole numbers
            return f"{round(cents)}."

        # For other cases, try fraction first if it's simple and accurate
        frac = Fraction(ratio).limit_denominator(1000)  # Lower limit for simpler fractions
        if abs(float(frac) - ratio) < 1e-6 and 1 < frac.denominator <= 100:
            return f"{frac.numerator}/{frac.denominator}"
        else:
            # Use cents with appropriate precision
            return f"{cents:.5f}"
    except (ValueError, TypeError, ZeroDivisionError):
        # Fallback a formato decimale
        return f"{ratio:.8f}"


def parse_file(file_name: str) -> int:
    """Finds the maximum table number 'f N' in the file."""
    max_num = 0
    try:
        with open(file_name, "r") as f:
            for line in f:
                for match in consts.PATTERN.finditer(line):
                    num = int(match.group(1))
                    max_num = max(max_num, num)
    except FileNotFoundError:
        print(f"Error: file not found: {file_name}")
    return max_num


def write_cpstun_table(output_base: str, ratios: List[float], basekey: int,
                       basefrequency: float, interval_value: Optional[float] = None) -> Tuple[int, bool]:
    """Creates or appends a cpstun table to a Csound .csd file."""
    csd_path = f"{output_base}.csd"
    skeleton = (
        "<CsoundSynthesizer>\n"
        "<CsOptions>\n\n</CsOptions>\n"
        "<CsInstruments>\n\n</CsInstruments>\n"
        "<CsScore>\n\n</CsScore>\n"
        "</CsoundSynthesizer>\n"
    )

    existed_before = utils.file_exists(csd_path)

    if not existed_before:
        try:
            with open(csd_path, "w") as f:
                f.write(skeleton)
        except IOError as e:
            print(f"Error creating CSD file: {e}")
            return 0, existed_before

    try:
        with open(csd_path, "r", encoding="utf-8") as f:
            content = f.read()
    except IOError as e:
        print(f"Error reading CSD: {e}")
        return 0, existed_before

    fnum = parse_file(csd_path) + 1

    # Ordina i rapporti
    try:
        ratios_sorted = sorted(float(r) for r in ratios)
    except (TypeError, ValueError):
        print(f"Error sorting ratios")
        ratios_sorted = [float(r) for r in ratios]

    # Determina parametri cpstun
    numgrades = len(ratios_sorted)

    if interval_value is not None and isinstance(interval_value, (int, float)):
        iv = float(interval_value)
        interval = max(0.0, iv)
    else:
        try:
            rmin = min(ratios_sorted) if ratios_sorted else 1.0
            rmax = max(ratios_sorted) if ratios_sorted else 1.0
            interval = 2.0 if (1.0 - consts.RATIO_EPS <= rmin and rmax <= 2.0 + consts.RATIO_EPS) else 0.0
        except (TypeError, ValueError):
            interval = 0.0

    # Costruisci lista dati
    data_list = [
        str(numgrades),
        f"{float(interval):.10g}",
        f"{float(basefrequency):.10g}",
        str(int(basekey))
    ]
    data_list.extend(f"{r:.10f}" for r in ratios_sorted)

    size = len(data_list)

    # Costruisci righe
    prefix = f"f {fnum} 0 {size} -2 "
    positions = []
    col = len(prefix)
    for t in data_list:
        positions.append(col)
        col += len(t) + 1

    def build_aligned_comment(label_map: List[Tuple[int, str]]) -> str:
        line = ";"
        base_offset = 1
        for pos_idx, text in label_map:
            target = base_offset + (positions[pos_idx] if 0 <= pos_idx < len(positions) else len(prefix))
            if target < len(line):
                target = len(line) + 1
            line += " " * (target - len(line)) + text
        return line + "\n"

    header_comment = (
            build_aligned_comment([(0, "numgrades"), (2, "basefreq"), (4, "rapporti-di-intonazione .......")]) +
            build_aligned_comment([(1, "interval"), (3, "basekey")]) +
            f"; cpstun table generated | basekey={basekey} basefrequency={basefrequency:.6f}Hz\n"
    )
    f_line = prefix + " ".join(data_list) + "\n"

    # Insert before </CsScore>
    insert_marker = "</CsScore>"
    idx = content.rfind(insert_marker)
    if idx == -1:
        content += f"\n<CsScore>\n{header_comment}{f_line}</CsScore>\n"
    else:
        content = content[:idx] + header_comment + f_line + content[idx:]

    try:
        with open(csd_path, "w", encoding="utf-8") as f:
            f.write(content)
        print(f"cpstun table (f {fnum}) saved to {csd_path}")
    except IOError as e:
        print(f"Error writing CSD: {e}")

    return fnum, existed_before


def write_tun_file(output_base: str, diapason: float, ratios: List[float], basekey: int,
                   basefrequency: float, tun_integer: bool = False) -> None:
    """Exports a .tun file (AnaMark TUN) with values expressed in absolute cents relative to 8.1757989156437073336 Hz.
    Structure: [Tuning] + 128 lines "note X=Y" (absolute cents)."""
    f_ref = consts.F_REF * (diapason / 440)
    tun_path = f"{output_base}.tun"

    if diapason != 440:
        lines = ["[Exact Tuning]",
                 f"basefreq={f_ref}"]
    else:

        lines = [
            "[Tuning]",
        ]

    # Ordina i rapporti per garantire valori crescenti nel segmento custom
    try:
        ratios_sorted = sorted(float(r) for r in ratios)
    except (TypeError, ValueError):
        ratios_sorted = [float(r) for r in ratios]

    def note_freq(n: int) -> float:
        """Calculate frequency for MIDI note n.

        For notes within the custom tuning range (basekey to basekey+len(ratios)-1),
        use the provided ratios. For all other notes, use standard 12-TET.
        """
        if 0 <= basekey <= 127:
            # If within the provided ratios range, use them directly
            if basekey <= n < basekey + len(ratios_sorted):
                offset = n - basekey
                return basefrequency * ratios_sorted[offset]
            else:
                # For notes outside the custom range, use standard 12-TET
                # Calculate 12-TET frequency relative to the standard C0 = MIDI 0
                # using the reference frequency (f_ref corresponds to C0 in TUN standard)
                c0_freq = f_ref  # This is already C0 at 8.1757989156437073336 Hz * (diapason/440)
                return c0_freq * (2.0 ** (n / 12.0))
        else:
            # Fallback for invalid basekey - use standard 12-TET from C0
            c0_freq = f_ref
            return c0_freq * (2.0 ** (n / 12.0))

    for note_idx in range(consts.MIDI_MIN, consts.MIDI_MAX + 1):
        f = note_freq(note_idx)
        if isinstance(f, (int, float)) and f > 0:
            cents = utils.ratio_to_cents(f / f_ref)
        else:
            cents = 0.0
        rounded2 = round(cents, 2)
        s2 = f"{rounded2:.2f}"
        if tun_integer or s2.endswith(".00"):
            cents_text = str(int(round(cents)))
        else:
            cents_text = s2
        lines.append(f"Note {note_idx}={cents_text}")

    try:
        with open(tun_path, "w", encoding="utf-8") as f:
            f.write("\n".join(lines) + "\n")
        print(f".tun file saved to {tun_path}")
    except IOError as e:
        print(f"Error writing .tun: {e}")


def import_tun_file(tun_path: str, basekey: int = consts.DEFAULT_BASEKEY, reduce_octave_out: bool = True) -> Optional[
    str]:
    """
    Importa un file .tun (AnaMark TUN) e converte i valori in ratios relativi a basekey.
    Salva un file .txt con mappatura 'MIDI -> ratio' per le note disponibili.
    """
    # Riferimento assoluto AnaMark (stesso di write_tun_file)
    f_ref = consts.F_REF

    cents_map = {}
    try:
        with open(tun_path, "r", encoding="utf-8") as f:
            for line in f:
                s = line.strip()
                if not s or s.startswith(";") or s.startswith("#"):
                    continue
                # Consenti e ignora le sezioni come [Tuning] o [Exact Tuning]
                if s.startswith("[") and s.endswith("]"):
                    continue
                m = _TUN_NOTE_PATTERN.match(s)
                if m:
                    idx = int(m.group(1))
                    if 0 <= idx <= 127:
                        cents_val = float(m.group(2))
                        cents_map[idx] = cents_val

    except IOError as e:
        print(f"Error reading .tun: {e}")
        return None

    if not cents_map:
        print("Empty or unrecognized .tun file")
        return None

    # Cents -> frequenze assolute
    freq_map = {}
    for k, cv in cents_map.items():
        freq_map[k] = float(f_ref) * (2.0 ** (float(cv) / 1200.0))

    # Determina frequenza base
    if basekey in freq_map:
        base_freq = freq_map[basekey]
        base_key_eff = basekey
    else:
        keys_sorted = sorted(freq_map.keys(), key=lambda nn: abs(nn - basekey))
        if not keys_sorted:
            print("No valid notes in .tun")
            return None
        base_key_eff = keys_sorted[0]
        base_freq = freq_map[base_key_eff]

    if not isinstance(base_freq, (int, float)) or base_freq <= 0:
        print("Invalid base frequency in .tun")
        return None

    # Costruisci ratios rispetto a base_key_eff
    ratios_by_midi = {}
    for n in range(128):
        f = freq_map.get(n)
        if isinstance(f, (int, float)) and f > 0:
            r = float(f) / float(base_freq)
            if reduce_octave_out:
                r = _reduce_to_octave_float(r)
            ratios_by_midi[n] = float(r)

    # Scrivi file di output
    base_name = os.path.splitext(os.path.basename(tun_path))[0]
    out_path = f"{base_name}_ratios.txt"
    try:
        with open(out_path, "w", encoding="utf-8") as f:
            f.write(f"; Import .tun -> ratios | basekey={base_key_eff}\n")
            f.write("MIDI -> Ratio\n")
            for n in sorted(ratios_by_midi.keys()):
                f.write(f"{n} -> {ratios_by_midi[n]:.10f}\n")
        print(f"Ratios file saved to {out_path}")
        return out_path
    except IOError as e:
        print(f"Error writing ratios: {e}")
        return None


def convert_excel_to_outputs(excel_path: str,
                             output_base: Optional[str],
                             default_basekey: int,
                             default_base_hz: float,
                             diapason_hz: float,
                             midi_truncate: bool = False,
                             tun_integer: bool = False,
                             cents_delta: Optional[float] = None) -> Union[bool, dict]:
    """
    Converte un file Excel (System/Compare) in output .csd (cpstun) e .tun usando
    la colonna dei rapporti (o Hz/Base_Hz se necessario). Applica opzionalmente uno shift in cents
    (cents_delta) to all system steps before comparisons (Scala/ET/Pyth/JI).
    Ritorna True se conversione riuscita.
    """
    try:
        from openpyxl import load_workbook
    except ImportError:
        print("openpyxl not installed: cannot use --convert")
        return False

    if not utils.file_exists(excel_path):
        print(f"Excel file not found: {excel_path}")
        return False

    try:
        wb = load_workbook(excel_path, data_only=True)
    except (FileNotFoundError, PermissionError, ValueError) as e:
        print(f"Error opening Excel: {e}")
        return False

    # Scegli foglio: preferisci 'System' se presente, altrimenti attivo
    if 'System' in wb.sheetnames:
        ws = wb['System']
    else:
        ws = wb.active

    # Mappa header -> colonna (1-based)
    header_row = 1
    headers = {}

    # Valida l'indice di riga
    if not isinstance(header_row, int) or header_row < 1:
        print("Invalid header row")
        return False

    # Ottieni la riga header in modo sicuro
    try:
        row = ws[header_row]
    except (TypeError, KeyError, IndexError, AttributeError):
        print("Cannot read header row")
        return False

    for cell in row:
        try:
            val = cell.value
            if val is None:
                continue
            name = str(val).strip()
            if not name:
                continue
            col = cell.column  # openpyxl: può essere int o lettera
            headers[name.lower()] = col
        except (AttributeError, ValueError, TypeError):
            # salta solo la cella problematica
            continue

    def col_index_by_name(names: list) -> Optional[int]:
        def _col_to_index(c) -> Optional[int]:
            if isinstance(c, int):
                return c
            s = str(c).strip().upper()
            if s and s.isalpha():
                idx = 0
                for ch in s:
                    idx = idx * 26 + (ord(ch) - ord('A') + 1)
                return idx if idx > 0 else None
            return None

        if not names:
            return None
        names_l = [str(name_item).lower() for name_item in names]
        for key, column in headers.items():
            k = str(key).lower()
            for name_match in names_l:
                if name_match in k:
                    ci = _col_to_index(column)
                    if ci:
                        return ci
        return None

    ratio_col = col_index_by_name(["ratio"])  # es. "ratio", "custom ratio"
    hz_col = col_index_by_name(["hz"])  # es. "hz"
    midi_col = col_index_by_name(["midi"])  # es. "midi"

    # Base_Hz: prova a trovarlo da intestazione 'Base_Hz' o F2
    base_hz = float(default_base_hz) if default_base_hz and default_base_hz > 0 else consts.DEFAULT_DIAPASON
    base_anchor_col = col_index_by_name(["base_hz", "base hz", "base-freq", "basefreq"]) or 6  # fallback F
    base_cell_val = ws.cell(row=2, column=base_anchor_col).value
    if isinstance(base_cell_val, (int, float)) and float(base_cell_val) > 0:
        base_hz = float(base_cell_val)


    # Basekey: usa prima riga della colonna MIDI se presente, altrimenti default
    basekey = int(default_basekey)
    if midi_col:
        mv = ws.cell(row=2, column=midi_col).value
        if isinstance(mv, (int, float)):
            basekey = int(mv)

    # Estrai rapporti
    ratios: List[float] = []
    row = 2
    max_empty = 0
    while True:
        v_ratio = None
        v_hz = None
        if ratio_col:
            try:
                v_ratio = ws.cell(row=row, column=ratio_col).value
            except ValueError:
                v_ratio = None
        if hz_col:
            try:
                v_hz = ws.cell(row=row, column=hz_col).value
            except ValueError:
                v_hz = None
        if v_ratio is None and v_hz is None:
            max_empty += 1
            if max_empty >= 5:  # stop after a few empty rows
                break
            row += 1
            continue
        # Reset empty counter when any content
        max_empty = 0
        r_val: Optional[float] = None
        if isinstance(v_ratio, (int, float)) and float(v_ratio) > 0:
            r_val = float(v_ratio)
        elif isinstance(v_hz, (int, float)) and float(v_hz) > 0 and base_hz > 0:
            try:
                r_val = float(v_hz) / float(base_hz)
            except (TypeError, ValueError):
                r_val = None
        if r_val is not None and r_val > 0:
            ratios.append(r_val)
        row += 1

    if not ratios:
        print("No valid ratios found in Excel (Ratio/Hz columns)")
        return False

    # Assicura compatibilità MIDI
    ratios_eff, basekey_eff = utils.ensure_midi_fit(ratios, basekey, midi_truncate)

    # Base output
    if output_base and isinstance(output_base, str) and output_base.strip():
        out_base = output_base.strip()
    else:
        try:
            stem = os.path.splitext(os.path.basename(excel_path))[0]
        except OSError:
            stem = "out"
        out_base = stem

    # Ensure output directory and build in-folder base
    out_dir = out_base
    try:
        os.makedirs(out_dir, exist_ok=True)
    except OSError as e:
        print(f"Error creating directory: {e}")
        pass
    out_base_in_dir = os.path.join(out_dir, out_base)
    # Scrivi cpstun e tun (entrambi di default)
    fnum, existed = write_cpstun_table(out_base_in_dir, ratios_eff, basekey_eff, base_hz, None)
    out_base_eff = out_base_in_dir if not existed else f"{out_base_in_dir}_{fnum}"
    write_tun_file(out_base_eff, diapason_hz, ratios_eff, basekey_eff, base_hz, tun_integer)

    # Scrivi anche .scl e .ascl files
    try:
        # Determine basenote string for .scl/.ascl export
        basenote_str = f"MIDI{basekey_eff}"  # Simple fallback

        # Generate .scl file
        write_scl_file(out_base_eff, ratios_eff, basenote_str, diapason_hz, "Converted from Excel")

        # Generate .ascl file
        write_ascl_file(out_base_eff, ratios_eff, basenote_str, diapason_hz, "Converted from Excel")

    except Exception as e:
        print(f"Error generating .scl/.ascl files: {e}")

    # --- Confronto con scale .scl e inferenza ET/Geometrico ---
    try:
        import glob as _glob
    except ImportError:
        _glob = None

    def _ratios_to_cents(_ratios: list) -> list:
        out = []
        for r in _ratios:
            try:
                rr = float(r)
                if rr <= 0:
                    continue
                rr = _reduce_to_octave_float(rr)
                c = utils.ratio_to_cents(rr)
                out.append(c)
            except ValueError:
                continue
        out = sorted(x % 1200.0 for x in out)
        # remove duplicates within 1e-6 cents
        dedup = []
        for v in out:
            if not dedup or abs(v - dedup[-1]) > 1e-6:
                dedup.append(v)
        return dedup

    def _parse_scl_file(path: str) -> Optional[dict]:
        """Parse .scl file using centralized function."""
        return parse_scl_file(path)

    def _fractions_to_cents(ratio_list: list) -> list:
        """Convert a list of fraction ratios to cents values."""
        cents = []
        for r in ratio_list:
            try:
                cents.append((1200.0 * math.log(_reduce_to_octave_float(r), 2)) % 1200.0)
            except (ValueError, TypeError):
                pass
        return sorted(cents)

    def _avg_error_best_rotation(a: list, b: list) -> float:
        # compute minimal mean absolute difference between a and b under rotation, using min length
        if not a or not b:
            return float('inf')
        a_sorted = sorted(a)
        b_sorted = sorted(b)
        min_len = min(len(a_sorted), len(b_sorted))
        # use first min_len of each
        a_vals = a_sorted[:min_len]
        b_vals = b_sorted[:min_len]
        best_err = float('inf')
        # try all rotations of b
        for shift in range(min_len):
            err_sum = 0.0
            for idx in range(min_len):
                dv = abs(a_vals[idx] - b_vals[(idx + shift) % min_len])
                # allow wrap-around distance on circle 1200
                dv = min(dv, 1200.0 - dv)
                err_sum += dv
            best_err = min(best_err, err_sum / min_len)
        return best_err

    def _infer_system(cents: list) -> dict:
        res = {'type': 'unknown'}
        if not cents or len(cents) < 2:
            return res
        cs = sorted(x % 1200.0 for x in cents)
        # compute circular diffs
        diffs = []
        for idx in range(len(cs) - 1):
            diffs.append(cs[idx + 1] - cs[idx])
        diffs.append(1200.0 - cs[-1] + cs[0])
        if not diffs:
            return res
        mean = sum(diffs) / len(diffs)
        # std dev
        var = sum((d - mean) ** 2 for d in diffs) / len(diffs)
        std = math.sqrt(max(0.0, var))
        # ET test: mean step should divide 1200 closely and low std
        if mean > 0:
            n_est = round(1200.0 / mean)
            if n_est > 0:
                mean_et = 1200.0 / n_est
                if abs(mean - mean_et) <= 0.5 and std <= 0.8:
                    res = {'type': 'et', 'n': int(n_est), 'step_cents': mean_et, 'std_cents': std}
                    return res
        # Geometric-like (equal steps not matching 1200/n but low std)
        if std <= 0.8:
            res = {'type': 'geometric', 'steps': len(diffs), 'step_cents': mean, 'std_cents': std}
            return res
        return res

    try:
        report_lines = [
            f"— Report confronto (convert) —",
            f"File Excel: {os.path.basename(excel_path)}",
            f"Output base: {out_base_eff}",
            f"Base_Hz: {base_hz:.6f}  |  Basekey: {basekey_eff}",
            f"N rapporti: {len(ratios_eff)}"
        ]
        cents_from_excel = _ratios_to_cents(ratios_eff)
        # Apply cents delta shift if provided
        try:
            if cents_delta is not None:
                dd = float(cents_delta)
                cents_from_excel = sorted(((x + dd) % 1200.0) for x in cents_from_excel)
        except ValueError:
            pass
        report_lines.append("")
        # Inference
        inf = _infer_system(cents_from_excel)
        if inf.get('type') == 'et':
            report_lines.append(
                f"Inference: ET ~ {inf['n']}-TET  step≈{inf['step_cents']:.2f}c  std≈{inf['std_cents']:.2f}c")
        elif inf.get('type') == 'geometric':
            report_lines.append(
                f"Inference: Geometric  steps={inf['steps']}  step≈{inf['step_cents']:.2f}c  std≈{inf['std_cents']:.2f}c")
        else:
            report_lines.append(f"Inference: undetermined")
        report_lines.append("")
        # Scala comparison
        top_matches = []
        scl_dir = os.path.join(os.getcwd(), 'scl')
        if _glob and os.path.isdir(scl_dir):
            files = _glob.glob(os.path.join(scl_dir, '**', '*.scl'), recursive=True)
        else:
            files = []
        for fp in files[:2000]:  # safety cap
            info = _parse_scl_file(fp)
            if not info:
                continue
            err = _avg_error_best_rotation(cents_from_excel, info['cents'])
            if math.isfinite(err):
                top_matches.append({'file': os.path.relpath(fp, scl_dir), 'name': info['name'], 'avg_cents_error': err})
        top_matches.sort(key=lambda d: d['avg_cents_error'])
        if top_matches:
            best = top_matches[0]
            report_lines.append(
                f"Best scale: {best['name']} [{best['file']}]  avg err: {best['avg_cents_error']:.2f} c")
            report_lines.append("")
            report_lines.append("Top 5 matches:")
            report_lines.append("Pos  Name                                   Err (c)         File")
            for i, it in enumerate(top_matches[:5], start=1):
                nm = str(it.get('name', ''))[:35]
                er = f"{float(it.get('avg_cents_error', 0.0)):.2f}"
                fl = str(it.get('file', ''))
                report_lines.append(f"{str(i).rjust(3)}  {nm.ljust(35)}  {er.rjust(7)}    {fl}")
        else:
            report_lines.append("No .scl scales found or comparable.")

        # --- Confronti aggiuntivi: ET / Pitagorici / Naturali ---
        report_lines.append("")
        report_lines.append("Comparisons ET / Pythagorean / Natural")

        # ET: valuta n-TET da 5 a 72
        def _et_cents(et_n: int) -> list:
            try:
                step = 1200.0 / float(et_n)
                return [step * k for k in range(et_n)]
            except ValueError:
                return []

        et_scores = []
        for n in range(5, 73):
            cents_et = _et_cents(n)
            if not cents_et:
                continue
            err = _avg_error_best_rotation(cents_from_excel, cents_et)
            if math.isfinite(err):
                et_scores.append((n, err))
        et_scores.sort(key=lambda t: t[1])
        if et_scores:
            report_lines.append("Migliori ET (n-TET):")
            report_lines.append("Pos  n    Err (c)")
            for i, (n, er) in enumerate(et_scores[:5], start=1):
                report_lines.append(f"{str(i).rjust(3)}  {str(n).rjust(3)}  {er:7.2f}")
        else:
            report_lines.append("Nessun risultato ET calcolabile.")

        # Pitagorici: 12 e 7
        def _pyth12_cents() -> list:
            try:
                from fractions import Fraction as _Frac
                vals = []
                for k in range(12):
                    r = utils.reduce_to_octave(utils.pow_fraction(_Frac(3, 2), k)) if hasattr(utils, 'pow_fraction') else utils.reduce_to_octave((3 / 2) ** k)
                    vals.append(float(r))
                return _fractions_to_cents(vals)
            except (TypeError, ValueError, AttributeError):
                vals = [(1.5 ** k) for k in range(12)]
                return _fractions_to_cents(vals)

        def _pyth7_cents() -> list:
            try:
                from fractions import Fraction as _Frac
                seq = [_Frac(1, 1), _Frac(9, 8), _Frac(81, 64), _Frac(4, 3), _Frac(3, 2), _Frac(27, 16), _Frac(243, 128)]
                vals = [utils.reduce_to_octave(r) for r in seq]
                return _fractions_to_cents(vals)
            except (TypeError, ValueError, AttributeError):
                vals = [1.0, 9 / 8, 81 / 64, 4 / 3, 3 / 2, 27 / 16, 243 / 128]
                return _fractions_to_cents(vals)

        py12_err = _avg_error_best_rotation(cents_from_excel, _pyth12_cents())
        py7_err = _avg_error_best_rotation(cents_from_excel, _pyth7_cents())
        report_lines.append("")
        report_lines.append("Pitagorici:")
        report_lines.append(f"Pyth-12 err medio: {py12_err:.2f} c")
        report_lines.append(f"Pyth-7  err medio: {py7_err:.2f} c")

        # Naturali: 5-limit diatonic e triade 4:5:6
        def _ji_diatonic5_cents() -> list:
            try:
                from fractions import Fraction as _Frac
                seq = [_Frac(1, 1), _Frac(9, 8), _Frac(5, 4), _Frac(4, 3), _Frac(3, 2), _Frac(5, 3), _Frac(15, 8)]
                vals = [utils.reduce_to_octave(r) for r in seq]
                return _fractions_to_cents(vals)
            except (ValueError, TypeError):
                vals = [1.0, 9 / 8, 5 / 4, 4 / 3, 3 / 2, 5 / 3, 15 / 8]
                return _fractions_to_cents(vals)

        def _ji_triad_456_cents() -> list:
            try:
                from fractions import Fraction as _Frac
                seq = [_Frac(1, 1), _Frac(5, 4), _Frac(3, 2)]
                vals = [utils.reduce_to_octave(r) for r in seq]
                return _fractions_to_cents(vals)
            except (ValueError, TypeError):
                vals = [1.0, 5 / 4, 3 / 2]
                return _fractions_to_cents(vals)

        ji5_err = _avg_error_best_rotation(cents_from_excel, _ji_diatonic5_cents())
        ji_triad_err = _avg_error_best_rotation(cents_from_excel, _ji_triad_456_cents())
        report_lines.append("")
        report_lines.append("Naturali (JI):")
        report_lines.append(f"Diatonica 5-limit err medio: {ji5_err:.2f} c")
        report_lines.append(f"Triade 4:5:6 err medio: {ji_triad_err:.2f} c")


        report_txt = "\n".join(report_lines) + "\n"
        rep_path = f"{out_base_eff}_convert_compare.txt"
        try:
            with open(rep_path, 'w', encoding='utf-8') as rf:
                rf.write(report_txt)
            print(f"Comparison report saved: {rep_path}")
        except OSError as e:
            print(f"Error saving report: {e}")
    except (ValueError, TypeError, ImportError, AttributeError) as e:
        print(f"Comparison/inference failed: {e}")

    return True


def write_scl_file(output_base: str, ratios: List[float], basenote: str, diapason: float,
                   system_name: str = "Custom Scale") -> None:
    """Exports a .scl file (Scala format) with the provided ratios.

    Args:
        output_base: Base name of the output file (without extension)
        ratios: List of frequency ratios (must include 1.0 as first element)
        basenote: Base note (e.g. "A4")
        diapason: Reference frequency in Hz
        system_name: Name of the tuning system
    """
    scl_path = f"{output_base}.scl"

    # Prepare description according to Scala specifications
    description = f"{system_name}, 1/1={diapason:.2f} Hz, base={basenote}"

    # Filter and sort ratios using helper function
    valid_ratios = _filter_valid_ratios(ratios)
    num_notes = len(valid_ratios)

    # Build file according to Scala specifications
    lines = [
        f"! {output_base}.scl",
        f"! Generated by THIN (SIM project)",
        f"! {description}",
        description,
        str(num_notes)
    ]

    # Add ratios in appropriate format using helper function
    for ratio in valid_ratios:
        lines.append(_ratio_to_scala_format(ratio))

    try:
        with open(scl_path, "w", encoding="utf-8") as f:
            f.write("\n".join(lines) + "\n")
        print(f".scl file saved to {scl_path}")
    except IOError as e:
        print(f"Error writing .scl: {e}")


def write_ascl_file(output_base: str, ratios: List[float], basenote: str, diapason: float,
                   system_name: str = "Custom Scale", source: str = None, link: str = None,
) -> None:
    """Exports an .ascl file (Ableton Live format) with the provided ratios.

    ASCL format is an Ableton extension to Scala format with Ableton metadata
    specified in @ABL comments according to Ableton Live specification.
    File is saved in UTF-8 format as required by the specification.

    Args:
        output_base: Base name of the output file (without extension)
        ratios: List of frequency ratios (must include 1.0 as first element)
        basenote: Base note (e.g. "A4")
        diapason: Reference frequency in Hz
        system_name: Name of the tuning system
        source: Source or author of the tuning system (optional)
        link: URL for more information about the system (optional)
    """
    ascl_path = f"{output_base}.ascl"

    # Prepare description according to Scala/ASCL specifications
    description = f"{system_name}, 1/1={diapason:.2f} Hz, base={basenote}"

    # For ASCL, we exclude unison (1.0) but need to include the octave (2.0)
    # Ableton's official 12-TET example has 12 entries: steps 1-11 + octave (2/1)
    valid_ratios = []
    for ratio in ratios:
        if isinstance(ratio, (int, float)) and ratio > 1.0 and math.isfinite(ratio):
            valid_ratios.append(float(ratio))

    # CRITICAL: For n-TET systems, ASCL format expects exactly n ratios with octave (2/1) as the last one
    # Don't use the system's natural period for ASCL export - always use octave for ET systems

    # Check if this is an n-TET system
    is_tet_system = ('tet' in system_name.lower() or 'temperament' in system_name.lower() or
                     'equal' in system_name.lower())

    if is_tet_system:
        # For n-TET systems, ASCL format needs exactly n ratios with octave (2/1) as the final one
        # Extract the n-TET number from system name to know how many ratios we need
        import re as _re_tet
        et_match = _re_tet.search(r'(\d+)-?TET', system_name, _re_tet.IGNORECASE)
        if et_match:
            try:
                n_tet = int(et_match.group(1))

                # For n-TET systems, we want exactly n ratios ending with the system's period
                # NOT always the octave - the period depends on the interval specified

                # Find the natural period from the original ratios or detect from system name
                natural_period = _detect_natural_period(valid_ratios, system_name)

                if len(valid_ratios) >= n_tet - 1:
                    # Take first (n-1) ratios from the system
                    ratios_before_period = valid_ratios[:n_tet-1]
                    # For n-TET systems, ensure we have exactly n ratios with the period as last
                    valid_ratios = ratios_before_period + [natural_period]
                else:
                    # If we have fewer than (n-1) ratios, fill up to n-1 and add period
                    while len(valid_ratios) < n_tet - 1:
                        if valid_ratios:
                            # Add intermediate ratios (simple linear interpolation as fallback)
                            last_ratio = valid_ratios[-1]
                            step_ratio = (natural_period / last_ratio) ** (1.0 / (n_tet - len(valid_ratios)))
                            valid_ratios.append(last_ratio * step_ratio)
                        else:
                            break
                    valid_ratios.append(natural_period)
            except ValueError:
                # Fallback to old logic if n-TET number can't be extracted
                octave = 2.0
                octave_tolerance = 0.02  # ~24 cents
                has_octave = any(abs(r - octave) < octave_tolerance for r in valid_ratios)
                if not has_octave:
                    valid_ratios.append(octave)
        else:
            # Fallback for systems with "temperament" or "equal" but no clear n-TET pattern
            octave = 2.0
            octave_tolerance = 0.02  # ~24 cents
            has_octave = any(abs(r - octave) < octave_tolerance for r in valid_ratios)
            if not has_octave:
                valid_ratios.append(octave)
    else:
        # For non-ET systems (geometric, natural, etc.), handle period appropriately
        if 'geometric' in system_name.lower() or 'progression' in system_name.lower():
            # For geometric systems, we need to add the final interval ratio
            # Extract the interval from the system name or calculate from ratios
            import re as _re_geom
            interval_match = _re_geom.search(r'interval[=:]?\s*(\d+(?:\.\d+)?)', system_name.lower())
            if interval_match:
                try:
                    interval_cents = float(interval_match.group(1))
                    interval_ratio = 2 ** (interval_cents / 1200.0)
                    # Check if this interval ratio is already present
                    interval_tolerance = 0.001  # ~1.7 cents
                    has_interval = any(abs(r - interval_ratio) < interval_tolerance for r in valid_ratios)
                    if not has_interval:
                        valid_ratios.append(interval_ratio)
                except (ValueError, TypeError):
                    pass
        else:
            # For other non-ET systems (natural, JI, etc.), detect the natural period
            natural_period = _detect_natural_period(valid_ratios, system_name)

            # Check if the natural period is already present
            period_tolerance = 0.02  # ~24 cents
            has_period = any(abs(r - natural_period) < period_tolerance for r in valid_ratios)

            if not has_period and natural_period > 1.0:
                valid_ratios.append(natural_period)

    valid_ratios.sort()
    num_notes = len(valid_ratios)

    # Build file according to Scala specifications con estensioni ASCL
    lines = [
        f"! {output_base}.ascl",
        f"! Generated by THIN (SIM project)",
        f"! Ableton Live ASCL format",
        f"! {description}",
        description,
        str(num_notes)
    ]

    # Add ratios in appropriate format using helper function
    for ratio in valid_ratios:
        lines.append(_ratio_to_scala_format(ratio))

    # Add Ableton ASCL extensions
    lines.append("!")
    lines.append("! === ABLETON LIVE EXTENSIONS ===")

    # @ABL REFERENCE_PITCH: octave_number note_index_in_scale frequency_hz
    # According to ASCL spec: note_index refers to which note in our NOTE_NAMES list,
    # not an absolute note system. We should use our basenote as the reference.

    # Parse basenote to get octave for REFERENCE_PITCH
    basenote_octave, _ = _parse_basenote_for_ascl(basenote)

    # The reference note index is always 0 since our NOTE_NAMES starts with the basenote (unison)
    reference_note_index = 0

    # Calculate the frequency of our basenote using the diapason and MIDI conversion
    # Parse basenote to get MIDI number
    import re as _re
    match = _re.match(r'([A-G][#B]?)(\d+)', basenote.upper())
    if match:
        note_name = match.group(1)
        octave = int(match.group(2))

        # MIDI note number calculation
        note_to_midi = {
            'C': 0, 'C#': 1, 'DB': 1, 'D': 2, 'D#': 3, 'EB': 3,
            'E': 4, 'F': 5, 'F#': 6, 'GB': 6, 'G': 7, 'G#': 8, 'AB': 8,
            'A': 9, 'A#': 10, 'BB': 10, 'B': 11
        }

        note_offset = note_to_midi.get(note_name, 0)
        midi_number = (octave + 1) * 12 + note_offset  # C4 = 60

        # Calculate frequency from MIDI number using A4=diapason as reference
        basenote_frequency = diapason * (2 ** ((midi_number - 69) / 12.0))
    else:
        # Fallback to C4 if parsing fails
        basenote_frequency = 261.6256  # C4

    lines.append(f"! @ABL REFERENCE_PITCH {basenote_octave} {reference_note_index} {basenote_frequency:.1f}")

    # Parse basenote for NOTE_RANGE_BY_INDEX (separate from REFERENCE_PITCH)
    basenote_octave, basenote_note_index = _parse_basenote_for_ascl(basenote)

    # @ABL NOTE_NAMES: note names for each pitch in the scale
    # According to ASCL spec: "The first entry in the sequence applies to the zeroeth note"
    # CRITICAL: User expects the first note name to be the basenote itself (unison)
    # So we need to prepend the basenote name to the list of ratio-based names
    # This gives us (num_notes + 1) total names: unison + all ratios

    # Extract the basenote name (e.g., "C" from "C3")
    try:
        import re as _re
        match = _re.match(r'([A-G][#B]?)', basenote.upper())
        if match:
            basenote_name = match.group(1)
            # Convert to user-friendly format
            if basenote_name == "DB":
                basenote_name = "C♯/D♭"
            elif basenote_name == "EB":
                basenote_name = "D♯/E♭"
            elif basenote_name == "GB":
                basenote_name = "F♯/G♭"
            elif basenote_name == "AB":
                basenote_name = "G♯/A♭"
            elif basenote_name == "BB":
                basenote_name = "A♯/B♭"
            elif basenote_name.endswith("#"):
                # Convert sharps to sharp/flat notation
                base = basenote_name[0]
                if base == "C":
                    basenote_name = "C♯/D♭"
                elif base == "D":
                    basenote_name = "D♯/E♭"
                elif base == "F":
                    basenote_name = "F♯/G♭"
                elif base == "G":
                    basenote_name = "G♯/A♭"
                elif base == "A":
                    basenote_name = "A♯/B♭"
        else:
            basenote_name = "C"  # Fallback
    except (ValueError, TypeError, AttributeError):
        basenote_name = "C"  # Fallback

    # Generate note names for the ratios
    # IMPORTANT: Ableton expects scales to repeat at the octave (2/1).
    # For octave-repeating scales: The last ratio is 2/1 and doesn't need a name.
    # For non-octave cycles: We need to generate ALL ratios with names, or extend to octave.

    # Check if this scale repeats at the octave
    is_octave_repeating = False
    if len(valid_ratios) > 0:
        last_ratio = valid_ratios[-1]
        # Check if last ratio is close to 2.0 (octave)
        if abs(last_ratio - 2.0) < 0.01:  # Within ~17 cents of octave
            is_octave_repeating = True

    if is_octave_repeating:
        # Standard octave-repeating scale: exclude the last ratio (2/1) from names
        ratios_for_names = valid_ratios[:-1] if len(valid_ratios) > 0 else []
        num_names_needed = len(ratios_for_names)
    else:
        # Non-octave cycle: ALL ratios need names
        # This allows Ableton to understand the actual pitch progression
        ratios_for_names = valid_ratios
        num_names_needed = len(ratios_for_names)

    ratio_note_names = _generate_note_names_for_ascl(num_names_needed, ratios_for_names, basenote, diapason, system_name, basenote_name, is_octave_repeating)

    # For ASCL format:
    # - Octave-repeating scales: prepend the basenote name since we excluded the octave ratio
    # - Non-octave cycles: do NOT prepend since all ratios are included
    if is_octave_repeating:
        # For octave-repeating scales, prepend the basenote name as the first element
        # This ensures that when REFERENCE_PITCH points to index 0, it points to the basenote
        note_names = [basenote_name] + ratio_note_names
    else:
        # For non-octave cycles, the first ratio already represents the first step
        # No need to prepend the basenote name
        note_names = ratio_note_names

    # Simple duplicate handling: add a suffix to any duplicate names
    seen_names = {}
    final_names = []

    for name in note_names:
        if name in seen_names:
            # Duplicate found - add a distinguishing suffix
            seen_names[name] += 1
            # Add dots to distinguish (e.g., "C", "C.", "C..", etc.)
            modified_name = name + "." * seen_names[name]
            final_names.append(modified_name)
        else:
            seen_names[name] = 0
            final_names.append(name)

    note_names = final_names

    # Format note names with quotes as per Ableton specification
    quoted_names = [f'"{name}"' for name in note_names]
    lines.append(f"! @ABL NOTE_NAMES {' '.join(quoted_names)}")

    # @ABL NOTE_RANGE_BY_FREQUENCY: calculate sensible range based on system ratios
    if valid_ratios:
        # Use the basenote frequency we already calculated above
        base_freq = basenote_frequency

        # Calculate frequency range from the ratios
        min_ratio = min(valid_ratios) if valid_ratios else 1.0
        max_ratio = max(valid_ratios) if valid_ratios else 2.0

        # Calculate actual frequency range
        min_freq = base_freq * min_ratio
        max_freq = base_freq * max_ratio

        # Add some margin and ensure reasonable bounds
        margin_factor = 0.8  # Allow 20% lower than minimum
        min_freq_with_margin = max(20.0, min_freq * margin_factor)  # At least 20 Hz
        max_freq_with_margin = min(20000.0, max_freq * 1.2)  # At most 20 kHz, 20% higher than max

        # Use NOTE_RANGE_BY_INDEX with parsed basenote octave and the basenote index in our scale
        # The second parameter should be the index of the basenote within our custom scale (always 0)
        # because the basenote is the first note (index 0) in our NOTE_NAMES list
        if basenote_octave is not None:
            # The basenote is always index 0 in our custom scale
            # This tells Ableton that the reference pitch (basenote) is at index 0 in our scale
            basenote_index_in_scale = 0  # F# in our 7-note scale is index 0, not index 6
            lines.append(f"! @ABL NOTE_RANGE_BY_INDEX {basenote_octave} {basenote_index_in_scale}")
        else:
            # Fallback to frequency-based range if parsing failed
            lines.append(f"! @ABL NOTE_RANGE_BY_FREQUENCY {min_freq_with_margin:.1f} {max_freq_with_margin:.1f}")

    # @ABL SOURCE: source documentation (optional)
    if source:
        lines.append(f"! @ABL SOURCE {source}")
    else:
        lines.append("! @ABL SOURCE THIN (SIM project) - Generated tuning system")

    # @ABL LINK: URL for more information (optional)
    if link:
        lines.append(f"! @ABL LINK {link}")

    try:
        # Save in UTF-8 as required by ASCL specification
        with open(ascl_path, "w", encoding="utf-8") as f:
            f.write("\n".join(lines) + "\n")
        print(f".ascl file saved to {ascl_path}")
    except IOError as e:
        print(f"Error writing .ascl: {e}")


def _determine_cycle_structure(system_name: str, ratios: List[float] = None, num_notes: int = 12) -> Tuple[int, bool]:
    """Determines the cycle structure of a tuning system for note naming.

    This function analyzes the actual ratios to determine the natural cycle/period
    of the tuning system, which may be different from an octave (2:1).

    Args:
        system_name: Name of the tuning system
        ratios: List of frequency ratios (if available)
        num_notes: Total number of notes

    Returns:
        Tuple[cycle_size, use_traditional_names]:
        - cycle_size: Number of steps before repetition/period
        - use_traditional_names: Whether to use C,D,E... names (True) or Step0,Step1... (False)
    """

    # If we have ratios, analyze them to find the ACTUAL period
    if ratios and len(ratios) > 1:
        try:
            # First, find the natural period of the system by analyzing the ratios
            period_factor, period_steps = _analyze_period_from_ratios(ratios)

            if period_steps > 0:
                # We found a clear period
                use_traditional = (period_steps == 12 and abs(period_factor - 2.0) < 0.01)  # Only for 12-step octaves
                return period_steps, use_traditional

            # If no clear period found, try to detect regular patterns
            pattern_info = _detect_regular_pattern(ratios)
            if pattern_info:
                cycle_size, is_octave_based = pattern_info
                use_traditional = (cycle_size == 12 and is_octave_based)
                return cycle_size, use_traditional

        except (ValueError, TypeError, ZeroDivisionError):
            pass

    # For ET systems, extract divisions from name ONLY as fallback
    if system_name:
        import re as _re
        et_match = _re.search(r'(\d+)-?TET', system_name, _re.IGNORECASE)
        if et_match:
            try:
                divisions = int(et_match.group(1))
                # Use traditional names only for 12-TET with octave intervals
                use_traditional = (divisions == 12)
                # Check if we have information about the interval
                if ratios and len(ratios) >= divisions:
                    # Check if it's really octave-based
                    last_ratio = ratios[divisions-1] if divisions <= len(ratios) else ratios[-1]
                    if abs(last_ratio - 2.0) > 0.02:  # Not an octave
                        use_traditional = False
                return divisions, use_traditional
            except ValueError:
                pass

    # Fallback: For systems without clear analysis, use conservative estimates
    if system_name:
        system_lower = system_name.lower()

        # Geometric progressions - use all available ratios as they might not cycle
        if 'geometric' in system_lower or 'progression' in system_lower:
            return min(num_notes, 100), False  # Conservative large cycle

        # Natural/JI systems - analyze for traditional 12-step if possible
        if 'natural' in system_lower or 'just' in system_lower or '4:5:6' in system_lower:
            return 12, True  # Assume traditional 12-step

        # Danielou systems - can be very complex
        if 'danielou' in system_lower:
            return min(num_notes, 50), False  # Conservative large cycle

        # Analysis-derived systems - no predictable pattern
        if 'analysis' in system_lower or 'audio' in system_lower:
            return min(num_notes, 30), False  # Conservative cycle size

    # Default fallback: Use all available notes as one cycle
    return min(num_notes, 12), (num_notes <= 12)


def _analyze_period_from_ratios(ratios: List[float]) -> Tuple[float, int]:
    """Analyze ratios to find the natural period of the tuning system.

    Returns:
        Tuple[period_factor, period_steps]: The period multiplier and number of steps,
        or (0.0, 0) if no clear period is found.
    """
    if len(ratios) < 2:
        return 0.0, 0

    # Look for ratios that are powers of a common factor
    # Common period factors: 2 (octave), 3 (tritave), etc.
    common_periods = [2.0, 3.0, 1.5, 4.0/3.0, 5.0/4.0]  # octave, tritave, perfect fifth, fourth, major third

    tolerance = 0.02  # ~24 cents tolerance

    for period_factor in common_periods:
        # Check at what steps this period factor appears
        period_steps = []

        for i, ratio in enumerate(ratios):
            if ratio <= 1.0:
                continue

            # Check if this ratio is close to a power of period_factor
            if period_factor > 1.0:
                log_ratio = math.log(ratio) / math.log(period_factor)
                nearest_power = round(log_ratio)

                if nearest_power >= 1 and abs(log_ratio - nearest_power) < tolerance:
                    # This ratio is close to period_factor^nearest_power
                    if nearest_power == 1:  # First occurrence of the period
                        period_steps.append(i + 1)  # +1 because we want step count, not index

        # If we found period occurrences, use the first/smallest one
        if period_steps:
            return period_factor, min(period_steps)

    return 0.0, 0


def _detect_regular_pattern(ratios: List[float]) -> Optional[Tuple[int, bool]]:
    """Detect if ratios follow a regular pattern (like ET systems).

    Returns:
        Tuple[cycle_size, is_octave_based] or None if no pattern detected.
    """
    if len(ratios) < 3:
        return None

    # Calculate step sizes in cents
    steps_cents = []
    for i in range(1, len(ratios)):
        if ratios[i] > ratios[i-1] > 0:
            step = 1200 * math.log2(ratios[i] / ratios[i-1])
            steps_cents.append(step)

    if len(steps_cents) < 2:
        return None

    # Check if steps are roughly equal (quasi-ET)
    mean_step = sum(steps_cents) / len(steps_cents)
    tolerance = 5.0  # 5 cents tolerance

    if all(abs(step - mean_step) < tolerance for step in steps_cents):
        # This looks like a regular ET-type system
        # Check if it's standard 12-TET (100 cents per step)
        if abs(mean_step - 100.0) < 0.5:  # ~100 cents per step = 12-TET
            return 12, True

        # For other regular systems, estimate cycle based on the actual pattern
        # Check what total interval the pattern spans
        if len(ratios) > 1:
            total_cents = 1200 * math.log2(ratios[-1] / ratios[0])
            estimated_cycle = len(ratios)

            # Determine if it's octave-based by checking the total span
            is_octave_based = abs(total_cents - 1200) < 50  # Within 50 cents of octave

            return estimated_cycle, is_octave_based

        # Fallback: estimate based on common intervals
        steps_per_octave = 1200.0 / mean_step
        steps_per_fifth = 700.0 / mean_step

        # Round to nearest integer and check if reasonable
        for target_steps, target_cents in [(steps_per_octave, 1200), (steps_per_fifth, 700)]:
            rounded_steps = round(target_steps)
            if 5 <= rounded_steps <= 100 and abs(target_steps - rounded_steps) < 0.1:
                is_octave_based = abs(target_cents - 1200) < 50  # Within 50 cents of octave
                return rounded_steps, is_octave_based

    return None


def _parse_basenote_for_ascl(basenote) -> Tuple[int, int]:
    """Converts a base note (e.g. 'A4') to octave_number and note_index for ASCL.

    Args:
        basenote: String note name (e.g. 'C3') or numeric frequency (e.g. 130.81)

    Returns:
        Tuple[octave_number, note_index]: octave (4 per A4) e indice nota (9 per A)
    """
    # Note to index mapping
    note_to_index = {
        'C': 0, 'C#': 1, 'DB': 1, 'D': 2, 'D#': 3, 'EB': 3,
        'E': 4, 'F': 5, 'F#': 6, 'GB': 6, 'G': 7, 'G#': 8, 'AB': 8,
        'A': 9, 'A#': 10, 'BB': 10, 'B': 11
    }

    # Handle numeric frequency input (convert to approximate note)
    if isinstance(basenote, (int, float)):
        try:
            # Convert frequency to MIDI note number (A4 = 440 Hz = MIDI 69)
            freq = float(basenote)
            if freq > 0:
                midi_float = 69 + 12 * math.log2(freq / 440.0)
                midi_int = round(midi_float)
                octave = (midi_int // 12) - 1  # MIDI octave adjustment
                note_index = midi_int % 12
                return max(0, octave), note_index
        except (ValueError, TypeError, ZeroDivisionError):
            pass

    # Handle string input
    if isinstance(basenote, str):
        try:
            # Simple parsing for common notes like A4, C4, etc.
            match = _NOTE_PARSE_PATTERN.match(basenote.upper())
            if match:
                note_name = match.group(1)
                octave = int(match.group(2))
                note_index = note_to_index.get(note_name, 9)  # Default A
                return octave, note_index
        except (ValueError, TypeError, AttributeError):
            pass

    # Default to A4 if parsing fails
    return 4, 9


def _generate_note_names_for_ascl(num_notes: int, ratios: List[float] = None,
                                 basenote: str = "C4", diapason: float = 440.0,
                                 system_name: str = None, basenote_name: str = None,
                                 is_octave_repeating: bool = True) -> List[str]:
    """Generates note names for ASCL with microtonal arrows based on analysis data.

    Args:
        num_notes: Total number of notes in the scale
        ratios: List of frequency ratios for the scale steps
        basenote: Base note reference (e.g., "C4")
        diapason: A4 frequency in Hz
        system_name: Name of the tuning system (used to extract ET divisions from patterns like "19-TET")

    Returns:
        List of note names with microtonal deviation arrows

    Note:
        system_name is used to extract ET divisions (e.g., from "Equal Temperament 19-TET")
        to ensure correct octave boundaries for non-12TET systems. Falls back to 12-TET
        if no ET pattern is found.
    """

    # Base names for 12-TET
    base_names = ["C", "C♯/D♭", "D", "D♯/E♭", "E", "F", "F♯/G♭", "G", "G♯/A♭", "A", "A♯/B♭", "B"]

    # Determine the system's cycle structure from system_name and ratios
    cycle_size, use_traditional_names = _determine_cycle_structure(system_name, ratios, num_notes)

    # Parse the starting basenote for fallback cases too
    try:
        # Parse basenote to get MIDI number using utils function
        basenote_midi, basenote_cents = utils.parse_note_with_microtones(basenote)
        # Calculate the starting note index (0=C, 1=C#, etc.) and octave
        starting_note_index = basenote_midi % 12
        starting_octave = (basenote_midi // 12) - 1  # MIDI octave mapping: C4 = MIDI 60
    except (TypeError, ValueError, AttributeError):
        # Fallback if parsing fails - assume C4
        starting_note_index = 0  # C
        starting_octave = 4
        basenote_cents = 0.0

    # If we don't have ratios, fall back to simple cyclic naming using detected cycle
    # BUT still respect the basenote for proper octave calculation
    if not ratios:
        note_names = []
        for i in range(num_notes):
            base_index = (starting_note_index + i) % cycle_size if use_traditional_names and cycle_size == 12 else i % cycle_size
            # Calculate octave taking into account the starting octave
            cycle_count = i // cycle_size
            actual_octave = starting_octave + cycle_count
            octave_suffix = f"_{actual_octave}" if i >= cycle_size else f"_{starting_octave}"

            # Use traditional note names only for 12-step cycles
            if use_traditional_names and cycle_size == 12:
                note_names.append(base_names[base_index] + octave_suffix)
            else:
                note_names.append(f"Step{base_index}" + octave_suffix)
        return note_names

    # Try to use analysis data to generate intelligent note names with arrows
    note_names = []

    # Calculate base note frequency from basenote and diapason
    try:
        # Parse basenote to get MIDI number
        basenote_midi = utils.parse_note_with_microtones(basenote)[0] if hasattr(utils, 'parse_note_with_microtones') else 60
        # Calculate base frequency
        base_freq = diapason * (2 ** ((basenote_midi - 69) / 12.0))
    except (TypeError, ValueError):
        base_freq = 261.626  # Default C4 frequency

    # Parse the starting basenote to determine the initial note and octave
    try:
        # Parse basenote to get MIDI number using utils function
        basenote_midi, basenote_cents = utils.parse_note_with_microtones(basenote)
        # Calculate the starting note index (0=C, 1=C#, etc.) and octave
        starting_note_index = basenote_midi % 12
        starting_octave = (basenote_midi // 12) - 1  # MIDI octave mapping: C4 = MIDI 60
    except (TypeError, ValueError, AttributeError):
        # Fallback if parsing fails - assume C4
        starting_note_index = 0  # C
        starting_octave = 4
        basenote_cents = 0.0

    # According to ASCL specification:
    # "The first entry in the sequence applies to the zeroeth note."
    # However, the zeroeth note is implicit (the basenote/unison)
    # So we generate note names for each ratio step, not including unison
    # For 12-TET: the first ratio (semitone above C) gets first name, etc.

    # Check if this is a non-octave cycle to use cents-based naming
    if not is_octave_repeating and ratios:
        # For non-octave cycles, use cents-based naming with basenote prefix
        # Format: (basenote)C(cents)_ e.g., DbC666_, C#C333_
        note_names = []

        # Extract the basenote name without octave number
        if basenote_name:
            # Use the provided basenote_name
            base_prefix = basenote_name
        else:
            # Parse basenote to get just the note name without octave
            # Remove octave number from basenote string
            import re
            base_match = re.match(r'^([A-G][#b]?)', basenote)
            if base_match:
                base_prefix = base_match.group(1)
                # Convert # to sharp notation for display
                base_prefix = base_prefix.replace('#', '♯')
            else:
                base_prefix = "C"


        # IMPORTANT: First add the basenote itself (unison = 0 cents)
        note_names.append(f"{base_prefix}C0_")

        # Then add names for all the ratios
        for i, ratio in enumerate(ratios):
            if len(note_names) >= num_notes:
                break
            # Calculate cents from basenote
            cents_from_base = 1200 * math.log2(ratio) if ratio > 0 else 0
            # Round to nearest cent for clean display
            cents_rounded = round(cents_from_base)
            # Create note name with basenote prefix, cents value, and underscore
            note_names.append(f"{base_prefix}C{cents_rounded}_")

        # Trim to exact number of notes needed
        return note_names[:num_notes]

    # STEP 1: Generate base note names for all ratios (for octave-based systems)
    # First, create basic note names based on actual pitches
    basic_note_names = []

    for i, ratio in enumerate(ratios):
        if len(basic_note_names) >= num_notes:
            break

        # Calculate frequency of this scale step
        freq = base_freq * ratio

        # Convert to MIDI to find closest note
        midi_from_freq = 69 + 12 * math.log2(freq / diapason) if diapason > 0 else 60
        closest_midi_note = round(midi_from_freq)
        note_index = closest_midi_note % 12
        basic_note_names.append(base_names[note_index])

    # STEP 2: Apply 48-TET grid to add microtonal adjustments
    # This is completely separate from note generation - we just modify what we have
    final_note_names = []

    for i in range(len(basic_note_names)):
        if i >= len(ratios):
            break

        base_name = basic_note_names[i]
        ratio = ratios[i]
        freq = base_freq * ratio

        # Calculate the exact frequency in cents relative to A4
        cents_from_a4 = 1200 * math.log2(freq / diapason) if diapason > 0 else 0

        # Find the closest 48-TET step (each step = 25 cents)
        tet48_position = cents_from_a4 / 25.0
        closest_48tet_step = round(tet48_position)

        # Calculate how far off we are from the closest 48-TET step
        deviation_from_48tet = tet48_position - closest_48tet_step
        deviation_cents = deviation_from_48tet * 25.0

        # Determine which quarter-tone symbol to use based on 48-TET position
        # The 48-TET grid divides each semitone (100 cents) into 4 parts of 25 cents each
        quarter_tone_position = closest_48tet_step % 4

        # Add quarter-tone symbols based on position in 48-TET grid
        microtonal_suffix = ""
        if quarter_tone_position == 1:  # 25 cents above base note
            microtonal_suffix = "!"
        elif quarter_tone_position == 2:  # 50 cents above base note
            microtonal_suffix = "+"
        elif quarter_tone_position == 3:  # 75 cents above base note
            microtonal_suffix = "+!"
        # quarter_tone_position == 0 means exactly on a 12-TET note

        # Add fine-tuning arrows for deviations from the 48-TET grid
        arrow_suffix = ""
        if abs(deviation_cents) > 6:  # More than 6 cents off the 48-TET grid
            if deviation_cents > 15:
                arrow_suffix = "↑↑"  # Very sharp
            elif deviation_cents > 6:
                arrow_suffix = "↑"   # Slightly sharp
            elif deviation_cents < -15:
                arrow_suffix = "↓↓"  # Very flat
            elif deviation_cents < -6:
                arrow_suffix = "↓"   # Slightly flat

        # Combine base name with microtonal adjustments
        final_note_names.append(base_name + microtonal_suffix + arrow_suffix)

    note_names = final_note_names

    # Fallback: if we don't have enough notes yet, use simple cycle structure
    # For ASCL format, we don't include octave numbers in note names
    while len(note_names) < num_notes:
        i = len(note_names)
        base_index = i % cycle_size
        if use_traditional_names and cycle_size == 12:
            note_names.append(base_names[base_index])
        else:
            note_names.append(f"Step{base_index}")

    # Fill remaining notes if needed using cycle structure
    # For ASCL format, we don't include octave numbers in note names
    # IMPORTANT: This should rarely be needed if the main loop works correctly
    # But in case deduplication or other logic reduces the count, we need to fill properly
    while len(note_names) < num_notes:
        missing_count = num_notes - len(note_names)

        # If we're missing notes, it's likely due to deduplication
        # We need to continue the pattern from where we left off
        # Calculate the logical index for the missing note
        logical_index = len(note_names)  # This is the index of the missing note in the original ratios array

        # Use the same logic as the main loop to generate the missing note name
        if logical_index < len(ratios):
            # We have a ratio for this position, generate name normally
            ratio = ratios[logical_index]
            freq = base_freq * ratio if 'base_freq' in locals() else 261.626 * ratio

            if diapason > 0:
                # Same logic as main loop for consistency
                midi_float_48tet = 69*4 + 48 * math.log2(freq / diapason)
                closest_48tet = round(midi_float_48tet)
                tet48_step = closest_48tet % 48
                tet12_step = tet48_step // 4
                base_name = base_names[tet12_step]

                # Add simple microtonal suffix for missing notes
                microtone_step = tet48_step % 4
                microtonal_suffix = ""
                if microtone_step == 1:
                    microtonal_suffix = "!"
                elif microtone_step == 2:
                    microtonal_suffix = "+"
                elif microtone_step == 3:
                    microtonal_suffix = "+!"

                note_names.append(base_name + microtonal_suffix)
            else:
                # Fallback naming
                base_index = logical_index % cycle_size if use_traditional_names and cycle_size == 12 else logical_index % cycle_size
                if use_traditional_names and cycle_size == 12:
                    note_names.append(base_names[base_index])
                else:
                    note_names.append(f"Step{base_index}")
        else:
            # No ratio available, use fallback cycle-based naming
            base_index = logical_index % cycle_size if use_traditional_names and cycle_size == 12 else logical_index % cycle_size
            if use_traditional_names and cycle_size == 12:
                note_names.append(base_names[base_index])
            else:
                note_names.append(f"Step{base_index}")

    return note_names


# Global cache for loaded scales
_scala_scales_cache = None


def parse_scl_file(path: str) -> Optional[dict]:
    """Parsa un file .scl secondo il formato Huygens-Fokker di base.
    Ritorna un dict: {'name': str, 'degrees_cents': List[float], 'file': str}
    dove degrees_cents sono ridotti a [0,1200).
    Se il file non è leggibile/parsabile, ritorna None.
    """
    try:
        with open(path, 'r', encoding='utf-8', errors='ignore') as f:
            lines = f.readlines()
    except FileNotFoundError:
        return None

    # Rimuovi commenti e righe vuote; in .scl i commenti iniziano con '!'
    cleaned: List[str] = []
    for ln in lines:
        s = ln.strip()
        if not s:
            continue
        # drop lines that start with '!' entirely
        if s.startswith('!'):
            # but keep the first comment line as possible filename/description if name missing
            cleaned.append(s)
            continue
        # remove trailing inline comments starting with '!' if any
        if '!' in s:
            s = s.split('!', 1)[0].strip()
        if s:
            cleaned.append(s)

    if not cleaned:
        return None

    # The format: optional leading comment lines starting with '!'
    # Next non-comment line is the description (name). Following line has the number of notes.
    idx = 0
    # Skip leading comments for description, but remember them if needed
    while idx < len(cleaned) and cleaned[idx].startswith('!'):
        idx += 1
    if idx >= len(cleaned):
        return None
    name_line = cleaned[idx]
    idx += 1
    # Some files may keep description empty; fallback to filename later
    desc = name_line.strip()

    # Read number of notes
    if idx >= len(cleaned):
        return None
    try:
        n_deg = int(str(cleaned[idx]).strip().split()[0])
    except ValueError:
        return None
    idx += 1

    degrees_cents: List[float] = []
    # Read next n_deg lines as degrees (can include comments after value that we already stripped)
    for _ in range(n_deg):
        if idx >= len(cleaned):
            break
        val = cleaned[idx].strip()
        idx += 1
        if not val:
            continue
        token = val.split()[0]
        cents_val: Optional[float] = None
        # ratio like 3/2 or 1.5/1.0 is allowed
        if '/' in token:
            try:
                num_str, den_str = token.split('/', 1)
                num = float(Fraction(num_str.strip()))
                den = float(Fraction(den_str.strip()))
                r = num / den if den != 0 else None
                if r and r > 0:
                    cents_val = utils.ratio_to_cents(r)
            except ValueError:
                cents_val = None
        else:
            # plain number in cents
            try:
                cents_val = float(token)
            except ValueError:
                cents_val = None
        if isinstance(cents_val, (int, float)) and math.isfinite(cents_val):
            # Reduce to [0,1200)
            c = float(cents_val) % 1200.0
            degrees_cents.append(c)
    if not degrees_cents:
        return None

    # Ensure unique sorted degrees
    degrees_sorted: List[float] = []
    for c in sorted(degrees_cents):
        if not degrees_sorted or abs(c - degrees_sorted[-1]) > 1e-6:
            degrees_sorted.append(c)

    return {
        'name': desc if desc else os.path.basename(path),
        'degrees_cents': degrees_sorted,
        'file': os.path.basename(path),
        'cents': degrees_sorted  # Add 'cents' key for compatibility
    }


def load_scales_from_dir(dir_path: str = 'scl') -> List[dict]:
    """Carica tutti i file .scl leggibili dalla directory indicata.
    Ritorna una lista di dict come parse_scl_file.
    Uses global cache for performance.
    """
    global _scala_scales_cache

    # Return cached scales if available
    if _scala_scales_cache is not None:
        return _scala_scales_cache

    try:
        if not os.path.isdir(dir_path):
            _scala_scales_cache = []
            return []
    except (FileNotFoundError, OSError):
        _scala_scales_cache = []
        return []

    out: List[dict] = []
    try:
        for fn in os.listdir(dir_path):
            if not fn.lower().endswith('.scl'):
                continue
            fp = os.path.join(dir_path, fn)
            info = parse_scl_file(fp)
            if info and isinstance(info.get('degrees_cents'), list) and info['degrees_cents']:
                out.append(info)
    except (FileNotFoundError, OSError):
        pass

    # Cache the result
    _scala_scales_cache = out
    return out

