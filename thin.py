#!/usr/bin/env python3
# -*- coding: utf-8 -*-
# ruff: noqa: BLE001
# pylint: disable=broad-exception-caught
"""
THIN – Sistemi di intonazione musicale / Musical intonation systems (IT/EN)
Copyright (c) 2025 Luca Bimbi
Distribuito secondo la licenza MIT - vedi il file LICENSE per i dettagli /
Distributed under the MIT license – see the LICENSE file for details

Nome programma: THIN – Sistemi di intonazione musicale / Program name: THIN – Musical intonation systems
Versione / Version: PHI
Autore / Author: LUCA BIMBI
Data / Date: 2025-09-09
"""

import argparse
import sys
import math
import re
import os
import shutil
import threading
import time
from fractions import Fraction
from typing import List, Tuple, Optional, Union, Iterable, Any


# Metadati di modulo / Module metadata (IT/EN)
__program_name__ = "THIN"  # Nuovo nome in onore di Walter Branchi / New name in honor of Walter Branchi
__version__ = "PHI"  # Version string literal 'PHI'
__author__ = "LUCA BIMBI"
__date__ = "2025-09-09"  # La data distingue le release / The date distinguishes releases
__license__ = "MIT"  # Vedi file LICENSE / See LICENSE file

# Costanti / Constants (IT/EN)
ELLIS_CONVERSION_FACTOR = 1200 / math.log(2)
RATIO_EPS = 1e-9
DEFAULT_DIAPASON = 440.0
DEFAULT_BASEKEY = 60
DEFAULT_OCTAVE = 2.0
MAX_HARMONIC_HZ = 10000.0
MIN_SUBHARMONIC_HZ = 16.0
PROXIMITY_THRESHOLD_HZ = 17.0
MIDI_MIN = 0
MIDI_MAX = 127
MIDI_A4 = 69
SEMITONES_PER_OCTAVE = 12

# Costanti colonne Excel / Excel column constants (IT/EN)
CUSTOM_COLUMN = "D"
HARM_COLUMN = "E"
SUB_COLUMN = "G"
TET_COLUMN = "I"
F0_COLUMN = "L"
FORMANT_COLUMN = "M"

# Localizzazione semplice / Simple localization (IT/EN)
_LANG = "it"

# Pattern per parsing file .csd / Regex pattern for .csd parsing (IT/EN)
PATTERN = re.compile(r"\bf\s*(\d+)\b")

# Tipo per valori numerici (int, float o Fraction) / Numeric type alias (int, float or Fraction) (IT/EN)
Numeric = Union[int, float, Fraction]

def L(it_msg: str, en_msg: str) -> str:
    """Restituisce il messaggio nella lingua selezionata / Return message in selected language."""
    return it_msg if _LANG == "it" else en_msg


def _detect_lang_from_argv(argv: Optional[List[str]] = None) -> str:
    """Pre-scan di argv per estrarre --lang prima del parsing argparse."""
    try:
        av = list(argv if argv is not None else sys.argv[1:])
    except Exception:
        av = []
    # Support both "--lang it" and "--lang=it"
    for i, tok in enumerate(av):
        if tok == "--lang" and i + 1 < len(av):
            lang_val = av[i + 1].lower()
            if lang_val in ("it", "en"):
                return lang_val
        elif tok.startswith("--lang="):
            lang_val2 = tok.split("=", 1)[1].strip().lower()
            if lang_val2 in ("it", "en"):
                return lang_val2
    return "it"


# Terminal styling and screen utils
def _supports_ansi() -> bool:
    try:
        return sys.stdout.isatty()
    except Exception:
        return False

class Style:
    # Initialized for dark theme by default (can be remapped by apply_theme)
    if _supports_ansi():
        RESET = "\033[0m"
        BOLD = "\033[1m"
        DIM = "\033[2m"
        UNDER = "\033[4m"
        FG_CYAN = "\033[36m"
        FG_GREEN = "\033[32m"
        FG_YELLOW = "\033[33m"
        FG_MAGENTA = "\033[35m"
    else:
        RESET = ""
        BOLD = ""
        DIM = ""
        UNDER = ""
        FG_CYAN = ""
        FG_GREEN = ""
        FG_YELLOW = ""
        FG_MAGENTA = ""

# Theme management: auto/dark/light
_DEF_COLORS = {
    'dark': {
        'FG_CYAN': "\033[36m",
        'FG_GREEN': "\033[32m",
        'FG_YELLOW': "\033[33m",
        'FG_MAGENTA': "\033[35m",
    },
    'light': {
        # On light backgrounds, prefer higher-contrast hues
        'FG_CYAN': "\033[34m",     # blue
        'FG_GREEN': "\033[32m",
        'FG_YELLOW': "\033[31m",  # red instead of yellow (poor contrast on white)
        'FG_MAGENTA': "\033[35m",
    }
}

_current_theme = 'dark'

def _detect_terminal_background() -> str:
    """Heuristic detection of terminal background.
    Returns 'light' or 'dark'. Defaults to 'dark' if unknown.
    """
    # Allow explicit env override
    env_theme = os.environ.get('TERM_BG') or os.environ.get('TERMINAL_BACKGROUND')
    if env_theme:
        v = env_theme.strip().lower()
        if v in ('light', 'dark'):
            return v
    # COLORFGBG format like "15;0" or "0;15"; background is usually the last
    cfg = os.environ.get('COLORFGBG')
    if cfg:
        try:
            parts = [int(p) for p in re.split(r"[^0-9]+", cfg) if p.isdigit()]
            if parts:
                bg = parts[-1]
                # Common light color indices: >= 7 (light gray..white=15)
                return 'light' if bg >= 7 else 'dark'
        except Exception:
            pass
    # Could add more heuristics per TERM_PROGRAM if needed
    return 'dark'


def apply_theme(theme: str) -> None:
    """Apply theme by remapping Style colors. Accepts 'dark' or 'light'."""
    global _current_theme
    if not _supports_ansi():
        # No-op if ANSI unsupported
        _current_theme = theme
        return
    t = 'light' if str(theme).lower().startswith('light') else 'dark'
    cmap = _DEF_COLORS.get(t, _DEF_COLORS['dark'])
    try:
        Style.FG_CYAN = cmap['FG_CYAN']
        Style.FG_GREEN = cmap['FG_GREEN']
        Style.FG_YELLOW = cmap['FG_YELLOW']
        Style.FG_MAGENTA = cmap['FG_MAGENTA']
        _current_theme = t
    except Exception:
        _current_theme = 'dark'


# Terminal helpers to avoid redundancy and improve robustness
# Returns (cols, rows) with sane fallbacks
def _term_size(default_cols: int = 80, default_rows: int = 24) -> Tuple[int, int]:
    try:
        size = shutil.get_terminal_size(fallback=(int(default_cols), int(default_rows)))
        cols = int(size.columns)
        rows = int(size.lines)
        return (max(1, cols), max(1, rows))
    except Exception:
        return (int(default_cols), int(default_rows))

# Returns only columns with fallback
def _term_cols(default_cols: int = 80) -> int:
    cols, _rows = _term_size(default_cols, 24)
    return cols

# Clear current terminal line up to given columns
def _clear_line(cols: int) -> None:
    try:
        print("\r" + (" " * int(cols)) + "\r", end="")
    except Exception:
        pass


def _detect_theme_from_argv(argv: Optional[List[str]] = None) -> str:
    """Pre-scan argv to extract --theme before argparse. Returns 'auto'|'dark'|'light'."""
    try:
        av = list(argv if argv is not None else sys.argv[1:])
    except Exception:
        av = []
    for i, tok in enumerate(av):
        if tok == "--theme" and i + 1 < len(av):
            val = av[i + 1].strip().lower()
            if val in ("auto", "dark", "light"):
                return val
        elif tok.startswith("--theme="):
            val2 = tok.split("=", 1)[1].strip().lower()
            if val2 in ("auto", "dark", "light"):
                return val2
    return "auto"


def clear_screen() -> None:
    """Clear terminal screen at program start."""
    try:
        if _supports_ansi():
            print("\033[2J\033[H", end="")
        else:
            # Fallback: attempt platform clear
            os.system("cls" if os.name == "nt" else "clear")
    except Exception:
        pass


def stylize_help(help_text: str) -> str:
    """Add basic ANSI styling to help text (section headers, options)."""
    if not _supports_ansi():
        return help_text
    lines = help_text.splitlines()
    out = []
    for s in lines:
        st = s
        # Bold section headers that end with ':' and are not indented much
        if st.strip().endswith(":") and (len(st) - len(st.lstrip())) <= 2:
            st = f"{Style.BOLD}{Style.FG_CYAN}{st}{Style.RESET}"
        # Colorize option flags starting with two spaces then '-'
        elif st.startswith("  -") or st.startswith("  --"):
            # split flags and help by two spaces after flags area
            parts = st.split("  ", 1)
            if parts:
                flags = parts[0]
                rest = parts[1] if len(parts) > 1 else ""
                st = f"{Style.FG_GREEN}{Style.BOLD}{flags}{Style.RESET}  {rest}"
        # Emphasize "Esempi / Examples" line
        if st.lower().lstrip().startswith(("usage:", "uso:", "examples", "esempi")):
            st = f"{Style.BOLD}{Style.FG_MAGENTA}{st}{Style.RESET}"
        out.append(st)
    return "\n".join(out)


def print_banner() -> None:
    """Stampa sempre le info di programma: nome, versione, data, autore, licenza."""
    # Etichette localizzate
    lbl_ver = L("Versione", "Version")
    lbl_date = L("Rilascio", "Release")
    lbl_auth = L("Autore", "Author")
    lbl_lic = L("Licenza", "License")
    title = f"{Style.BOLD}{Style.FG_CYAN}{__program_name__}{Style.RESET}"
    info_line = f"{title}  |  {lbl_ver}: {__version__}  |  {lbl_date}: {__date__}  |  {lbl_auth}: {__author__}  |  {lbl_lic}: {__license__}"
    print(info_line)
    # Copyright / License line (IT/EN)
    copyright_line = L(
        "Copyright (c) 2025 Luca Bimbi. Distribuito secondo la licenza MIT. \nVedi il file LICENSE per i dettagli.",
        "Copyright (c) 2025 Luca Bimbi. Distributed under the MIT License. See the LICENSE file for details."
    )
    print(copyright_line)

def apply_cents(freq_hz: float, cents: float) -> float:
    """Applica offset in cents a una frequenza / Apply cents offset to a frequency."""
    return float(freq_hz) * (2.0 ** (float(cents) / 1200.0))

def is_fraction_type(value: Any) -> bool:
    """Verifica se value è una frazione."""
    return isinstance(value, Fraction)


def fraction_to_cents(ratio: Fraction) -> int:
    """Converte un rapporto razionale in cents."""
    if ratio.denominator == 0:
        raise ValueError(L("Denominatore zero nella frazione", "Zero denominator in fraction"))
    decimal = math.log(ratio.numerator / ratio.denominator)
    return round(decimal * ELLIS_CONVERSION_FACTOR)


def cents_to_fraction(cents: float) -> Fraction:
    """Converte cents in rapporto razionale approssimato."""
    return Fraction(math.exp(cents / ELLIS_CONVERSION_FACTOR)).limit_denominator(10000)


def reduce_to_octave(value: Numeric) -> Numeric:
    """Riduce un rapporto nell'ambito di un'ottava [1, 2)."""
    if isinstance(value, Fraction):
        two = Fraction(2, 1)
        while value >= two:
            value /= two
        while value < 1:
            value *= two
    else:
        value = float(value)
        while value >= 2.0:
            value /= 2.0
        while value < 1.0:
            value *= 2.0
    return value


def reduce_to_interval(value: Numeric, interval: Numeric) -> Numeric:
    """Riduce un rapporto nell'intervallo [1, interval)."""
    try:
        interval_num = float(interval)
    except (TypeError, ValueError):
        return value

    if not math.isfinite(interval_num) or interval_num <= 1.0:
        return value

    if isinstance(value, Fraction) and isinstance(interval, Fraction):
        one = Fraction(1, 1)
        while value >= interval:
            value /= interval
        while value < one:
            value *= interval
    else:
        v = float(value)
        i = float(interval)
        while v >= i:
            v /= i
        while v < 1.0:
            v *= i
        return v

    return value


def normalize_ratios(ratios: Iterable[Numeric], reduce_octave: bool = True) -> List[float]:
    """
    Normalizza una lista di rapporti:
    - Opzionalmente riduce nell'ottava [1, 2)
    - Elimina duplicati con tolleranza
    - Ordina in modo crescente
    """
    processed = []
    seen = []

    for r in ratios:
        v = reduce_to_octave(r) if reduce_octave else r
        v_float = float(v)

        # Controlla duplicati con tolleranza / Check duplicates with tolerance (IT/EN)
        if not any(abs(v_float - s) <= RATIO_EPS for s in seen):
            seen.append(v_float)
            processed.append(v_float)

    return sorted(processed)


def repeat_ratios(ratios: List[float], span: int, interval_factor: float) -> List[float]:
    """Ripete la serie di rapporti su un certo ambitus."""
    try:
        s = int(span)
    except (TypeError, ValueError):
        s = 1

    if s <= 1:
        return list(ratios)

    if not isinstance(interval_factor, (int, float)) or interval_factor <= 0:
        return list(ratios)

    out = []
    for k in range(s):
        factor_k = interval_factor ** k
        out.extend(float(r) * factor_k for r in ratios)

    return out


def pow_fraction(fr: Numeric, k: int) -> Numeric:
    """Eleva una Fraction/float ad un esponente intero."""
    k = int(k)
    return fr ** k if isinstance(fr, Fraction) else float(fr) ** k


def build_natural_ratios(a_max: int, b_max: int, reduce_octave: bool = True) -> List[float]:
    """Genera rapporti del sistema naturale (4:5:6)."""
    three_over_two = Fraction(3, 2)
    five_over_four = Fraction(5, 4)
    vals = []

    for a in range(-abs(a_max), abs(a_max) + 1):
        for b in range(-abs(b_max), abs(b_max) + 1):
            r = pow_fraction(three_over_two, a) * pow_fraction(five_over_four, b)
            vals.append(r)

    return normalize_ratios(vals, reduce_octave=reduce_octave)


def build_danielou_ratios(full_grid: bool = False, reduce_octave: bool = True) -> List[float]:
    """Genera rapporti del sistema Danielou."""
    six_over_five = Fraction(6, 5)
    three_over_two = Fraction(3, 2)
    vals = []

    if full_grid:
        # Mappatura basata sugli appunti di Danielou / Mapping based on Danielou's notes (IT/EN)
        series_spec = {
            0: (5, 5),  # Prima serie su 1/1
            1: (3, 4),  # Seconda serie su 6/5
            -1: (4, 3),  # Terza serie su 5/3
            2: (4, 4),  # Quarta serie su 36/25
            -2: (4, 4),  # Quinta serie su 25/18
            3: (4, 0),  # Sesta serie su 216/125
            -3: (0, 4),  # Settima serie su 125/108
        }

        for a, (desc_n, asc_n) in series_spec.items():
            # Quinte discendenti / Descending fifths (IT/EN)
            for b in range(-desc_n, 0):
                r = pow_fraction(six_over_five, a) * pow_fraction(three_over_two, b)
                vals.append(r)
            # Centro serie / Series center (IT/EN)
            vals.append(pow_fraction(six_over_five, a))
            # Quinte ascendenti / Ascending fifths (IT/EN)
            for b in range(1, asc_n + 1):
                r = pow_fraction(six_over_five, a) * pow_fraction(three_over_two, b)
                vals.append(r)
    else:
        # Sottoinsieme dimostrativo / Demonstrative subset (IT/EN)
        vals.append(Fraction(1, 1))
        # Asse delle quinte / Circle (axis) of fifths (IT/EN)
        for b in range(-5, 6):
            vals.append(pow_fraction(three_over_two, b))
        # Terze minori armoniche / Harmonic minor thirds (IT/EN)
        for k in range(1, 4):
            vals.append(pow_fraction(six_over_five, k))
        # Seste maggiori armoniche / Harmonic major sixths (IT/EN)
        five_over_three = Fraction(5, 3)
        for k in range(1, 4):
            vals.append(pow_fraction(five_over_three, k))

    normalized = normalize_ratios(vals, reduce_octave=reduce_octave)

    if full_grid and reduce_octave:
        # Costruisci lista finale di 53 gradi
        base = list(normalized)
        if not base or abs(base[0] - 1.0) > RATIO_EPS:
            base.append(1.0)
            base = sorted(base)

        base_52_unique = []
        for v in base[:52]:
            if not base_52_unique or abs(v - base_52_unique[-1]) > RATIO_EPS:
                base_52_unique.append(v)

        # Estendi se necessario
        if len(base_52_unique) < 52 and len(base) > len(base_52_unique):
            for v in base[len(base_52_unique):]:
                if all(abs(v - u) > RATIO_EPS for u in base_52_unique):
                    base_52_unique.append(v)
                    if len(base_52_unique) >= 52:
                        break

        normalized = base_52_unique[:52] + [2.0]

    return normalized


def danielou_from_exponents(a: int, b: int, c: int, reduce_octave: bool = True) -> List[float]:
    """Calcola un singolo rapporto del sistema Danielou."""
    six_over_five = Fraction(6, 5)
    three_over_two = Fraction(3, 2)
    two = Fraction(2, 1)

    r = (pow_fraction(six_over_five, a) *
         pow_fraction(three_over_two, b) *
         pow_fraction(two, c))

    if reduce_octave:
        r = reduce_to_octave(r)

    return [float(r)]


def ratio_et(index: int, cents: Union[int, Fraction]) -> float:
    """Calcola il rapporto della radice per un temperamento equabile."""
    decimal_number = math.exp(cents / ELLIS_CONVERSION_FACTOR)
    return decimal_number ** (1 / index)


def int_or_fraction(value: str) -> Union[int, Fraction]:
    """Parser per interi o frazioni."""
    try:
        return int(value)
    except ValueError:
        try:
            return Fraction(value)
        except ValueError:
            raise argparse.ArgumentTypeError(
                f"'{value}' non è un intero o frazione valida. / '{value}' is not a valid integer or fraction."
            )


def parse_interval_value(value: str) -> Union[Fraction, float]:
    """Parsa un valore di intervallo per --geometric.

    Regola: un intero senza suffisso è interpretato come cents; frazioni e float sono rapporti.
    Esempi: 700 -> cents; 700c -> cents; 3/2 -> rapporto; 2.0 -> rapporto.
    """
    if value is None:
        return Fraction(2, 1)

    s = str(value).strip().lower()

    # Controlla suffissi cents / Check 'cents' suffixes (IT/EN)
    cents_suffixes = ['cents', 'cent', 'c']
    for suffix in cents_suffixes:
        if s.endswith(suffix):
            num_str = s[:-len(suffix)].strip()
            try:
                cents_val = float(num_str)
            except ValueError:
                raise argparse.ArgumentTypeError(f"Valore cents non valido: '{value}'")
            ratio = cents_to_fraction(cents_val)
            if float(ratio) <= 1.0:
                raise argparse.ArgumentTypeError("L'intervallo in cents deve essere > 0")
            return ratio

    # Prova int o frazione / Try int or fraction (IT/EN)
    try:
        parsed_val = int_or_fraction(s)
        if isinstance(parsed_val, int):
            # Intero puro => cents / Pure integer => cents (IT/EN)
            cents_val = float(parsed_val)
            ratio = cents_to_fraction(cents_val)
            if float(ratio) <= 1.0:
                raise argparse.ArgumentTypeError("L'intervallo in cents deve essere > 0")
        else:
            ratio = parsed_val
            if float(ratio) <= 1.0:
                raise argparse.ArgumentTypeError("L'intervallo (rapporto) deve essere > 1")
        return ratio
    except argparse.ArgumentTypeError:
        # Prova float come rapporto / Try float as ratio (IT/EN)
        try:
            f = float(s)
            if not math.isfinite(f) or f <= 1.0:
                raise argparse.ArgumentTypeError("L'intervallo (rapporto) deve essere > 1")
            return f
        except ValueError:
            raise argparse.ArgumentTypeError(f"Intervallo non valido: '{value}'")


def parse_danielou_tuple(value: str) -> Tuple[int, int, int]:
    """Parser per terne Danielou 'a,b,c'."""
    if value is None:
        return (0, 0, 1)

    s = str(value).strip()

    # Rimuove parentesi esterne / Remove outer parentheses (IT/EN)
    if s and s[0] in '[({' and s[-1] in '])}':
        s = s[1:-1].strip()

    # Normalizza separatori
    for sep in (':', ';'):
        s = s.replace(sep, ',')

    # Split e pulizia
    s = s.strip(',')
    parts = [p.strip() for p in s.split(',') if p.strip()]

    # Caso singolo intero
    if len(parts) == 1:
        try:
            return (int(parts[0]), 0, 1)
        except ValueError:
            raise argparse.ArgumentTypeError("Formato non valido per --danielou / Invalid format for --danielou")

    if len(parts) != 3:
        raise argparse.ArgumentTypeError("Formato non valido per --danielou. Usa 'a,b,c'")

    try:
        a = int(parts[0])
        b = int(parts[1])
        c = int(parts[2])
        return (a, b, c)
    except ValueError:
        raise argparse.ArgumentTypeError("Esponenti non validi per --danielou")


def note_name_or_frequency(value: str) -> Union[float, str]:
    """Parser per nome nota o frequenza."""
    try:
        return float(value)
    except (ValueError, TypeError):
        return str(value)


def convert_note_name_to_midi(note_name: str) -> int:
    """Converte nome nota in valore MIDI. Supporta # e b/B. / Convert note name to MIDI (supports # and b/B)."""
    note_map = {'C': 0, 'D': 2, 'E': 4, 'F': 5, 'G': 7, 'A': 9, 'B': 11}
    s = note_name.strip()
    if not s:
        raise ValueError(L("Nome nota vuoto", "Empty note name"))

    s_up = s.upper()
    note = s_up[0]
    if note not in note_map:
        raise ValueError(L(f"Nome nota non valido: {note_name}", f"Invalid note name: {note_name}"))

    alteration = 0
    idx = 1

    # gestisci # e b/B / handle # and b
    if idx < len(s_up):
        if s_up[idx] == '#':
            alteration += 1
            idx += 1
        elif s_up[idx] in ('B', 'b'):
            alteration -= 1
            idx += 1

    try:
        octave = int(s_up[idx:])
    except (ValueError, IndexError):
        raise ValueError(L(f"Formato ottava non valido in: {note_name}",
                           f"Invalid octave format in: {note_name}"))

    midi_value = (octave + 1) * SEMITONES_PER_OCTAVE + note_map[note] + alteration

    if not (MIDI_MIN <= midi_value <= MIDI_MAX):
        raise ValueError(L(f"Nota MIDI fuori range: {midi_value}",
                           f"MIDI note out of range: {midi_value}"))

    return midi_value


def parse_note_with_microtones(note_name: str) -> Tuple[int, float]:
    """
    Parsing nota con microtoni per basenote e fondamentali.
    Simboli: # (diesis), b/B (bemolle), + (quarto di tono su, +50c), - (quarto giù, -50c),
             ! (ottavo su, +25c), . (ottavo giù, -25c).
    Restituisce (midi_int, cents_offset). / Returns (midi_int, cents_offset).
    """
    note_map = {'C': 0, 'D': 2, 'E': 4, 'F': 5, 'G': 7, 'A': 9, 'B': 11}
    s = note_name.strip()
    if not s:
        raise ValueError(L("Nome nota vuoto", "Empty note name"))
    s_up = s.upper()
    note = s_up[0]
    if note not in note_map:
        raise ValueError(L(f"Nome nota non valida: {note_name}", f"Invalid note name: {note_name}"))
    idx = 1
    semitones = 0
    cents = 0.0

    # Accidenti base / basic accidentals
    while idx < len(s_up) and s_up[idx] in ('#', 'B', 'b', '+', '-', '!', '.'):
        ch = s_up[idx]
        if ch == '#':
            semitones += 1
        elif ch in ('B', 'b'):
            semitones -= 1
        elif ch == '+':
            cents += 50.0
        elif ch == '-':
            cents -= 50.0
        elif ch == '!':
            cents += 25.0
        elif ch == '.':
            cents -= 25.0
        idx += 1

    try:
        octave = int(s_up[idx:])
    except (ValueError, IndexError):
        raise ValueError(L(f"Formato ottava non valido in: {note_name}",
                           f"Invalid octave format in: {note_name}"))

    midi_value = (octave + 1) * SEMITONES_PER_OCTAVE + note_map[note] + semitones
    if not (MIDI_MIN <= midi_value <= MIDI_MAX):
        raise ValueError(L(f"Nota MIDI fuori range: {midi_value}",
                           f"MIDI note out of range: {midi_value}"))

    return midi_value, cents


def convert_midi_to_hz(midi_value: int, diapason_hz: float = DEFAULT_DIAPASON) -> float:
    """Converte valore MIDI in frequenza Hz."""
    return diapason_hz * (2 ** ((float(midi_value) - MIDI_A4) / SEMITONES_PER_OCTAVE))


def midi_to_note_name_12tet(midi_value: int) -> str:
    """Converte un valore MIDI in nome nota 12-TET (con #) con ottava, es. A4, C#3."""
    try:
        m = int(midi_value)
    except Exception:
        return ""
    if m < MIDI_MIN or m > MIDI_MAX:
        m = max(MIDI_MIN, min(MIDI_MAX, m))
    names = ["C", "C#", "D", "D#", "E", "F", "F#", "G", "G#", "A", "A#", "B"]
    name = names[m % SEMITONES_PER_OCTAVE]
    octave = (m // SEMITONES_PER_OCTAVE) - 1
    return f"{name}{octave}"


def file_exists(file_path: str) -> bool:
    """Verifica esistenza file."""
    exists = os.path.exists(file_path)
    it_word = "esistente" if exists else "non esistente"
    en_word = "exists" if exists else "not found"
    print(L(f"File {it_word}: {file_path}", f"File {en_word}: {file_path}"))
    return exists


def parse_file(file_name: str) -> int:
    """Trova il massimo numero di tabella 'f N' nel file."""
    max_num = 0
    try:
        with open(file_name, "r") as file:
            for line in file:
                for match in PATTERN.finditer(line):
                    num = int(match.group(1))
                    max_num = max(max_num, num)
    except FileNotFoundError:
        print(L(f"Errore: file non trovato: {file_name}", f"Error: file not found: {file_name}"))
    return max_num


def ensure_midi_fit(ratios: List[float], basekey: int,
                    prefer_truncate: bool) -> Tuple[List[float], int]:
    """Assicura che i rapporti rientrino nel range MIDI."""
    n = len(ratios)

    try:
        bk = int(basekey)
    except (TypeError, ValueError):
        bk = 0

    if n > 128:
        if not prefer_truncate:
            print(L(f"WARNING: numero di passi ({n}) eccede 128. Verranno mantenuti solo i primi 128.",
                     f"WARNING: number of steps ({n}) exceeds 128. Only the first 128 will be kept."))
            prefer_truncate = True
        n = 128
        ratios = ratios[:n]

    if prefer_truncate:
        bk = max(MIDI_MIN, min(MIDI_MAX, bk))
        max_len = 128 - bk
        if len(ratios) > max_len:
            print(L(f"WARNING: serie eccede il limite MIDI. Troncati {len(ratios) - max_len} passi.",
                    f"WARNING: series exceeds the MIDI limit. Truncated {len(ratios) - max_len} steps."))
            ratios = ratios[:max_len]
        return ratios, bk
    else:
        allowed_min = MIDI_MIN
        allowed_max = MIDI_MAX - (n - 1)
        if allowed_max < allowed_min:
            allowed_max = allowed_min

        bk_eff = bk
        if bk < allowed_min or bk > allowed_max:
            bk_eff = max(allowed_min, min(allowed_max, bk))
            print(L(f"WARNING: basekey adattata da {bk} a {bk_eff} per includere tutti i passi.",
                     f"WARNING: basekey adjusted from {bk} to {bk_eff} to include all steps."))

        return ratios, bk_eff


def write_cpstun_table(output_base: str, ratios: List[float], basekey: int,
                       basefrequency: float, interval_value: Optional[float] = None) -> Tuple[int, bool]:
    """Crea o appende una tabella cpstun in un file Csound .csd."""
    csd_path = f"{output_base}.csd"
    skeleton = (
        "<CsoundSynthesizer>\n"
        "<CsOptions>\n\n</CsOptions>\n"
        "<CsInstruments>\n\n</CsInstruments>\n"
        "<CsScore>\n\n</CsScore>\n"
        "</CsoundSynthesizer>\n"
    )

    existed_before = file_exists(csd_path)

    if not existed_before:
        try:
            with open(csd_path, "w") as f:
                f.write(skeleton)
        except IOError as e:
            print(L(f"Errore creazione file CSD: {e}", f"Error creating CSD file: {e}"))
            return 0, existed_before

    try:
        with open(csd_path, "r", encoding="utf-8") as f:
            content = f.read()
    except IOError as e:
        print(L(f"Errore lettura CSD: {e}", f"Error reading CSD: {e}"))
        return 0, existed_before

    fnum = parse_file(csd_path) + 1

    # Ordina i rapporti / Sort ratios (IT/EN)
    try:
        ratios_sorted = sorted(float(r) for r in ratios)
    except (TypeError, ValueError):
        ratios_sorted = [float(r) for r in ratios]

    # Determina parametri cpstun / Determine cpstun parameters (IT/EN)
    numgrades = len(ratios_sorted)

    if interval_value is not None and isinstance(interval_value, (int, float)):
        interval = 0.0 if float(interval_value) <= 0 else float(interval_value)
    else:
        try:
            rmin = min(ratios_sorted) if ratios_sorted else 1.0
            rmax = max(ratios_sorted) if ratios_sorted else 1.0
            interval = 2.0 if (rmin >= 1.0 - RATIO_EPS and rmax <= 2.0 + RATIO_EPS) else 0.0
        except (TypeError, ValueError):
            interval = 0.0

    # Costruisci lista dati / Build data list (IT/EN)
    data_list = [
        str(numgrades),
        f"{float(interval):.10g}",
        f"{float(basefrequency):.10g}",
        str(int(basekey))
    ]
    data_list.extend(f"{r:.10f}" for r in ratios_sorted)

    size = len(data_list)

    # Costruisci righe / Build lines (IT/EN)
    prefix = f"f {fnum} 0 {size} -2 "
    positions = []
    col = len(prefix)
    for t in data_list:
        positions.append(col)
        col += len(t) + 1

    def build_aligned_comment(label_map: List[Tuple[int, str]]) -> str:
        line = ";"
        base_offset = 1
        for pos_idx, text in label_map:
            target = base_offset + (positions[pos_idx] if 0 <= pos_idx < len(positions) else len(prefix))
            if target < len(line):
                target = len(line) + 1
            line += " " * (target - len(line)) + text
        return line + "\n"

    header_comment = (
            build_aligned_comment([(0, "numgrades"), (2, "basefreq"), (4, "rapporti-di-intonazione .......")]) +
            build_aligned_comment([(1, "interval"), (3, "basekey")]) +
            f"; tabella cpstun generata | basekey={basekey} basefrequency={basefrequency:.6f}Hz\n"
    )
    f_line = prefix + " ".join(data_list) + "\n"

    # Inserisci prima di </CsScore> / Insert before </CsScore> (IT/EN)
    insert_marker = "</CsScore>"
    idx = content.rfind(insert_marker)
    if idx == -1:
        content += f"\n<CsScore>\n{header_comment}{f_line}</CsScore>\n"
    else:
        content = content[:idx] + header_comment + f_line + content[idx:]

    try:
        with open(csd_path, "w", encoding="utf-8") as f:
            f.write(content)
        print(L(f"Tabella cpstun (f {fnum}) salvata in {csd_path}", f"cpstun table (f {fnum}) saved to {csd_path}"))
    except IOError as e:
        print(L(f"Errore scrittura CSD: {e}", f"Error writing CSD: {e}"))

    return fnum, existed_before

def write_tun_file(output_base: str, diapason: float, ratios: List[float], basekey: int,
                   basefrequency: float, tun_integer: bool = False) -> None:
    """Esporta un file .tun (AnaMark TUN) con valori espressi in cents assoluti riferiti a 8.1757989156437073336 Hz.
    Struttura: [Tuning] + 128 righe "note X=Y" (cents assoluti)."""
    # ATTENZIONE! Deve recuperare il valore di diapason o prendere il --diapason da riga di comando
    # e nel convertire il file .TUN deve essere in tal caso [Exact Tuning]
    # basefreq = 8.1757989156437073336 * (diapasonHz/440)
    # Riferimento assoluto AnaMark / AnaMark absolute reference (IT/EN)
    f_ref = 8.1757989156437073336*(diapason/440)

    tun_path = f"{output_base}.tun"

    if diapason != 440:
        lines = ["[Exact Tuning]",
                 f"basefreq={f_ref}"]
    else:

        # logica -> diapason = 440?
        lines = [
        "[Tuning]",
        ]

    def tet_freq(offset_semitones: int) -> float:
        return basefrequency * (2.0 ** (offset_semitones / 12.0))

    # Ordina i rapporti per garantire valori crescenti nel segmento custom / Sort ratios to ensure ascending custom segment (IT/EN)
    try:
        ratios_sorted = sorted(float(r) for r in ratios)
    except (TypeError, ValueError):
        ratios_sorted = [float(r) for r in ratios]

    def note_freq(n: int) -> float:
        if 0 <= basekey <= 127 and basekey <= n < basekey + len(ratios_sorted):
            idx = n - basekey
            if 0 <= idx < len(ratios_sorted):
                return basefrequency * float(ratios_sorted[idx])
            else:
                return tet_freq(n - basekey)
        else:
            return tet_freq(n - basekey)



    for note_idx in range(MIDI_MAX):
        f = note_freq(note_idx)
        if isinstance(f, (int, float)) and f > 0:
            cents = 1200.0 * math.log2(f / f_ref)
        else:
            cents = 0.0 if note_idx == 0 else 0.0
        if tun_integer:
            cents_text = str(int(round(cents)))
        else:
            # Format with two decimals, but if the decimal part is .00 print only the integer
            rounded2 = round(cents, 2)
            s2 = f"{rounded2:.2f}"
            if s2.endswith(".00"):
                cents_text = str(int(round(rounded2)))
            else:
                cents_text = s2
        lines.append(f"Note {note_idx}={cents_text}")

    try:
        with open(tun_path, "w", encoding="utf-8") as f:
            f.write("\n".join(lines) + "\n")
        print(L(f"File .tun salvato in {tun_path}", f".tun file saved to {tun_path}"))
    except IOError as e:
        print(L(f"Errore scrittura .tun: {e}", f"Error writing .tun: {e}"))


def import_tun_file(tun_path: str, basekey: int = DEFAULT_BASEKEY, reduce_octave_out: bool = True) -> Optional[str]:
    """
    Importa un file .tun (AnaMark TUN) e converte i valori in ratios relativi a basekey.
    Salva un file .txt con mappatura 'MIDI -> ratio' per le note disponibili.
    """
    # Riferimento assoluto AnaMark (stesso di write_tun_file) / AnaMark absolute reference (same as write_tun_file) (IT/EN)
    f_ref = 8.1757989156437073336

    cents_map = {}
    try:
        with open(tun_path, "r", encoding="utf-8") as f:
            for line in f:
                s = line.strip()
                if not s or s.startswith(";") or s.startswith("#"):
                    continue
                # Consenti e ignora le sezioni come [Tuning] / Allow and ignore sections like [Tuning] (IT/EN)
                if s.startswith("[") and s.endswith("]"):
                    continue
                m = re.match(r'^(?:[Nn]ote|[Kk]ey)\s+(\d{1,3})\s*=\s*([+-]?\d+(?:\.\d+)?)', s)
                if m:
                    idx = int(m.group(1))
                    if 0 <= idx <= 127:
                        try:
                            cents_val = float(m.group(2))
                            cents_map[idx] = cents_val
                        except Exception:
                            pass
    except IOError as e:
        print(L(f"Errore lettura .tun: {e}", f"Error reading .tun: {e}"))
        return None

    if not cents_map:
        print(L("File .tun vuoto o non riconosciuto", "Empty or unrecognized .tun file"))
        return None

    # Cents -> frequenze assolute / Cents -> absolute frequencies (IT/EN)
    freq_map = {}
    for k, cv in cents_map.items():
        try:
            freq_map[k] = float(f_ref) * (2.0 ** (float(cv) / 1200.0))
        except Exception:
            continue

    # Determina frequenza base / Determine base frequency (IT/EN)
    if basekey in freq_map:
        base_freq = freq_map[basekey]
        base_key_eff = basekey
    else:
        keys_sorted = sorted(freq_map.keys(), key=lambda n: abs(n - basekey))
        if not keys_sorted:
            print(L("Nessuna nota valida nel .tun", "No valid notes in .tun"))
            return None
        base_key_eff = keys_sorted[0]
        base_freq = freq_map[base_key_eff]

    if not isinstance(base_freq, (int, float)) or base_freq <= 0:
        print(L("Frequenza base non valida nel .tun", "Invalid base frequency in .tun"))
        return None

    # Costruisci ratios rispetto a base_key_eff / Build ratios relative to base_key_eff (IT/EN)
    ratios_by_midi = {}
    for n in range(128):
        f = freq_map.get(n)
        if isinstance(f, (int, float)) and f > 0:
            r = float(f) / float(base_freq)
            if reduce_octave_out:
                try:
                    r = float(reduce_to_octave(r))
                except Exception:
                    pass
            ratios_by_midi[n] = float(r)

    # Scrivi file di output / Write output file (IT/EN)
    base_name = os.path.splitext(os.path.basename(tun_path))[0]
    out_path = f"{base_name}_ratios.txt"
    try:
        with open(out_path, "w", encoding="utf-8") as f:
            f.write(f"; Import .tun -> ratios | basekey={base_key_eff}\n")
            f.write("MIDI -> Ratio\n")
            for n in sorted(ratios_by_midi.keys()):
                f.write(f"{n} -> {ratios_by_midi[n]:.10f}\n")
        print(L(f"File ratios salvato in {out_path}", f"Ratios file saved to {out_path}"))
        return out_path
    except IOError as e:
        print(L(f"Errore scrittura ratios: {e}", f"Error writing ratios: {e}"))
        return None


def export_system_tables(output_base: str, ratios: List[float], basekey: int,
                         basenote_hz: float) -> None:
    """Esporta tabelle del sistema generato."""
    # Calcola e ordina / Compute and sort (IT/EN)
    computed = [(basenote_hz * float(r), i, basekey + i, float(r))
                for i, r in enumerate(ratios)]
    computed.sort(key=lambda t: t[0])

    # Export testo / Text export (IT/EN)
    txt_path = f"{output_base}_system.txt"
    headers = ["Step", "MIDI", "Ratio", "Hz"]
    try:
        rows = [[str(i), str(basekey + i), f"{r:.10f}", f"{hz:.6f}"]
                for i, (hz, _, _, r) in enumerate(computed)]

        # Calcola larghezze colonne / Compute column widths (IT/EN)
        widths = [len(h) for h in headers]
        for row in rows:
            for c, val in enumerate(row):
                widths[c] = max(widths[c], len(val))

        def fmt(vals: List[str]) -> str:
            return "  ".join(str(value).ljust(widths[i]) for i, value in enumerate(vals))

        with open(txt_path, "w", encoding="utf-8") as f:
            f.write(fmt(headers) + "\n")
            for row in rows:
                f.write(fmt(row) + "\n")
        print(L(f"Esportato: {txt_path}", f"Exported: {txt_path}"))
    except IOError as e:
        print(L(f"Errore scrittura {txt_path}: {e}", f"Write error {txt_path}: {e}"))

    # Export Excel (opzionale)
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill

        wb = Workbook()
        ws = wb.active
        ws.title = "System"
        ws.append(headers)

        header_fill = PatternFill(start_color="FFDDDDDD", end_color="FFDDDDDD",
                                  fill_type="solid")
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.fill = header_fill

        # Inserisci ancora di riferimento Base_Hz per formule di Ratio
        # Put Base_Hz anchor for Ratio formulas
        ws.cell(row=1, column=6, value="Base_Hz")
        ws.cell(row=2, column=6, value=float(basenote_hz))

        # Popola righe: Ratio come formula basata su Hz/Base_Hz
        for i, (hz, _, _, _r) in enumerate(computed):
            row_idx = i + 2  # a partire dalla riga 2 (dopo header)
            # Step, MIDI
            ws.cell(row=row_idx, column=1, value=i)
            ws.cell(row=row_idx, column=2, value=basekey + i)
            # Ratio =IFERROR(Dn/$F$2, "")
            ws.cell(row=row_idx, column=3, value=f"=IFERROR(D{row_idx}/$F$2,\"\")")
            # Hz
            ws.cell(row=row_idx, column=4, value=float(hz))

        xlsx_path = f"{output_base}_system.xlsx"
        wb.save(xlsx_path)
        print(L(f"Esportato: {xlsx_path}", f"Exported: {xlsx_path}"))
    except ImportError:
        print(L("openpyxl non installato: export Excel saltato", "openpyxl not installed: Excel export skipped"))
    except Exception as e:
        print(L(f"Errore export Excel: {e}", f"Excel export error: {e}"))


def analyze_audio(audio_path: str,
                   method: str = "lpc",
                   frame_size: int = 1024,
                   hop_size: int = 512,
                   lpc_order: int = 12,
                   window_type: str = "hamming",
                   use_hq: bool = False) -> Optional[dict]:
    """Analizza un file WAV per F0 e formanti usando librosa.
    Ritorna un dict: { 'f0_hz': float|None, 'formants': [(freq_hz, rel_amp_0_1), ...] }
    Se librosa non è disponibile o l'analisi fallisce, ritorna None.
    """
    try:
        import numpy as _np
        import librosa as _lb
        import warnings as _warnings
        try:
            from scipy.signal import find_peaks as _find_peaks  # optional
        except (ImportError, ModuleNotFoundError):
            _find_peaks = None
    except (ImportError, ModuleNotFoundError) as e:
        print(L(f"librosa non disponibile ({e}): installa 'librosa' per l'analisi audio.",
                f"librosa not available ({e}): install 'librosa' for audio analysis."))
        return None

    try:
        # Load mono audio at native sampling rate
        y, sr = _lb.load(audio_path, sr=None, mono=True)
    except (FileNotFoundError, ValueError, RuntimeError) as e:
        print(L(f"Errore nel caricamento audio: {e}", f"Audio loading error: {e}"))
        return None

    # Prepare window function
    win_type_in = (window_type or "hamming").lower()
    if win_type_in not in ("hamming", "hanning"):
        win_type_in = "hamming"
    win = _lb.filters.get_window("hann" if win_type_in == "hanning" else win_type_in, frame_size, fftbins=True)
    # Initialize optional CREPE timing and mask variables
    time_c = None  # type: ignore[assignment]
    mask = None    # type: ignore[assignment]

    # Pitch estimation: CREPE when use_hq else pYIN/YIN; robust fallbacks
    f0_hz: Optional[float] = None
    f0_vals = _np.array([], dtype=float)
    try:
        fmin = 50.0
        fmax = max(2000.0, sr / 4.0)
        # Auto-adapt fmin so at least two periods of fmin fit into the frame
        required_min_fmin = (2.0 * sr) / float(frame_size) if frame_size and sr else fmin
        fmin_eff = fmin
        try:
            fmin_eff = max(float(fmin), float(required_min_fmin) + 1e-6)
        except Exception:
            fmin_eff = fmin
        if fmin_eff >= fmax:
            fmin_eff = 0.5 * fmax
        if fmin_eff > fmin:
            pass
            #print(L(
            #    f"Adatto automaticamente fmin da {fmin:.3f} a {fmin_eff:.3f} Hz per frame_length={frame_size}",
            #    f"Auto-adapting fmin from {fmin:.3f} to {fmin_eff:.3f} Hz for frame_length={frame_size}"
            #))
        # Use the effective fmin for subsequent pYIN/YIN calls
        fmin = fmin_eff
        if use_hq:
            try:
                import crepe as _cr
                # CREPE expects 16kHz
                y16 = _lb.resample(y, orig_sr=sr, target_sr=16000) if sr != 16000 else y
                time_c, freq_c, conf_c, _act = _cr.predict(y16, 16000, viterbi=True, step_size=10, model_capacity='full', verbose=0)
                # Filter by confidence
                mask = conf_c > 0.5
                f0_vals = _np.asarray(freq_c[mask], dtype=float)
            except (ImportError, ModuleNotFoundError, RuntimeError, ValueError):
                # Fallback to pYIN/YIN
                with _warnings.catch_warnings():
                    _warnings.simplefilter("ignore")
                    try:
                        f0_series, _, _ = _lb.pyin(y, fmin=fmin, fmax=fmax, frame_length=frame_size, hop_length=hop_size, sr=sr)
                        f0_vals = _np.asarray([v for v in _np.nan_to_num(f0_series, nan=_np.nan) if _np.isfinite(v) and v > 0])
                    except (ValueError, RuntimeError):
                        f0_track = _lb.yin(y, fmin=fmin, fmax=fmax, frame_length=frame_size, hop_length=hop_size, sr=sr)
                        f0_vals = _np.asarray([v for v in f0_track if _np.isfinite(v) and v > 0])
        else:
            with _warnings.catch_warnings():
                _warnings.simplefilter("ignore")
                try:
                    f0_series, _, _ = _lb.pyin(y, fmin=fmin, fmax=fmax, frame_length=frame_size, hop_length=hop_size, sr=sr)
                    f0_vals = _np.asarray([v for v in _np.nan_to_num(f0_series, nan=_np.nan) if _np.isfinite(v) and v > 0])
                except (ValueError, RuntimeError):
                    f0_track = _lb.yin(y, fmin=fmin, fmax=fmax, frame_length=frame_size, hop_length=hop_size, sr=sr)
                    f0_vals = _np.asarray([v for v in f0_track if _np.isfinite(v) and v > 0])
        if f0_vals.size > 0:
            arr = f0_vals[_np.isfinite(f0_vals)]
            if arr.size > 0:
                med = _np.median(arr)
                mad = _np.median(_np.abs(arr - med))
                if mad > 0:
                    sel = _np.abs(arr - med) <= (3.0 * mad)
                    arr_f = arr[sel]
                    if arr_f.size == 0:
                        arr_f = arr
                else:
                    lo = _np.percentile(arr, 5.0)
                    hi = _np.percentile(arr, 95.0)
                    arr_f = arr[(arr >= lo) & (arr <= hi)] if hi > lo else arr
                f0_hz = float(_np.median(arr_f))
    except Exception:
        f0_hz = None

    formant_freqs: list = []
    formant_amps: list = []

    try:
        m = (method or "lpc").lower()
        if m == "lpc":
            # Iterate frames and compute LPC roots
            y_pad = _np.pad(y, (frame_size//2, frame_size//2), mode='reflect') if y.size >= frame_size else _np.pad(y, (0, frame_size - y.size), mode='constant')
            collect_f = []
            for start in range(0, max(1, len(y_pad) - frame_size + 1), hop_size):
                frame = y_pad[start:start+frame_size]
                if frame.size < frame_size:
                    break
                w = frame * win
                # Pre-emphasis can help formant detection
                w = _np.append(w[0], w[1:] - 0.97 * w[:-1])
                try:
                    a = _lb.lpc(w, order=max(2, int(lpc_order)))
                except Exception:
                    continue
                try:
                    roots = _np.roots(a)
                except Exception:
                    continue
                for r in roots:
                    if _np.abs(r) < 1.0 - 1e-6:
                        ang = _np.angle(r)
                        if ang > 0:
                            f = float(ang * sr / (2 * _np.pi))
                            if 90.0 <= f <= sr/2 - 100:
                                collect_f.append(f)
            if collect_f:
                # Cluster nearby frequencies (~70 Hz) and count as amplitude proxy
                freqs_sorted = _np.sort(_np.array(collect_f, dtype=float))
                centers = []
                counts = []
                for f in freqs_sorted:
                    if not centers or abs(f - centers[-1]) > 70.0:
                        centers.append(f)
                        counts.append(1)
                    else:
                        centers[-1] = (centers[-1] * counts[-1] + f) / (counts[-1] + 1)
                        counts[-1] += 1
                maxc = max(counts) if counts else 1
                formant_freqs = list(centers)
                formant_amps = [c / maxc for c in counts]
        elif m == "specenv":
            # Spectral envelope approximation via STFT average and peak picking
            S = _np.abs(_lb.stft(y, n_fft=frame_size, hop_length=hop_size, window=win))
            if S.size == 0:
                return {"f0_hz": f0_hz, "formants": []}
            mag = _np.mean(S, axis=1)
            # Frequency axis in Hz
            freqs_axis = _np.linspace(0.0, sr/2.0, num=mag.shape[0])
            # Peak picking
            if _find_peaks is not None:
                peaks, _ = _find_peaks(mag, distance=max(1, int((200.0/(sr/2.0))*mag.shape[0])))
                # take top 10 by magnitude
                if peaks.size > 10:
                    idx = _np.argsort(mag[peaks])[-10:]
                    peaks = peaks[idx]
            else:
                # Fallback: take largest bins
                peaks = _np.argpartition(mag, -10)[-10:]
            # Map to Hz and normalize amplitudes
            amps = mag[peaks]
            if amps.size > 0:
                amps = amps / float(amps.max())
            fsel = freqs_axis[peaks]
            keep = (fsel >= 150.0) & (fsel <= sr/2.0 - 100.0)
            formant_freqs = list(map(float, fsel[keep]))
            formant_amps = list(map(float, amps[keep]))
        elif m == "sfft":
            # Single long-window FFT analysis
            n_fft = 1
            if y.size > 0:
                n_fft = int(2 ** _np.ceil(_np.log2(max(frame_size*4, y.size))))
            n_fft = max(2048, min(n_fft, 1<<18))  # safety cap
            # Apply window
            w = _lb.filters.get_window("hann", min(len(y), n_fft), fftbins=True)
            y_seg = y[:min(len(y), n_fft)] * w
            spec = _np.abs(_np.fft.rfft(y_seg, n=n_fft))
            freqs_axis = _np.fft.rfftfreq(n_fft, d=1.0/sr)
            if spec.size == 0:
                return {"f0_hz": f0_hz, "formants": []}
            mag = spec
            if _find_peaks is not None:
                peaks, _ = _find_peaks(mag, distance=max(1, int((200.0/(sr/2.0))*mag.shape[0])))
                if peaks.size > 12:
                    idx = _np.argsort(mag[peaks])[-12:]
                    peaks = peaks[idx]
            else:
                peaks = _np.argpartition(mag, -12)[-12:]
            amps = mag[peaks]
            if amps.size > 0:
                amps = amps / float(amps.max())
            fsel = freqs_axis[peaks]
            keep = (fsel >= 150.0) & (fsel <= sr/2.0 - 100.0)
            formant_freqs = list(map(float, fsel[keep]))
            formant_amps = list(map(float, amps[keep]))
        elif m == "cqt":
            # Constant-Q analysis averaged over time
            try:
                fmin = 55.0
                bins_per_octave = 48
                n_bins = int(_np.ceil(bins_per_octave * _np.log2((sr/2.0) / fmin)))
                if n_bins <= 0:
                    return {"f0_hz": f0_hz, "formants": []}
                C = _np.abs(_lb.cqt(y, sr=sr, hop_length=hop_size, fmin=fmin, n_bins=n_bins, bins_per_octave=bins_per_octave))
                if C.size == 0:
                    return {"f0_hz": f0_hz, "formants": []}
                mag = _np.mean(C, axis=1)
                freqs_axis = _lb.cqt_frequencies(n_bins=n_bins, fmin=fmin, bins_per_octave=bins_per_octave)
                if _find_peaks is not None:
                    peaks, _ = _find_peaks(mag, distance=max(1, int(bins_per_octave/2)))
                    if peaks.size > 12:
                        idx = _np.argsort(mag[peaks])[-12:]
                        peaks = peaks[idx]
                else:
                    peaks = _np.argpartition(mag, -12)[-12:]
                amps = mag[peaks]
                if amps.size > 0:
                    amps = amps / float(amps.max())
                fsel = _np.array(freqs_axis)[peaks]
                keep = (fsel >= 150.0) & (fsel <= sr/2.0 - 100.0)
                formant_freqs = list(map(float, fsel[keep]))
                formant_amps = list(map(float, amps[keep]))
            except Exception:
                # Fallback to STFT if CQT fails
                S = _np.abs(_lb.stft(y, n_fft=frame_size, hop_length=hop_size, window=win))
                if S.size == 0:
                    return {"f0_hz": f0_hz, "formants": []}
                mag = _np.mean(S, axis=1)
                freqs_axis = _np.linspace(0.0, sr/2.0, num=mag.shape[0])
                peaks = _np.argpartition(mag, -10)[-10:]
                amps = mag[peaks]
                if amps.size > 0:
                    amps = amps / float(amps.max())
                fsel = freqs_axis[peaks]
                keep = (fsel >= 150.0) & (fsel <= sr/2.0 - 100.0)
                formant_freqs = list(map(float, fsel[keep]))
                formant_amps = list(map(float, amps[keep]))
        else:
            # Default to STFT specenv if unknown method
            S = _np.abs(_lb.stft(y, n_fft=frame_size, hop_length=hop_size, window=win))
            if S.size == 0:
                return {"f0_hz": f0_hz, "formants": []}
            mag = _np.mean(S, axis=1)
            freqs_axis = _np.linspace(0.0, sr/2.0, num=mag.shape[0])
            if _find_peaks is not None:
                peaks, _ = _find_peaks(mag, distance=max(1, int((200.0/(sr/2.0))*mag.shape[0])))
                if peaks.size > 10:
                    idx = _np.argsort(mag[peaks])[-10:]
                    peaks = peaks[idx]
            else:
                peaks = _np.argpartition(mag, -10)[-10:]
            amps = mag[peaks]
            if amps.size > 0:
                amps = amps / float(amps.max())
            fsel = freqs_axis[peaks]
            keep = (fsel >= 150.0) & (fsel <= sr/2.0 - 100.0)
            formant_freqs = list(map(float, fsel[keep]))
            formant_amps = list(map(float, amps[keep]))
    except Exception as e:
        print(L(f"Errore analisi audio: {e}", f"Audio analysis error: {e}"))
        return None

    # Normalize amplitudes 0..1
    pairs = []
    if formant_freqs:
        maxamp = max(formant_amps) if formant_amps else 1.0
        if maxamp <= 0:
            maxamp = 1.0
        for f, a in zip(formant_freqs, formant_amps):
            pairs.append((float(f), max(0.0, min(1.0, float(a) / maxamp))))
    # Build f0_times if possible
    f0_times = []
    try:
        if time_c is not None and isinstance(time_c, (_np.ndarray, list)):
            # CREPE path with optional confidence mask
            tarr = _np.asarray(time_c)
            if isinstance(mask, _np.ndarray) and mask.shape[0] == tarr.shape[0]:
                tmask = tarr[mask]
            else:
                tmask = tarr
            f0_times = [float(t) for t in tmask.tolist()[:len(f0_vals)]]
        else:
            # Estimate frame times from hop_size and sr using librosa
            if isinstance(sr, (int, float)) and sr > 0:
                n_frames = int(_np.ceil(len(y) / float(hop_size))) if hop_size and hop_size > 0 else 0
                if n_frames > 0:
                    t = _lb.times_like(_np.zeros(n_frames), sr=sr, hop_length=hop_size)
                    if len(t) > 0:
                        # simple resample to match f0_vals length
                        idx = _np.linspace(0, max(1, len(t) - 1), num=len(f0_vals)) if len(f0_vals) > 0 else []
                        f0_times = [float(t[int(round(i))]) for i in idx] if len(f0_vals) > 0 else []
    except Exception:
        f0_times = []
    return {"f0_hz": f0_hz, "formants": pairs, "f0_series": [float(x) for x in (f0_vals.tolist() if 'f0_vals' in locals() else [])], "f0_times": f0_times}


def export_comparison_tables(output_base: str, ratios: List[float], basekey: int,
                             basenote_hz: float, diapason_hz: float,
                             compare_fund_hz: Optional[float] = None,
                             _tet_align: str = "same",
                             subharm_fund_hz: Optional[float] = None,
                             tet_divisions: int = 12,
                             analysis_result: Optional[dict] = None,
                             delta_threshold_hz: float = 0.0) -> None:
    """Esporta tabelle di confronto con TET (12/24/48), serie armonica, subarmonica e analisi audio opzionale. /
    Export comparison with TET (12/24/48), harmonic, subharmonic, and optional audio analysis."""

    def freq_to_note_name(freq: float, a4_hz: float) -> str:
        """Converte frequenza in nome nota 12-TET (solo per labeling). /
        Convert frequency to 12-TET note name (labeling only)."""
        try:
            if not (freq > 0 and a4_hz > 0):
                return ""
            midi = int(round(MIDI_A4 + SEMITONES_PER_OCTAVE * math.log2(freq / a4_hz)))
        except (ValueError, OverflowError):
            return ""
        midi = max(MIDI_MIN, min(MIDI_MAX, midi))
        names = ["C", "C#", "D", "D#", "E", "F", "F#", "G", "G#", "A", "A#", "B"]
        name = names[midi % SEMITONES_PER_OCTAVE]
        octave = (midi // SEMITONES_PER_OCTAVE) - 1
        return f"{name}{octave}"

    def tet_step_index(freq: float, base_freq: float, divs: int) -> int:
        """Restituisce l'indice del passo TET più vicino da base_freq. /
        Return nearest TET step index from base_freq."""
        if freq <= 0 or base_freq <= 0:
            return 0
        return int(round(divs * math.log2(freq / base_freq)))

    base_cmp = compare_fund_hz if compare_fund_hz is not None else basenote_hz
    sub_base = subharm_fund_hz if subharm_fund_hz is not None else diapason_hz

    # Prepara dati ordinati
    computed = [(basenote_hz * float(r), i, basekey + i, float(r))
                for i, r in enumerate(ratios)]
    computed.sort(key=lambda t: t[0])

    custom_list = computed
    n_rows = len(custom_list)

    # Serie armonica
    harm_vals = []
    harm_n = 1
    while True:
        harm_freq_val = base_cmp * harm_n
        if harm_freq_val > MAX_HARMONIC_HZ:
            break
        harm_vals.append(harm_freq_val)
        harm_n += 1

    # Serie subarmonica
    sub_desc = []
    m = 1
    while True:
        sub_freq_val = sub_base / m
        if sub_freq_val < MIN_SUBHARMONIC_HZ:
            break
        sub_desc.append(sub_freq_val)
        m += 1
    sub_vals = list(reversed(sub_desc))

    # Filtro subarmoniche
    min_custom = custom_list[0][0] if n_rows > 0 else base_cmp
    cutoff_sub = max(MIN_SUBHARMONIC_HZ, min_custom)
    sub_vals = [v for v in sub_vals if v >= cutoff_sub]

    # Allineamento sequenze
    def align_sequence(seq: List[float], customs: List[float]) -> List[Optional[float]]:
        out: List[Optional[float]] = [None] * len(customs)
        p = 0
        for j in range(len(customs)):
            low = customs[j]
            high = customs[j + 1] if j + 1 < len(customs) else float('inf')
            if p < len(seq) and low <= seq[p] < high:
                out[j] = seq[p]
                p += 1
        return out

    custom_hz_list = [c[0] for c in custom_list]
    harm_aligned = align_sequence(harm_vals, custom_hz_list)
    sub_aligned = align_sequence(sub_vals, custom_hz_list)

    # Export testo
    txt_path = f"{output_base}_compare.txt"
    try:
        # Prepara allineamento analisi audio / Prepare audio analysis alignment
        audio_f0_idx = None
        audio_f0_val = None
        audio_formant_map = {}
        if analysis_result is not None:
            f0_val = analysis_result.get('f0_hz') if isinstance(analysis_result, dict) else None
            formants = analysis_result.get('formants') if isinstance(analysis_result, dict) else None
            if f0_val and f0_val > 0 and custom_hz_list:
                # indice più vicino / nearest index
                audio_f0_idx = int(min(range(len(custom_hz_list)), key=lambda k: abs(custom_hz_list[k] - f0_val)))
                audio_f0_val = float(f0_val)
            if formants and custom_hz_list:
                # Escludi formanti sotto f0 / Exclude formants below f0
                _ff_list = formants
                if f0_val and f0_val > 0:
                    _ff_list = [(ff, amp) for (ff, amp) in formants if ff >= f0_val]
                for (ff, amp) in sorted(_ff_list, key=lambda x: x[0]):
                    idx = int(min(range(len(custom_hz_list)), key=lambda k: abs(custom_hz_list[k] - ff)))
                    if idx not in audio_formant_map or amp > audio_formant_map[idx][1]:
                        audio_formant_map[idx] = (float(ff), float(max(0.0, min(1.0, amp))))

        headers = [
            "Step", "MIDI", "Ratio",
            "Custom_Hz", "Harmonic_Hz", "DeltaHz_Harm",
            "Subharm_Hz", "DeltaHz_Sub",
            "TET_Hz", "TET_Note", "DeltaHz_TET",
            "AudioF0_Hz", "AudioFormant_Hz", "Formant_RelAmp",
            "DeltaHz_F0", "DeltaHz_Formant"
        ]

        rows = []
        for row_i, (custom_hz, _step_idx, _midi, r) in enumerate(custom_list):
            harm_val = harm_aligned[row_i]
            sub_val = sub_aligned[row_i]

            # N-TET (12/24/48)
            if custom_hz > 0 and base_cmp > 0:
                tet_val = base_cmp * (2.0 ** (tet_step_index(custom_hz, base_cmp, tet_divisions) / tet_divisions))
            else:
                tet_val = base_cmp

            # Formatta valori
            harm_str = f"{harm_val:.6f}" if harm_val is not None else ""
            d_har_str = f"{(custom_hz - harm_val):.6f}" if harm_val is not None else ""
            approx = "≈" if harm_val is not None and abs(custom_hz - harm_val) < PROXIMITY_THRESHOLD_HZ else ""

            sub_str = f"{sub_val:.6f}" if sub_val is not None else ""
            d_sub_str = f"{(custom_hz - sub_val):.6f}" if sub_val is not None else ""

            tet_str = f"{tet_val:.6f}"
            tet_note = freq_to_note_name(tet_val, diapason_hz)
            d_tet_str = f"{(custom_hz - tet_val):.6f}"

            # Audio alignments for this row
            f0_str = f"{audio_f0_val:.6f}" if (audio_f0_idx is not None and audio_f0_idx == row_i and audio_f0_val) else ""
            af = audio_formant_map.get(row_i)
            a_formant_str = f"{af[0]:.6f}" if af else ""
            a_amp_str = f"{af[1]:.3f}" if af else ""
            d_f0_str = f"{(custom_hz - float(audio_f0_val)):.6f}" if f0_str else ""
            d_form_str = f"{(custom_hz - af[0]):.6f}" if af else ""

            # Apply delta threshold filtering (text output)
            if delta_threshold_hz and delta_threshold_hz > 0:
                try:
                    if harm_val is not None and abs(custom_hz - harm_val) < delta_threshold_hz:
                        harm_str = ""
                        d_har_str = ""
                    if sub_val is not None and abs(custom_hz - sub_val) < delta_threshold_hz:
                        sub_str = ""
                        d_sub_str = ""
                    if tet_val is not None and abs(custom_hz - tet_val) < delta_threshold_hz:
                        tet_str = ""
                        tet_note = ""
                        d_tet_str = ""
                    if f0_str:
                        if abs(custom_hz - float(audio_f0_val)) < delta_threshold_hz:
                            f0_str = ""
                            d_f0_str = ""
                    if af:
                        if abs(custom_hz - af[0]) < delta_threshold_hz:
                            a_formant_str = ""
                            a_amp_str = ""
                            d_form_str = ""
                except Exception:
                    pass

            rows.append([
                str(row_i), str(basekey + row_i), f"{r:.10f}",
                f"{custom_hz:.6f}{approx}",
                f"{harm_str}{approx if harm_str else ''}",
                d_har_str, sub_str, d_sub_str,
                tet_str, tet_note, d_tet_str,
                f0_str, a_formant_str, a_amp_str,
                d_f0_str, d_form_str
            ])

        # Calcola larghezze
        widths = [len(h) for h in headers]
        for row in rows:
            for c, val in enumerate(row):
                widths[c] = max(widths[c], len(val))

        def fmt(vals: List[str]) -> str:
            return "  ".join(str(value).ljust(widths[i]) for i, value in enumerate(vals))

        with open(txt_path, "w", encoding="utf-8") as f:
            f.write(fmt(headers) + "\n")
            for row in rows:
                f.write(fmt(row) + "\n")
        print(L(f"Esportato: {txt_path}", f"Exported: {txt_path}"))
    except IOError as e:
        print(f"Errore scrittura {txt_path}: {e}")

    # Export Excel (opzionale)
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
        from openpyxl.utils import get_column_letter

        wb = Workbook()
        ws = wb.active
        ws.title = "Compare"

        headers_xl = [
            "Step", "MIDI", "Ratio",
            "Custom_Hz", "Harmonic_Hz", "|DeltaHz_Harm|",
            "Subharm_Hz", "|DeltaHz_Sub|",
            "TET_Hz", "TET_Note", "|DeltaHz_TET|",
            "AudioF0_Hz", "AudioFormant_Hz", "Formant_RelAmp",
            "|DeltaHz_F0|", "|DeltaHz_Formant|"
        ]
        ws.append(headers_xl)

        header_fill = PatternFill(start_color="FFCCE5FF", end_color="FFCCE5FF",
                                  fill_type="solid")
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.fill = header_fill

        # Colori serie
        fill_custom = PatternFill(start_color="FFFFCCCC", end_color="FFFFCCCC",
                                  fill_type="solid")
        fill_harm = PatternFill(start_color="FFCCFFCC", end_color="FFCCFFCC",
                                fill_type="solid")
        fill_sub = PatternFill(start_color="FFFFFFCC", end_color="FFFFFFCC",
                               fill_type="solid")
        fill_tet = PatternFill(start_color="FFCCE5FF", end_color="FFCCE5FF",
                               fill_type="solid")

        # Prepara mappa analisi audio per Excel
        audio_f0_idx = None
        audio_formant_map = {}
        if analysis_result is not None:
            f0_val = analysis_result.get('f0_hz') if isinstance(analysis_result, dict) else None
            formants = analysis_result.get('formants') if isinstance(analysis_result, dict) else None
            if f0_val and f0_val > 0 and custom_hz_list:
                audio_f0_idx = int(min(range(len(custom_hz_list)), key=lambda k: abs(custom_hz_list[k] - f0_val)))
            if formants and custom_hz_list:
                # Escludi formanti sotto f0 / Exclude formants below f0
                _ff_list = formants
                if f0_val and f0_val > 0:
                    _ff_list = [(ff, amp) for (ff, amp) in formants if ff >= f0_val]
                for (ff, amp) in sorted(_ff_list, key=lambda x: x[0]):
                    idx = int(min(range(len(custom_hz_list)), key=lambda k: abs(custom_hz_list[k] - ff)))
                    if idx not in audio_formant_map or amp > audio_formant_map[idx][1]:
                        audio_formant_map[idx] = (float(ff), float(max(0.0, min(1.0, amp))))

        for row_i, (custom_hz, _step_idx, _midi, r) in enumerate(custom_list):
            harm_val = harm_aligned[row_i]
            sub_val = sub_aligned[row_i]

            if custom_hz > 0 and base_cmp > 0:
                tet_val = base_cmp * (2.0 ** (tet_step_index(custom_hz, base_cmp, tet_divisions) / tet_divisions))
            else:
                tet_val = base_cmp

            # Threshold for gating
            thr = float(delta_threshold_hz) if isinstance(delta_threshold_hz, (int, float)) else 0.0
            idx = row_i + 2
            # Harmonic: write as formula gating against threshold using literal value
            if harm_val is not None and thr > 0:
                harm_cell = f"=IF(ABS({CUSTOM_COLUMN}{idx}-{harm_val:.6f})>={thr:.6f},{harm_val:.6f},\"\")"
                d_har_cell = f"=IF(ABS({CUSTOM_COLUMN}{idx}-{harm_val:.6f})>={thr:.6f},ABS({CUSTOM_COLUMN}{idx}-{harm_val:.6f}),\"\")"
            else:
                harm_cell = harm_val if harm_val is not None else None
                d_har_cell = (f"=IF({HARM_COLUMN}{idx}<>\"\",ABS({CUSTOM_COLUMN}{idx}-{HARM_COLUMN}{idx}),\"\")")
            # Subharmonic gating
            if sub_val is not None and thr > 0:
                sub_cell = f"=IF(ABS({CUSTOM_COLUMN}{idx}-{sub_val:.6f})>={thr:.6f},{sub_val:.6f},\"\")"
                d_sub_cell = f"=IF(ABS({CUSTOM_COLUMN}{idx}-{sub_val:.6f})>={thr:.6f},ABS({CUSTOM_COLUMN}{idx}-{sub_val:.6f}),\"\")"
            else:
                sub_cell = sub_val if sub_val is not None else None
                d_sub_cell = (f"=IF({SUB_COLUMN}{idx}<>\"\",ABS({CUSTOM_COLUMN}{idx}-{SUB_COLUMN}{idx}),\"\")")
            # TET gating
            if thr > 0:
                tet_cell = f"=IF(ABS({CUSTOM_COLUMN}{idx}-{tet_val:.6f})>={thr:.6f},{tet_val:.6f},\"\")"
                d_tet_cell = f"=IF(ABS({CUSTOM_COLUMN}{idx}-{tet_val:.6f})>={thr:.6f},ABS({CUSTOM_COLUMN}{idx}-{tet_val:.6f}),\"\")"
            else:
                tet_cell = tet_val
                d_tet_cell = (f"=IF({TET_COLUMN}{idx}<>\"\",ABS({CUSTOM_COLUMN}{idx}-{TET_COLUMN}{idx}),\"\")")
            tet_note = freq_to_note_name(tet_val, diapason_hz)

            # Valori analisi per questa riga (use actual audio values, not custom)
            f0_cell_val = None
            if (audio_f0_idx is not None and audio_f0_idx == row_i and isinstance(analysis_result.get('f0_hz'), (int, float))):
                f0_cell_val = float(analysis_result.get('f0_hz'))
            a_formant = audio_formant_map.get(row_i)
            a_formant_hz = a_formant[0] if a_formant else None
            a_formant_amp = a_formant[1] if a_formant else None

            # Delta formulas for audio with threshold gating
            if thr > 0:
                d_f0_cell = (f"=IF({F0_COLUMN}{idx}<>\"\",IF(ABS({CUSTOM_COLUMN}{idx}-{F0_COLUMN}{idx})>={thr:.6f},ABS({CUSTOM_COLUMN}{idx}-{F0_COLUMN}{idx}),\"\"),\"\")")
                d_form_cell = (f"=IF({FORMANT_COLUMN}{idx}<>\"\",IF(ABS({CUSTOM_COLUMN}{idx}-{FORMANT_COLUMN}{idx})>={thr:.6f},ABS({CUSTOM_COLUMN}{idx}-{FORMANT_COLUMN}{idx}),\"\"),\"\")")
            else:
                d_f0_cell = (f"=IF({F0_COLUMN}{idx}<>\"\",ABS({CUSTOM_COLUMN}{idx}-{F0_COLUMN}{idx}),\"\")")
                d_form_cell = (f"=IF({FORMANT_COLUMN}{idx}<>\"\",ABS({CUSTOM_COLUMN}{idx}-{FORMANT_COLUMN}{idx}),\"\")")

            ws.append([row_i, basekey + row_i, r, custom_hz, harm_cell, d_har_cell,
                       sub_cell, d_sub_cell, tet_cell, tet_note, d_tet_cell,
                       (f0_cell_val if thr <= 0 else (f"=IF(ABS({CUSTOM_COLUMN}{idx}-{f0_cell_val:.6f})>={thr:.6f},{f0_cell_val:.6f},\"\")" if f0_cell_val is not None else None)),
                       (a_formant_hz if (thr <= 0 or a_formant_hz is None) else f"=IF(ABS({CUSTOM_COLUMN}{idx}-{a_formant_hz:.6f})>={thr:.6f},{a_formant_hz:.6f},\"\")"),
                       a_formant_amp,
                       d_f0_cell, d_form_cell])

            row = ws.max_row
            ws.cell(row=row, column=4).fill = fill_custom
            ws.cell(row=row, column=5).fill = fill_harm
            ws.cell(row=row, column=7).fill = fill_sub
            ws.cell(row=row, column=9).fill = fill_tet
            ws.cell(row=row, column=10).fill = fill_tet

        # Foglio opzionale per analisi diapason / Optional sheet for diapason analysis
        if analysis_result and isinstance(analysis_result, dict) and (analysis_result.get('diapason_est') or analysis_result.get('f0_list') or analysis_result.get('ratio_clusters')):
            ws2 = wb.create_sheet(title="Diapason")
            ws2.freeze_panes = "A2"
            # Riepilogo
            try:
                a4_est_val = float(analysis_result.get('diapason_est')) if analysis_result.get('diapason_est') is not None else ""
            except (ValueError, TypeError):
                a4_est_val = ""
            ws2.append(["A4_est_Hz", a4_est_val])
            ws2.append(["Basenote_Hz", float(basenote_hz)])
            ws2.append(["Starting_Diapason_Hz", float(diapason_hz)])
            try:
                base_hint = float(analysis_result.get('base_hint_hz')) if analysis_result.get('base_hint_hz') is not None else ""
            except (ValueError, TypeError):
                base_hint = ""
            ws2.append(["Base_hint_Hz_for_ratios", base_hint])
            # Scala match (from scl/)
            try:
                _sc_info = analysis_result.get('scala_match_info') if isinstance(analysis_result, dict) else None
            except Exception:
                _sc_info = None
            if _sc_info:
                ws2.append(["Scala_match_name", _sc_info.get('name','')])
                ws2.append(["Scala_match_file", _sc_info.get('file','')])
                try:
                    ws2.append(["Scala_match_avg_error_cents", float(_sc_info.get('avg_cents_error'))])
                except Exception:
                    ws2.append(["Scala_match_avg_error_cents", ""])    
                # Optional: steps mapping
                _sc_steps = analysis_result.get('scala_match_steps') if isinstance(analysis_result, dict) else None
                if _sc_steps:
                    ws2.append(["Scala_Map", "(index)", "Ratio", "Hz_from_Base(utente)", "Hz_from_Base(stimato)", "Cents", "Count"])
                    for c in range(1, 8):
                        cell = ws2.cell(row=ws2.max_row, column=c)
                        cell.font = Font(bold=True)
                        cell.fill = header_fill
                    # determine estimated base for steps (re-use base_est_for_steps after it's computed below if available)
                    _base_est_for_steps = None
                    try:
                        if isinstance(analysis_result, dict) and analysis_result.get('basenote_est_hz'):
                            _base_est_for_steps = float(analysis_result.get('basenote_est_hz'))
                    except Exception:
                        _base_est_for_steps = None
                    if _base_est_for_steps is None:
                        try:
                            _a4_est_try = float(analysis_result.get('diapason_est')) if analysis_result.get('diapason_est') is not None else None
                            if _a4_est_try and float(diapason_hz) > 0:
                                _base_est_for_steps = float(basenote_hz) * (_a4_est_try / float(diapason_hz))
                        except Exception:
                            _base_est_for_steps = None
                    for (idx_step, ratio_c, cnt) in _sc_steps:
                        try:
                            _hz_user = float(basenote_hz) * float(ratio_c)
                            _hz_est = (float(_base_est_for_steps) * float(ratio_c)) if isinstance(_base_est_for_steps, (int, float)) and _base_est_for_steps > 0 else ""
                            _cents = 1200.0 * math.log2(float(ratio_c))
                            ws2.append([None, int(idx_step), float(ratio_c), float(_hz_user), _hz_est if _hz_est != "" else "", float(_cents), int(cnt)])
                        except Exception:
                            continue
            # Top-5 Scala matches table
            try:
                _sc_top = analysis_result.get('scala_top_matches') if isinstance(analysis_result, dict) else None
            except Exception:
                _sc_top = None
            if _sc_top:
                ws2.append(["Scala_Top5", "Rank", "Name", "File", "AvgError_cents"] )
                # style header
                for c in range(1, 6):
                    cell = ws2.cell(row=ws2.max_row, column=c)
                    cell.font = Font(bold=True)
                    cell.fill = header_fill
                for i, item in enumerate(_sc_top[:5], start=1):
                    try:
                        nm = item.get('name','')
                        fl = item.get('file','')
                        er = float(item.get('avg_cents_error')) if item.get('avg_cents_error') is not None else ""
                    except Exception:
                        nm, fl, er = '', '', ""
                    ws2.append([None, i, nm, fl, er if er != "" else ""]) 
            # Scala within-threshold block (optional)
            try:
                _sc_within = analysis_result.get('scala_within_matches') if isinstance(analysis_result, dict) else None
                _sc_thr = analysis_result.get('scala_within_threshold_cents') if isinstance(analysis_result, dict) else None
            except Exception:
                _sc_within, _sc_thr = None, None
            if _sc_within:
                ws2.append(["Scala_Within", "Rank", "Name", "File", "AvgError_cents", "Threshold_cents", (_sc_thr if isinstance(_sc_thr, (int, float)) else "")])
                for c in range(1, 8):
                    cell = ws2.cell(row=ws2.max_row, column=c)
                    cell.font = Font(bold=True)
                    cell.fill = header_fill
                for i, item in enumerate(_sc_within, start=1):
                    try:
                        nm = item.get('name','')
                        fl = item.get('file','')
                        er = float(item.get('avg_cents_error')) if item.get('avg_cents_error') is not None else ""
                    except Exception:
                        nm, fl, er = '', '', ""
                    ws2.append([None, i, nm, fl, er if er != "" else "", (_sc_thr if isinstance(_sc_thr, (int, float)) else "")])
            # Basenote estimated from analysis (12-TET suggestion)
            try:
                bn_est_hz = float(analysis_result.get('basenote_est_hz')) if analysis_result.get('basenote_est_hz') is not None else ""
            except (ValueError, TypeError):
                bn_est_hz = ""
            try:
                bn_est_midi = int(analysis_result.get('basenote_est_midi')) if analysis_result.get('basenote_est_midi') is not None else ""
            except (ValueError, TypeError):
                bn_est_midi = ""
            bn_est_name = analysis_result.get('basenote_est_name_12tet') if isinstance(analysis_result.get('basenote_est_name_12tet'), str) else ""
            ws2.append(["Basenote_est_Hz", bn_est_hz])
            ws2.append(["Basenote_est_12TET", bn_est_name, bn_est_midi])
            ws2.append([])
            # Elenco F0 e rapporti (with side-by-side TET comparisons)
            ws2.append(["F0_Hz", "Ratio_to_Basenote_Reduced", "Hz_from_Ratio", "TET_user_Hz", "|Delta_TET_user|", "TET_est_Hz", "|Delta_TET_est|"])
            f0_list_raw = analysis_result.get('f0_list') or []
            # Sort F0 ascending for scale-like presentation
            f0_list_sorted = sorted([float(x) for x in f0_list_raw if isinstance(x, (int, float)) or (isinstance(x, str) and x)])
            # Parse A4 stimato se disponibile
            try:
                a4_est_val_f = float(analysis_result.get('diapason_est')) if analysis_result.get('diapason_est') is not None else None
            except (ValueError, TypeError):
                a4_est_val_f = None
            for ff in f0_list_sorted:
                # Compute ratio to basenote reduced to [1,2)
                rr = None
                try:
                    if float(basenote_hz) > 0:
                        rr = float(reduce_to_octave(ff / float(basenote_hz)))
                except Exception:
                    rr = None
                hz_from_r = (rr * basenote_hz) if isinstance(rr, (int, float)) else None
                # TET user (12-TET anchored to user diapason)
                def nearest_tet(freq: float, a4: float) -> float:
                    try:
                        m = int(round(69 + 12 * math.log2(freq / a4)))
                        return a4 * (2.0 ** ((m - 69) / 12.0))
                    except Exception:
                        return 0.0
                tet_user = nearest_tet(ff, float(diapason_hz)) if (ff > 0 and float(diapason_hz) > 0) else 0.0
                d_user = abs(ff - tet_user) if tet_user > 0 else ""
                tet_est = nearest_tet(ff, a4_est_val_f) if (a4_est_val_f and ff > 0) else 0.0
                d_est = abs(ff - tet_est) if tet_est > 0 else ""
                ws2.append([
                    float(ff),
                    float(rr) if isinstance(rr, (int, float)) else "",
                    float(hz_from_r) if isinstance(hz_from_r, (int, float)) else "",
                    float(tet_user) if tet_user > 0 else "",
                    float(d_user) if isinstance(d_user, (int, float)) else "",
                    float(tet_est) if tet_est > 0 else "",
                    float(d_est) if isinstance(d_est, (int, float)) else ""
                ])
            ws2.append([])
            # Cluster dei rapporti
            ws2.append(["ClusterCenter_Ratio", "Count", "Hz_from_Base"])
            clusters = analysis_result.get('ratio_clusters') or []
            for item in clusters:
                try:
                    center, count = item
                except (ValueError, TypeError):
                    continue
                ws2.append([float(center), int(count), float(center) * float(basenote_hz)])

            # Inferred tuning system and steps
            ws2.append([])
            tun_info = analysis_result.get('tuning_inferred') if isinstance(analysis_result, dict) else None
            steps = analysis_result.get('scale_steps') if isinstance(analysis_result, dict) else None
            if tun_info:
                ws2.append(["Tuning_inferred", tun_info.get('name', ''), "AvgError_cents", tun_info.get('avg_cents_error', '')])
            else:
                ws2.append(["Tuning_inferred", "", "AvgError_cents", ""]) 
            ws2.append(["Inferred_Steps", "(index)", "Ratio", "Hz_from_Base(utente)", "Hz_from_Base(stimato)", "Cents", "Count"])
            # Header style
            header_fill2 = PatternFill(start_color="FFEEEEEE", end_color="FFEEEEEE", fill_type="solid")
            for c in range(1, 8):
                cell = ws2.cell(row=ws2.max_row, column=c)
                cell.font = Font(bold=True)
                cell.fill = header_fill2
            # Determine estimated basenote for inferred steps (if available)
            base_est_for_steps = None
            try:
                # Prefer explicit basenote estimated from analysis
                if isinstance(analysis_result, dict) and analysis_result.get('basenote_est_hz'):
                    base_est_for_steps = float(analysis_result.get('basenote_est_hz'))
            except Exception:
                base_est_for_steps = None
            if base_est_for_steps is None:
                # Derive from estimated A4 if present
                try:
                    a4_est_try = float(analysis_result.get('diapason_est')) if isinstance(analysis_result, dict) and analysis_result.get('diapason_est') is not None else None
                    if a4_est_try and float(diapason_hz) > 0:
                        base_est_for_steps = float(basenote_hz) * (a4_est_try / float(diapason_hz))
                except Exception:
                    base_est_for_steps = None
            if steps:
                for (idx_step, ratio_c, cnt) in steps:
                    try:
                        hz_user = float(basenote_hz) * float(ratio_c)
                        hz_est = (float(base_est_for_steps) * float(ratio_c)) if isinstance(base_est_for_steps, (int, float)) and base_est_for_steps > 0 else ""
                        cents = 1200.0 * math.log2(float(ratio_c))
                        ws2.append([None, int(idx_step), float(ratio_c), float(hz_user), hz_est if hz_est != "" else "", float(cents), int(cnt)])
                    except Exception:
                        continue
            # Comparative tuning (if available)
            comp_info = analysis_result.get('tuning_comparative') if isinstance(analysis_result, dict) else None
            comp_steps = analysis_result.get('scale_steps_comp') if isinstance(analysis_result, dict) else None
            ws2.append([])
            if comp_info:
                ws2.append(["Tuning_comparative", comp_info.get('name',''), "AvgError_cents", comp_info.get('avg_cents_error','')])
            else:
                ws2.append(["Tuning_comparative", "", "AvgError_cents", ""]) 
            ws2.append(["Comparative_Steps", "(index)", "Ratio", "Hz_from_Base(utente)", "Hz_from_Base(stimato)", "Cents", "Count"])
            for c in range(1, 8):
                cell = ws2.cell(row=ws2.max_row, column=c)
                cell.font = Font(bold=True)
                cell.fill = header_fill2
            if comp_steps:
                for (idx_step, ratio_c, cnt) in comp_steps:
                    try:
                        hz_user = float(basenote_hz) * float(ratio_c)
                        hz_est = (float(base_est_for_steps) * float(ratio_c)) if isinstance(base_est_for_steps, (int, float)) and base_est_for_steps > 0 else ""
                        cents = 1200.0 * math.log2(float(ratio_c))
                        ws2.append([None, int(idx_step), float(ratio_c), float(hz_user), hz_est if hz_est != "" else "", float(cents), int(cnt)])
                    except Exception:
                        continue

            # --- Sezioni di confronto: 12-TET e Pitagorico (per diapason utente ed eventuale stimato) ---
            ws2.append([])
            ws2.append(["Riferimenti di confronto / Reference systems"])
            ws2.cell(row=ws2.max_row, column=1).font = Font(bold=True)

            def add_section(title: str, rows: list) -> None:
                ws2.append([title])
                ws2.cell(row=ws2.max_row, column=1).font = Font(bold=True)
                ws2.append(["Step", "Ratio", "Hz", "Cents"])
                header_fill2 = PatternFill(start_color="FFEEEEEE", end_color="FFEEEEEE", fill_type="solid")
                for c in range(1, 5):
                    cell = ws2.cell(row=ws2.max_row, column=c)
                    cell.font = Font(bold=True)
                    cell.fill = header_fill2
                for r in rows:
                    ws2.append(r)
                ws2.append([])

            def build_rows_for_base(base_hz: float) -> dict:
                # 12-TET rows
                tet_rows = []
                for k in range(12):
                    ratio = 2.0 ** (k / 12.0)
                    hz = float(base_hz) * ratio
                    cents = 100.0 * k
                    tet_rows.append([k, float(ratio), float(hz), float(cents)])
                # Pythagorean 12
                try:
                    from fractions import Fraction as _Fr
                    ratios_py12 = []
                    for k in range(12):
                        r = pow_fraction(_Fr(3, 2), k)
                        r = reduce_to_octave(r)
                        ratios_py12.append(float(r))
                    ratios_py12 = normalize_ratios(ratios_py12, reduce_octave=True)
                except Exception:
                    ratios_py12 = [float(reduce_to_octave((3.0/2.0) ** k)) for k in range(12)]
                    ratios_py12 = normalize_ratios(ratios_py12, reduce_octave=True)
                py12_rows = []
                for i, r in enumerate(ratios_py12):
                    hz = float(base_hz) * float(r)
                    cents = 1200.0 * math.log2(float(r))
                    py12_rows.append([i, float(r), float(hz), float(cents)])
                # Pythagorean 7
                try:
                    from fractions import Fraction as _Fr2
                    diatonic_ratios = [
                        _Fr2(1,1), _Fr2(9,8), _Fr2(81,64), _Fr2(4,3),
                        _Fr2(3,2), _Fr2(27,16), _Fr2(243,128)
                    ]
                    diatonic_ratios = [reduce_to_octave(r) for r in diatonic_ratios]
                except Exception:
                    diatonic_ratios = [1.0, 9/8, 81/64, 4/3, 3/2, 27/16, 243/128]
                    diatonic_ratios = [float(reduce_to_octave(r)) for r in diatonic_ratios]
                diatonic_ratios = sorted([float(r) for r in diatonic_ratios])
                py7_rows = []
                for i, r in enumerate(diatonic_ratios):
                    hz = float(base_hz) * float(r)
                    cents = 1200.0 * math.log2(float(r))
                    py7_rows.append([i, float(r), float(hz), float(cents)])
                return {"tet": tet_rows, "py12": py12_rows, "py7": py7_rows}

            # Build for user diapason (current basenote_hz)
            rows_user = build_rows_for_base(float(basenote_hz))
            add_section(f"12-TET (utente A4={float(diapason_hz):.2f} Hz)", rows_user["tet"])
            add_section("Pitagorico 12 (utente)", rows_user["py12"])
            add_section("Pitagorico 7 (utente)", rows_user["py7"])

            # If estimated diapason/basenote is available, build second set
            rows_est = None
            base_est_hz = None
            try:
                a4_est_val_f = float(a4_est_val) if a4_est_val != "" else None
            except (ValueError, TypeError):
                a4_est_val_f = None
            # Prefer basenote estimated from audio if available
            bn_est_hz = None
            try:
                if isinstance(analysis_result, dict) and analysis_result.get('basenote_est_hz'):
                    bn_est_hz = float(analysis_result.get('basenote_est_hz'))
            except Exception:
                bn_est_hz = None
            if isinstance(bn_est_hz, (int, float)) and bn_est_hz > 0:
                base_est_hz = bn_est_hz
                rows_est = build_rows_for_base(base_est_hz)
                add_section(f"12-TET (stimato A4={a4_est_val_f:.2f} Hz)", rows_est["tet"])
                add_section("Pitagorico 12 (stimato)", rows_est["py12"])
                add_section("Pitagorico 7 (stimato)", rows_est["py7"])
            elif a4_est_val_f and float(diapason_hz) > 0:
                base_est_hz = float(basenote_hz) * (a4_est_val_f / float(diapason_hz))
                rows_est = build_rows_for_base(base_est_hz)
                add_section(f"12-TET (stimato A4={a4_est_val_f:.2f} Hz)", rows_est["tet"])
                add_section("Pitagorico 12 (stimato)", rows_est["py12"])
                add_section("Pitagorico 7 (stimato)", rows_est["py7"])

            # Migliora formattazione Diapason: header, bordi, zebra, auto-larghezze
            header_fill_dia = PatternFill(start_color="FFEEEEEE", end_color="FFEEEEEE", fill_type="solid")
            thin_side = Side(style="thin", color="FFB0B0B0")
            thin_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
            zebra_fill = PatternFill(start_color="FFF9F9F9", end_color="FFF9F9F9", fill_type="solid")

            # Bold per etichette di riepilogo (colonna A) fino alla prima riga vuota
            r = 1
            while r <= ws2.max_row:
                row_cells = list(ws2[r])
                if all((c.value is None or str(c.value) == "") for c in row_cells):
                    break
                # Grassetto su prima colonna per righe con almeno 2 colonne compilate
                if len(row_cells) >= 2 and row_cells[0].value not in (None, ""):
                    row_cells[0].font = Font(bold=True)
                r += 1

            # Helper: trova riga con un determinato header nella prima cella
            def find_row_with_label(label: str) -> Optional[int]:
                for rr in range(1, ws2.max_row + 1):
                    v = ws2.cell(row=rr, column=1).value
                    if isinstance(v, str) and v.strip() == label:
                        return rr
                return None

            # Applica formato tabellare a blocchi: header + zebra + bordi fino alla riga vuota successiva
            def format_block(header_row: int) -> None:
                if not header_row:
                    return
                # Header styling su tutta la riga usata
                max_c = ws2.max_column
                for c in range(1, max_c + 1):
                    cell = ws2.cell(row=header_row, column=c)
                    if cell.value not in (None, ""):
                        cell.font = Font(bold=True)
                        cell.fill = header_fill_dia
                        cell.border = thin_border
                        cell.alignment = Alignment(vertical="center")
                # Dati fino alla riga vuota
                rr = header_row + 1
                stripe = False
                while rr <= ws2.max_row:
                    row_vals = [ws2.cell(row=rr, column=cc).value for cc in range(1, max_c + 1)]
                    if all(v in (None, "") for v in row_vals):
                        break
                    stripe = not stripe
                    for cc in range(1, max_c + 1):
                        cell = ws2.cell(row=rr, column=cc)
                        if stripe:
                            cell.fill = zebra_fill
                        cell.border = thin_border
                        cell.alignment = Alignment(vertical="center")
                    rr += 1

            # Formatta blocchi noti
            for label in ["F0_Hz", "ClusterCenter_Ratio", "Inferred_Steps", "Comparative_Steps", "Scala_Map", "Scala_Top5", "Scala_Within"]:
                hr = find_row_with_label(label)
                if hr:
                    format_block(hr)

            # Bordi generali per tutte le celle non vuote
            for rr in range(1, ws2.max_row + 1):
                for cc in range(1, ws2.max_column + 1):
                    cell = ws2.cell(row=rr, column=cc)
                    if cell.value not in (None, ""):
                        # Assign border unconditionally to avoid StyleProxy hashing issues
                        cell.border = thin_border

            # Auto-larghezza colonne in base al contenuto (limite max 40)
            col_max = ws2.max_column
            for c in range(1, col_max + 1):
                max_len = 0
                for rr in range(1, ws2.max_row + 1):
                    v = ws2.cell(row=rr, column=c).value
                    if v is None:
                        continue
                    s = str(v)
                    if len(s) > max_len:
                        max_len = len(s)
                width = max(10, min(40, int(max_len * 1.2) + 2))
                ws2.column_dimensions[get_column_letter(c)].width = width

            # ---- Esporta anche un file di testo con le stesse informazioni (suffix _diapason.txt) ----
            txt_diap_path = f"{output_base}_diapason.txt"
            try:
                lines = []
                lines.append("DIAPASON – Riferimenti di confronto / Reference systems")
                lines.append("")
                # Riepilogo
                lines.append(f"A4_utente (Hz): {float(diapason_hz):.6f}")
                if 'a4_est_val_f' in locals() and a4_est_val_f:
                    lines.append(f"A4_stimato (Hz): {a4_est_val_f:.6f}")
                else:
                    lines.append("A4_stimato (Hz): ")
                lines.append(f"Basenote_Hz (utente): {float(basenote_hz):.6f}")
                if 'base_est_hz' in locals() and isinstance(base_est_hz, (int, float)):
                    lines.append(f"Basenote_Hz (stimato): {float(base_est_hz):.6f}")
                else:
                    lines.append("Basenote_Hz (stimato): ")
                # Suggested basenote (12-TET) from analysis, if any
                try:
                    bn_name = analysis_result.get('basenote_est_name_12tet') if isinstance(analysis_result, dict) else None
                    bn_midi = analysis_result.get('basenote_est_midi') if isinstance(analysis_result, dict) else None
                except Exception:
                    bn_name = None
                    bn_midi = None
                if bn_name or (bn_midi is not None):
                    nm = bn_name or ""
                    mid = f" (MIDI {int(bn_midi)})" if isinstance(bn_midi, (int, float)) else ""
                    lines.append(f"Basenote_12TET (stimato): {nm}{mid}")
                else:
                    lines.append("Basenote_12TET (stimato): ")
                lines.append("")

                # Scala match section
                try:
                    _sc_info = analysis_result.get('scala_match_info') if isinstance(analysis_result, dict) else None
                except Exception:
                    _sc_info = None
                lines.append("Scala_match:")
                if _sc_info:
                    lines.append(f"  name: {_sc_info.get('name','')}")
                    lines.append(f"  file: {_sc_info.get('file','')}")
                    try:
                        lines.append(f"  avg_error_cents: {float(_sc_info.get('avg_cents_error')):.2f}")
                    except Exception:
                        lines.append("  avg_error_cents: ")
                else:
                    lines.append("  name: ")
                    lines.append("  file: ")
                    lines.append("  avg_error_cents: ")
                lines.append("")

                # Scala top-5 section
                try:
                    _sc_top = analysis_result.get('scala_top_matches') if isinstance(analysis_result, dict) else None
                except Exception:
                    _sc_top = None
                lines.append("Scala_top5:")
                if _sc_top:
                    for i, item in enumerate(_sc_top[:5], start=1):
                        try:
                            nm = item.get('name','')
                            fl = item.get('file','')
                            er = float(item.get('avg_cents_error'))
                            lines.append(f"  {i}. {nm} [{fl}] — {er:.2f} cents")
                        except Exception:
                            lines.append(f"  {i}. ")
                else:
                    lines.append("  ")
                lines.append("")
                # Scala within-threshold section
                try:
                    _sc_within = analysis_result.get('scala_within_matches') if isinstance(analysis_result, dict) else None
                    _sc_thr = analysis_result.get('scala_within_threshold_cents') if isinstance(analysis_result, dict) else None
                except Exception:
                    _sc_within, _sc_thr = None, None
                if _sc_within:
                    thr_str = (f"{float(_sc_thr):.2f}" if isinstance(_sc_thr, (int,float)) else "")
                    lines.append(f"Scala_within (<= {thr_str} cents):")
                    for i, item in enumerate(_sc_within, start=1):
                        try:
                            nm = item.get('name','')
                            fl = item.get('file','')
                            er = float(item.get('avg_cents_error'))
                            lines.append(f"  {i}. {nm} [{fl}] — {er:.2f} cents")
                        except Exception:
                            lines.append(f"  {i}. ")
                    lines.append("")
 
                 # Inferred tuning section in text
                tun_info = analysis_result.get('tuning_inferred') if isinstance(analysis_result, dict) else None
                steps = analysis_result.get('scale_steps') if isinstance(analysis_result, dict) else None
                lines.append("Tuning_inferred:")
                if tun_info:
                    lines.append(f"  name: {tun_info.get('name','')}")
                    try:
                        lines.append(f"  avg_error_cents: {float(tun_info.get('avg_cents_error')):.2f}")
                    except Exception:
                        lines.append("  avg_error_cents: ")
                else:
                    lines.append("  name: ")
                    lines.append("  avg_error_cents: ")
                # Table of inferred steps
                lines.append("")
                lines.append("Inferred steps (index, Ratio, Hz_from_Base(utente), Hz_from_Base(stimato), Cents, Count):")
                if steps:
                    # Determine estimated basenote for inferred steps (if available)
                    base_est_for_steps = None
                    try:
                        if isinstance(analysis_result, dict) and analysis_result.get('basenote_est_hz'):
                            base_est_for_steps = float(analysis_result.get('basenote_est_hz'))
                    except Exception:
                        base_est_for_steps = None
                    if base_est_for_steps is None:
                        try:
                            a4_est_try = float(analysis_result.get('diapason_est')) if isinstance(analysis_result, dict) and analysis_result.get('diapason_est') is not None else None
                            if a4_est_try and float(diapason_hz) > 0:
                                base_est_for_steps = float(basenote_hz) * (a4_est_try / float(diapason_hz))
                        except Exception:
                            base_est_for_steps = None
                    headers_inf = ["Idx","Ratio","Hz_user","Hz_est","Cents","Count"]
                    table_inf = []
                    for (idx_step, ratio_c, cnt) in steps:
                        try:
                            hz_user = float(basenote_hz) * float(ratio_c)
                            hz_est = (float(base_est_for_steps) * float(ratio_c)) if isinstance(base_est_for_steps, (int, float)) and base_est_for_steps > 0 else None
                            cents = 1200.0 * math.log2(float(ratio_c))
                            table_inf.append([str(int(idx_step)), f"{float(ratio_c):.10f}", f"{hz_user:.6f}", (f"{hz_est:.6f}" if isinstance(hz_est, (int, float)) else ""), f"{cents:.2f}", str(int(cnt))])
                        except Exception:
                            continue
                    widths_inf = [len(h) for h in headers_inf]
                    for row in table_inf:
                        for i,v in enumerate(row):
                            widths_inf[i] = max(widths_inf[i], len(v))
                    def fmt_inf(cols):
                        return "  ".join(str(cols[i]).ljust(widths_inf[i]) for i in range(len(cols)))
                    lines.append(fmt_inf(headers_inf))
                    for row in table_inf:
                        lines.append(fmt_inf(row))
                else:
                    lines.append("(none)")
                lines.append("")

                # Comparative tuning section in text
                comp_info = analysis_result.get('tuning_comparative') if isinstance(analysis_result, dict) else None
                comp_steps = analysis_result.get('scale_steps_comp') if isinstance(analysis_result, dict) else None
                lines.append("Tuning_comparative:")
                if comp_info:
                    lines.append(f"  name: {comp_info.get('name','')}")
                    try:
                        lines.append(f"  avg_error_cents: {float(comp_info.get('avg_cents_error')):.2f}")
                    except Exception:
                        lines.append("  avg_error_cents: ")
                else:
                    lines.append("  name: ")
                    lines.append("  avg_error_cents: ")
                lines.append("")
                lines.append("Comparative steps (index, Ratio, Hz_from_Base(utente), Hz_from_Base(stimato), Cents, Count):")
                if comp_steps:
                    # Determine estimated basenote for comparative steps (independent of inferred block)
                    base_est_for_steps2 = None
                    try:
                        if isinstance(analysis_result, dict) and analysis_result.get('basenote_est_hz'):
                            base_est_for_steps2 = float(analysis_result.get('basenote_est_hz'))
                    except Exception:
                        base_est_for_steps2 = None
                    if base_est_for_steps2 is None:
                        try:
                            a4_est_try2 = float(analysis_result.get('diapason_est')) if isinstance(analysis_result, dict) and analysis_result.get('diapason_est') is not None else None
                            if a4_est_try2 and float(diapason_hz) > 0:
                                base_est_for_steps2 = float(basenote_hz) * (a4_est_try2 / float(diapason_hz))
                        except Exception:
                            base_est_for_steps2 = None
                    headers_inf = ["Idx","Ratio","Hz_user","Hz_est","Cents","Count"]
                    table_inf = []
                    for (idx_step, ratio_c, cnt) in comp_steps:
                        try:
                            hz_user = float(basenote_hz) * float(ratio_c)
                            hz_est = (float(base_est_for_steps2) * float(ratio_c)) if isinstance(base_est_for_steps2, (int, float)) and base_est_for_steps2 > 0 else None
                            cents = 1200.0 * math.log2(float(ratio_c))
                            table_inf.append([str(int(idx_step)), f"{float(ratio_c):.10f}", f"{hz_user:.6f}", (f"{hz_est:.6f}" if isinstance(hz_est, (int, float)) else ""), f"{cents:.2f}", str(int(cnt))])
                        except Exception:
                            continue
                    widths_inf = [len(h) for h in headers_inf]
                    for row in table_inf:
                        for i,v in enumerate(row):
                            widths_inf[i] = max(widths_inf[i], len(v))
                    def fmt_inf(cols):
                        return "  ".join(str(cols[i]).ljust(widths_inf[i]) for i in range(len(cols)))
                    lines.append(fmt_inf(headers_inf))
                    for row in table_inf:
                        lines.append(fmt_inf(row))
                else:
                    lines.append("(none)")
                lines.append("")

                def add_block(title: str, rows: list):
                    headers = ["Step", "Ratio", "Hz", "Cents"]
                    safe_rows = rows if isinstance(rows, list) else []
                    # Prepare rows as strings
                    table = [[str(r[0]), f"{float(r[1]):.10f}", f"{float(r[2]):.6f}", f"{float(r[3]):.2f}"] for r in safe_rows]
                    widths = [len(h) for h in headers]
                    for row in table:
                        for i, v in enumerate(row):
                            widths[i] = max(widths[i], len(v))
                    def fmt_row(cols):
                        return "  ".join(str(cols[i]).ljust(widths[i]) for i in range(len(cols)))
                    lines.append(title)
                    lines.append(fmt_row(headers))
                    for row in table:
                        lines.append(fmt_row(row))
                    lines.append("")

                # Sezioni per utente
                add_block(f"12-TET (utente A4={float(diapason_hz):.2f} Hz)", rows_user["tet"])
                add_block("Pitagorico 12 (utente)", rows_user["py12"])
                add_block("Pitagorico 7 (utente)", rows_user["py7"])
                # Sezioni per stimato, se disponibli
                if isinstance(rows_est, dict):
                    add_block(f"12-TET (stimato A4={a4_est_val_f:.2f} Hz)", rows_est.get("tet"))
                    add_block("Pitagorico 12 (stimato)", rows_est.get("py12"))
                    add_block("Pitagorico 7 (stimato)", rows_est.get("py7"))

                with open(txt_diap_path, "w", encoding="utf-8") as f_txt:
                    f_txt.write("\n".join(lines) + "\n")
                print(L(f"Esportato: {txt_diap_path}", f"Exported: {txt_diap_path}"))
            except IOError as e:
                print(L(f"Errore scrittura {txt_diap_path}: {e}", f"Write error {txt_diap_path}: {e}"))

        xlsx_path = f"{output_base}_compare.xlsx"
        wb.save(xlsx_path)
        print(L(f"Esportato: {xlsx_path}", f"Exported: {xlsx_path}"))
    except ImportError:
        print(L("openpyxl non installato: export Excel saltato", "openpyxl not installed: Excel export skipped"))
    except Exception as e:
        print(L(f"Errore export Excel: {e}", f"Excel export error: {e}"))


def page_lines(lines: List[str], rows_per_page: Optional[int] = None) -> None:
    """Simple pager: prints lines with a --More-- prompt, Enter continues, q quits.
    Caps the page height at 80 rows by requirement.
    """
    # Determine terminal rows
    if rows_per_page is None:
        cols, term_rows = _term_size(80, 24)
        rows = term_rows - 1
    else:
        rows = int(rows_per_page)
        cols = _term_cols(80)

    # Cap at 80 rows and enforce minimum
    rows = max(5, min(80, rows))

    i = 0
    n = len(lines)
    while i < n:
        # print up to rows lines
        end = min(n, i + rows)
        for j in range(i, end):
            print(lines[j])
        i = end
        if i < n:
            base_prompt = L("--More-- (invio=continua, q=esci)", "--More-- (enter=continue, q=quit)")
            prompt = f"{Style.DIM}{Style.FG_YELLOW}{base_prompt}{Style.RESET}" if _supports_ansi() else base_prompt
            try:
                print(prompt, end="", flush=True)
            except Exception:
                pass
            # Wait for Enter to continue or 'q' to quit
            while True:
                try:
                    ch = sys.stdin.read(1)
                except (OSError, ValueError, AttributeError):
                    try:
                        ch = input()
                    except EOFError:
                        ch = "q"
                    ch = ch[0] if ch else "\n"
                if ch.lower() == 'q' or ch in ('\n', '\r'):
                    break
            # clear prompt line according to terminal width
            try:
                _clear_line(cols)
            except Exception:
                pass
            if ch.lower() == 'q':
                break


def print_step_hz_table(ratios: List[float], basenote_hz: float) -> None:
    """Stampa tabella multi-colonna con Step/Hz."""
    pairs = [(str(i + 1), f"{basenote_hz * float(r):.2f}")
             for i, r in enumerate(ratios)]

    if not pairs:
        page_lines(["Step Hz"])
        return

    # Calcola larghezze
    step_w = max(len("Step"), max(len(p[0]) for p in pairs))
    hz_w = max(len("Hz"), max(len(p[1]) for p in pairs))
    cell_w = step_w + 1 + hz_w
    gap = 3

    # Larghezza terminale
    try:
        term_w = _term_cols(80)
    except Exception:
        term_w = 80

    # Calcola colonne
    cols = max(1, (term_w + gap) // (cell_w + gap)) if cell_w < term_w else 1
    rows_count = math.ceil(len(pairs) / cols)

    # Compose lines instead of printing directly
    lines_out: List[str] = []
    header_cell = "Step".ljust(step_w) + " " + "Hz".ljust(hz_w)
    lines_out.append((" " * gap).join([header_cell] * cols))

    for r in range(rows_count):
        line_cells = []
        for c in range(cols):
            idx = c * rows_count + r
            if idx < len(pairs):
                s, h = pairs[idx]
                cell = s.rjust(step_w) + " " + h.rjust(hz_w)
                line_cells.append(cell)
        if line_cells:
            lines_out.append((" " * gap).join(line_cells))

    page_lines(lines_out)


def process_tuning_system(args: argparse.Namespace, _basenote: float) -> Optional[Tuple[List[float], float]]:
    """Processa il sistema di intonazione e ritorna (ratios, interval) o None."""

    # Sistema naturale
    if args.natural:
        try:
            a_max = int(args.natural[0])
            b_max = int(args.natural[1])
            if a_max < 0 or b_max < 0:
                print(L("A_MAX e B_MAX devono essere >= 0", "A_MAX and B_MAX must be >= 0"))
                return None
            ratios = build_natural_ratios(a_max, b_max, not args.no_reduce)
            return ratios, DEFAULT_OCTAVE
        except (TypeError, ValueError):
            print("Valori non validi per --natural")
            return None

    # Sistema Danielou
    if args.danielou_all:
        ratios = build_danielou_ratios(True, not args.no_reduce)
        return ratios, DEFAULT_OCTAVE

    if args.danielou is not None:
        ratios = []
        for (a, b, c) in args.danielou:
            ratios.extend(danielou_from_exponents(a, b, c, not args.no_reduce))
        return ratios, DEFAULT_OCTAVE

    # Sistema geometrico
    if args.geometric:
        parts = list(args.geometric)
        if len(parts) != 3:
            print("Uso: --geometric GEN STEPS INTERVAL")
            return None

        # Parse STEPS
        try:
            steps = int(parts[1])
        except (TypeError, ValueError):
            print("Numero di passi non valido")
            return None
        if steps <= 0:
            print("Numero di passi deve essere > 0")
            return None

        # Parse INTERVAL
        try:
            interval_ratio = parse_interval_value(parts[2])
        except argparse.ArgumentTypeError as e:
            print(f"Intervallo non valido: {e}")
            return None

        # Parse GEN as int or fraction, fallback to float
        try:
            gen_val = int_or_fraction(parts[0])
            gen_ratio = Fraction(gen_val, 1) if isinstance(gen_val, int) else gen_val
        except (argparse.ArgumentTypeError, ValueError) as e:
            try:
                gen_ratio = float(parts[0])
            except ValueError:
                print(f"Errore parsing geometrico: {e}")
                return None

        if float(gen_ratio) <= 0:
            print("Il generatore deve essere > 0")
            return None

        ratios = []
        for i in range(steps):
            r = pow_fraction(gen_ratio, i) if isinstance(gen_ratio, Fraction) else float(gen_ratio) ** i
            if not args.no_reduce:
                r = reduce_to_interval(r, interval_ratio)
            ratios.append(float(r))

        return ratios, float(interval_ratio)

    # Temperamento equabile (default)
    if args.et:
        index, cents = args.et
        if index <= 0 or (isinstance(cents, (int, float)) and cents <= 0):
            print("Indice o cents non valido")
            return None

        if is_fraction_type(cents):
            cents = fraction_to_cents(cents)


        try:
            ratio = ratio_et(index, cents)
        except ZeroDivisionError:
            print("Errore divisione per zero")
            return None

        ratios = [float(ratio ** i) for i in range(index)]
        interval_factor = math.exp(cents / ELLIS_CONVERSION_FACTOR)
        return ratios, interval_factor

    return None


def convert_excel_to_outputs(excel_path: str,
                              output_base: Optional[str],
                              default_basekey: int,
                              default_base_hz: float,
                              diapason_hz: float,
                              midi_truncate: bool = False,
                              tun_integer: bool = False) -> bool:
    """
    Converte un file Excel (System/Compare) in output .csd (cpstun) e .tun usando
    la colonna dei rapporti (o Hz/Base_Hz se necessario).
    Ritorna True se conversione riuscita.
    """
    # ATTENZIONE! Deve recuperare il valore di diapason o prendere il --diapason da riga di comando
    # e nel convertire il file .TUN deve essere in tal caso [Exact Tuning] 
    # basefreq = 8.1757989156437073336 * (diapasonHz/440)
    try:
        from openpyxl import load_workbook
    except ImportError:
        print(L("openpyxl non installato: impossibile usare --convert",
                "openpyxl not installed: cannot use --convert"))
        return False

    if not file_exists(excel_path):
        print(L(f"File Excel non trovato: {excel_path}", f"Excel file not found: {excel_path}"))
        return False

    try:
        wb = load_workbook(excel_path, data_only=True)
    except Exception as e:
        print(L(f"Errore apertura Excel: {e}", f"Error opening Excel: {e}"))
        return False

    # Scegli foglio: preferisci 'System' se presente, altrimenti attivo
    if 'System' in wb.sheetnames:
        ws = wb['System']
    else:
        ws = wb.active

    # Mappa header -> colonna (1-based)
    header_row = 1
    headers = {}
    try:
        for cell in ws[header_row]:
            if cell.value is None:
                continue
            name = str(cell.value).strip()
            if name:
                headers[name.lower()] = cell.column if isinstance(cell.column, int) else cell.column
    except Exception:
        headers = {}

    def col_index_by_name(names: list) -> Optional[int]:
        def _col_to_index(c) -> Optional[int]:
            if isinstance(c, int):
                return c
            try:
                s = str(c).strip()
                if not s:
                    return None
                # Convert Excel letters to 1-based index
                idx = 0
                for ch in s.upper():
                    if 'A' <= ch <= 'Z':
                        idx = idx * 26 + (ord(ch) - ord('A') + 1)
                    else:
                        return None
                return idx if idx > 0 else None
            except Exception:
                return None
        for key, col in headers.items():
            for nm in names:
                if nm in key:
                    ci = _col_to_index(col)
                    if ci:
                        return ci
        return None

    ratio_col = col_index_by_name(["ratio"])  # es. "ratio", "custom ratio"
    hz_col = col_index_by_name(["hz"])  # es. "hz"
    midi_col = col_index_by_name(["midi"])  # es. "midi"

    # Base_Hz: prova a trovarlo da intestazione 'Base_Hz' o F2
    base_hz = float(default_base_hz) if default_base_hz and default_base_hz > 0 else DEFAULT_DIAPASON
    base_anchor_col = col_index_by_name(["base_hz", "base hz", "base-freq", "basefreq"]) or 6  # fallback F
    try:
        base_cell_val = ws.cell(row=2, column=base_anchor_col).value
        if isinstance(base_cell_val, (int, float)) and float(base_cell_val) > 0:
            base_hz = float(base_cell_val)
    except Exception:
        pass

    # Basekey: usa prima riga della colonna MIDI se presente, altrimenti default
    basekey = int(default_basekey)
    if midi_col:
        try:
            mv = ws.cell(row=2, column=midi_col).value
            if isinstance(mv, (int, float)):
                basekey = int(mv)
        except Exception:
            pass

    # Estrai rapporti
    ratios: List[float] = []
    row = 2
    max_empty = 0
    while True:
        v_ratio = None
        v_hz = None
        if ratio_col:
            try:
                v_ratio = ws.cell(row=row, column=ratio_col).value
            except Exception:
                v_ratio = None
        if hz_col:
            try:
                v_hz = ws.cell(row=row, column=hz_col).value
            except Exception:
                v_hz = None
        if v_ratio is None and v_hz is None:
            max_empty += 1
            if max_empty >= 5:  # stop after a few empty rows
                break
            row += 1
            continue
        # Reset empty counter when any content
        max_empty = 0
        r_val: Optional[float] = None
        if isinstance(v_ratio, (int, float)) and float(v_ratio) > 0:
            r_val = float(v_ratio)
        elif isinstance(v_hz, (int, float)) and float(v_hz) > 0 and base_hz > 0:
            try:
                r_val = float(v_hz) / float(base_hz)
            except Exception:
                r_val = None
        if r_val is not None and r_val > 0:
            ratios.append(r_val)
        row += 1

    if not ratios:
        print(L("Nessun rapporto valido trovato in Excel (colonne Ratio/Hz)",
                "No valid ratios found in Excel (Ratio/Hz columns)"))
        return False

    # Assicura compatibilità MIDI
    ratios_eff, basekey_eff = ensure_midi_fit(ratios, basekey, midi_truncate)

    # Base output
    if output_base and isinstance(output_base, str) and output_base.strip():
        out_base = output_base.strip()
    else:
        try:
            stem = os.path.splitext(os.path.basename(excel_path))[0]
        except Exception:
            stem = "out"
        out_base = stem

    # Scrivi cpstun e tun (entrambi di default)
    fnum, existed = write_cpstun_table(out_base, ratios_eff, basekey_eff, base_hz, None)
    write_tun_file(out_base if not existed else f"{out_base}_{fnum}", diapason_hz, ratios_eff, basekey_eff, base_hz, tun_integer)

    return True


def infer_tuning_system_general(cluster_centers: List[Tuple[float, int]]):
    """
    Inferenza generale del sistema di intonazione a partire dai centroidi dei cluster (ratio, count).
    Restituisce (tuning_info: Optional[dict], scale_steps: List[Tuple[int, float, int]]).
    tuning_info: {'name': str, 'avg_cents_error': float, 'params': {...}}
    scale_steps: lista di tuple (indice_step, ratio in [1,2), count)
    IT/EN comments included.
    """
    try:
        import math as _math
    except Exception:  # pragma: no cover
        _math = math
    # Converti cluster in cents [0,1200)
    centers = []  # (cents, count)
    for item in (cluster_centers or []):
        try:
            r, c = item
            if r and r > 0:
                cents = 1200.0 * _math.log2(float(r))
                # riduci in [0,1200)
                cents = cents % 1200.0
                centers.append((float(cents), int(c)))
        except Exception:
            continue
    if not centers:
        return None, []

    def circ_diff(a: float, b: float, mod: float = 1200.0) -> float:
        d = abs(float(a) - float(b)) % mod
        return d if d <= (mod * 0.5) else (mod - d)

    def score_candidate(steps_cents: List[float]) -> Tuple[float, List[Tuple[int, float, int]]]:
        """Valuta una candidata lista di passi (in cents, ridotti in [0,1200)).
        Ritorna (errore_medio_pesato, mapping) dove mapping = [(idx, ratio, count)]."""
        if not steps_cents:
            return float('inf'), []
        sc = sorted([(s % 1200.0) for s in steps_cents])
        total_werr = 0.0
        total_w = 0
        mapping: List[Tuple[int, float, int]] = []
        for (ci, cnt) in centers:
            try:
                # nearest index in circular metric
                idx = min(range(len(sc)), key=lambda k: circ_diff(ci, sc[k], 1200.0))
            except ValueError:
                continue
            err = circ_diff(ci, sc[idx], 1200.0)
            total_werr += float(err) * max(1, int(cnt))
            total_w += max(1, int(cnt))
            # store ratio folded in [1,2)
            ratio = 2.0 ** ((sc[idx] % 1200.0) / 1200.0)
            mapping.append((int(idx), float(ratio), int(cnt)))
        avg = total_werr / float(total_w or 1)
        return avg, mapping

    best_name = None
    best_err = float('inf')
    best_map: List[Tuple[int, float, int]] = []
    best_params = {}

    # Candidati noti / Known families (for continuity)
    try:
        tet12 = [k * (1200.0 / 12.0) for k in range(12)]
        err, mp = score_candidate(tet12)
        if err < best_err:
            best_name, best_err, best_map = "12-TET", err, mp
            best_params = {"n": 12}
    except Exception:
        pass
    # Pitagorico 12 / Pyth-12
    try:
        from fractions import Fraction as _Fr
        p12_rat = [reduce_to_octave(pow_fraction(_Fr(3, 2), k)) for k in range(12)]
        p12 = [1200.0 * _math.log2(float(r)) for r in normalize_ratios(p12_rat, reduce_octave=True)]
        err, mp = score_candidate(p12)
        if err < best_err:
            best_name, best_err, best_map = "Pythagorean-12", err, mp
            best_params = {}
    except Exception:
        pass
    # Pitagorico 7 / Pyth-7
    try:
        from fractions import Fraction as _Fr2
        diat = [_Fr2(1,1), _Fr2(9,8), _Fr2(81,64), _Fr2(4,3), _Fr2(3,2), _Fr2(27,16), _Fr2(243,128)]
        diat = [reduce_to_octave(r) for r in diat]
        p7 = sorted([1200.0 * _math.log2(float(r)) for r in diat])
        err, mp = score_candidate(p7)
        if err < best_err:
            best_name, best_err, best_map = "Pythagorean-7", err, mp
            best_params = {}
    except Exception:
        pass

    # n-TET generico / General n-TET (5..72)
    for n in range(5, 73):
        try:
            steps = [k * (1200.0 / float(n)) for k in range(n)]
            err, mp = score_candidate(steps)
            if err < best_err:
                best_name, best_err, best_map = f"{n}-TET", err, mp
                best_params = {"n": n}
        except Exception:
            continue

    # Rank-1 con generatore / Rank-1 generated with free generator
    periods = [1200.0, 1901.955000865, 2400.0]  # octave, tritave, double octave
    for P in periods:
        try:
            # Limit generator search so steps <= 72 and not too small
            g_min = max(20.0, P / 72.0)
            g_max = max(g_min + 1.0, min(1200.0, P / 3.0))
            step_cents = 1.0  # 1 cent resolution
            g = g_min
            while g <= g_max + 1e-9:
                try:
                    max_k = int(_math.floor(P / g)) + 1
                    steps = [(k * g) % 1200.0 for k in range(max(1, max_k))]
                    # Dedup near-equal cents
                    steps_u = []
                    for s in sorted(steps):
                        if not steps_u or abs(s - steps_u[-1]) > 1e-6:
                            steps_u.append(s)
                    err, mp = score_candidate(steps_u)
                    if err < best_err:
                        best_name, best_err, best_map = (f"Rank-1 gen={g:.2f}c period={P:.2f}c" if P != 1200.0 else f"Rank-1 gen={g:.2f}c"), err, mp
                        best_params = {"generator_cents": round(g, 4), "period_cents": round(P, 4)}
                except Exception:
                    pass
                g += step_cents
        except Exception:
            continue

    # Prepara output / Prepare output
    if best_name is None:
        return None, []
    # Normalizza: assicurati che esista step indice 0 a ratio 1.0
    try:
        has_zero = any(int(s) == 0 for (s, _, __) in best_map)
    except Exception:
        has_zero = False
    steps_out = list(best_map)
    if has_zero:
        steps_out = [(0 if int(s)==0 else int(s), (1.0 if int(s)==0 else float(r)), int(c)) for (s,r,c) in steps_out]
    else:
        steps_out = [(0, 1.0, 0)] + steps_out
    steps_out = sorted(steps_out, key=lambda t: (int(t[0]), float(t[1])))
    info = {"name": best_name, "avg_cents_error": float(best_err), "params": best_params}
    return info, steps_out


def infer_tuning_system_with_comparative(cluster_centers: List[Tuple[float, int]]):
    """
    Come infer_tuning_system_general ma ritorna anche una soluzione comparativa
    di famiglia diversa: (primary_info, primary_steps, comp_info, comp_steps).
    Famiglie: 'TET', 'Pyth', 'Rank1'.
    """
    try:
        import math as _math
    except Exception:
        _math = math

    # Prepara centers in cents [0,1200)
    centers = []
    for item in (cluster_centers or []):
        try:
            r, c = item
            if r and r > 0:
                cents = 1200.0 * _math.log2(float(r))
                centers.append((float(cents % 1200.0), int(c)))
        except Exception:
            continue
    if not centers:
        return None, [], None, []

    def circ_diff(a: float, b: float, mod: float = 1200.0) -> float:
        d = abs(float(a) - float(b)) % mod
        return d if d <= (mod * 0.5) else (mod - d)

    def score_candidate(steps_cents: List[float]):
        if not steps_cents:
            return float('inf'), []
        sc = sorted([(s % 1200.0) for s in steps_cents])
        total_werr = 0.0
        total_w = 0
        mapping: List[Tuple[int, float, int]] = []
        for (ci, cnt) in centers:
            try:
                idx = min(range(len(sc)), key=lambda k: circ_diff(ci, sc[k], 1200.0))
            except ValueError:
                continue
            err = circ_diff(ci, sc[idx], 1200.0)
            total_werr += float(err) * max(1, int(cnt))
            total_w += max(1, int(cnt))
            ratio = 2.0 ** ((sc[idx] % 1200.0) / 1200.0)
            mapping.append((int(idx), float(ratio), int(cnt)))
        avg = total_werr / float(total_w or 1)
        return avg, mapping

    def norm_steps_out(map_in: List[Tuple[int, float, int]]):
        try:
            has_zero = any(int(s) == 0 for (s, _, __) in map_in)
        except Exception:
            has_zero = False
        steps = list(map_in)
        if has_zero:
            steps = [(0 if int(s)==0 else int(s), (1.0 if int(s)==0 else float(r)), int(c)) for (s,r,c) in steps]
        else:
            steps = [(0, 1.0, 0)] + steps
        return sorted(steps, key=lambda t: (int(t[0]), float(t[1])))

    cands = []  # list of dicts with name,family,err,map,params

    # TETs (including 12-TET)
    for n in range(5, 73):
        try:
            steps = [k * (1200.0 / float(n)) for k in range(n)]
            err, mp = score_candidate(steps)
            cands.append({"name": f"{n}-TET", "family": "TET", "err": err, "map": mp, "params": {"n": n}})
        except Exception:
            pass

    # Pythagorean families
    try:
        from fractions import Fraction as _Fr
        p12_rat = [reduce_to_octave(pow_fraction(_Fr(3, 2), k)) for k in range(12)]
        p12 = [1200.0 * _math.log2(float(r)) for r in normalize_ratios(p12_rat, reduce_octave=True)]
        err, mp = score_candidate(p12)
        cands.append({"name": "Pythagorean-12", "family": "Pyth", "err": err, "map": mp, "params": {}})
    except Exception:
        pass
    try:
        from fractions import Fraction as _Fr2
        diat = [_Fr2(1,1), _Fr2(9,8), _Fr2(81,64), _Fr2(4,3), _Fr2(3,2), _Fr2(27,16), _Fr2(243,128)]
        diat = [reduce_to_octave(r) for r in diat]
        p7 = sorted([1200.0 * _math.log2(float(r)) for r in diat])
        err, mp = score_candidate(p7)
        cands.append({"name": "Pythagorean-7", "family": "Pyth", "err": err, "map": mp, "params": {}})
    except Exception:
        pass

    # Rank-1 generated systems (octave, tritave, double octave)
    periods = [1200.0, 1901.955000865, 2400.0]
    for P in periods:
        try:
            g_min = max(20.0, P / 72.0)
            g_max = max(g_min + 1.0, min(1200.0, P / 3.0))
            g = g_min
            while g <= g_max + 1e-9:
                try:
                    max_k = int(_math.floor(P / g)) + 1
                    steps = [(k * g) % 1200.0 for k in range(max(1, max_k))]
                    steps_u = []
                    for s in sorted(steps):
                        if not steps_u or abs(s - steps_u[-1]) > 1e-6:
                            steps_u.append(s)
                    err, mp = score_candidate(steps_u)
                    name = (f"Rank-1 gen={g:.2f}c period={P:.2f}c" if P != 1200.0 else f"Rank-1 gen={g:.2f}c")
                    cands.append({"name": name, "family": "Rank1", "err": err, "map": mp, "params": {"generator_cents": round(g, 4), "period_cents": round(P, 4)}})
                except Exception:
                    pass
                g += 1.0
        except Exception:
            pass

    if not cands:
        return None, [], None, []

    # Pick primary and complementary comparative
    primary = min(cands, key=lambda d: float(d.get("err", float('inf'))))
    comp = None
    fam = primary.get("family")
    # Complementary rule: TET <-> Rank1; Pyth -> TET (fallback: any different family)
    try:
        if fam == "TET":
            pool = [d for d in cands if d.get("family") == "Rank1"]
        elif fam == "Rank1":
            pool = [d for d in cands if d.get("family") == "TET"]
        else:
            pool = [d for d in cands if d.get("family") == "TET"]
        if not pool:
            pool = [d for d in cands if d.get("family") != fam]
        comp = min(pool, key=lambda d: float(d.get("err", float('inf')))) if pool else None
    except Exception:
        comp = None

    prim_info = {"name": primary["name"], "avg_cents_error": float(primary["err"]), "params": primary.get("params", {})}
    prim_steps = norm_steps_out(primary["map"])

    if comp is not None:
        comp_info = {"name": comp["name"], "avg_cents_error": float(comp["err"]), "params": comp.get("params", {})}
        comp_steps = norm_steps_out(comp["map"])
    else:
        comp_info, comp_steps = None, []

    return prim_info, prim_steps, comp_info, comp_steps


# --- Scala (.scl) parsing and matching utilities ---
# IT: Parser minimale per file .scl (Scala) e matching con i centroidi.
# EN: Minimal parser for .scl (Scala) files and matching with cluster centers.

def parse_scl_file(path: str) -> Optional[dict]:
    """Parsa un file .scl secondo il formato Huygens-Fokker di base.
    Ritorna un dict: {'name': str, 'degrees_cents': List[float], 'file': str}
    dove degrees_cents sono ridotti a [0,1200).
    Se il file non è leggibile/parsabile, ritorna None.
    """
    try:
        with open(path, 'r', encoding='utf-8', errors='ignore') as f:
            lines = f.readlines()
    except Exception:
        return None

    # Rimuovi commenti e righe vuote; in .scl i commenti iniziano con '!'
    cleaned: List[str] = []
    for ln in lines:
        s = ln.strip()
        if not s:
            continue
        # drop lines that start with '!' entirely
        if s.startswith('!'):
            # but keep the first comment line as possible filename/description if name missing
            cleaned.append(s)
            continue
        # remove trailing inline comments starting with '!' if any
        if '!' in s:
            s = s.split('!', 1)[0].strip()
        if s:
            cleaned.append(s)

    if not cleaned:
        return None

    # The format: optional leading comment lines starting with '!'
    # Next non-comment line is the description (name). Following line has the number of notes.
    idx = 0
    # Skip leading comments for description, but remember them if needed
    while idx < len(cleaned) and cleaned[idx].startswith('!'):
        idx += 1
    if idx >= len(cleaned):
        return None
    name_line = cleaned[idx]
    idx += 1
    # Some files may keep description empty; fallback to filename later
    desc = name_line.strip()

    # Read number of notes
    if idx >= len(cleaned):
        return None
    try:
        n_deg = int(str(cleaned[idx]).strip().split()[0])
    except Exception:
        return None
    idx += 1

    degrees_cents: List[float] = []
    # Read next n_deg lines as degrees (can include comments after value that we already stripped)
    for _ in range(n_deg):
        if idx >= len(cleaned):
            break
        val = cleaned[idx].strip()
        idx += 1
        if not val:
            continue
        token = val.split()[0]
        cents_val: Optional[float] = None
        # ratio like 3/2 or 1.5/1.0 is allowed
        if '/' in token:
            try:
                num_str, den_str = token.split('/', 1)
                num = float(Fraction(num_str.strip()))
                den = float(Fraction(den_str.strip()))
                r = num / den if den != 0 else None
                if r and r > 0:
                    cents_val = 1200.0 * math.log2(float(r))
            except Exception:
                cents_val = None
        else:
            # plain number in cents
            try:
                cents_val = float(token)
            except Exception:
                cents_val = None
        if isinstance(cents_val, (int, float)) and math.isfinite(cents_val):
            # Reduce to [0,1200)
            c = float(cents_val) % 1200.0
            degrees_cents.append(c)
    if not degrees_cents:
        return None

    # Ensure unique sorted degrees
    degrees_sorted: List[float] = []
    for c in sorted(degrees_cents):
        if not degrees_sorted or abs(c - degrees_sorted[-1]) > 1e-6:
            degrees_sorted.append(c)

    return {
        'name': desc if desc else os.path.basename(path),
        'degrees_cents': degrees_sorted,
        'file': os.path.basename(path),
    }


def load_scales_from_dir(dir_path: str = 'scl') -> List[dict]:
    """Carica tutti i file .scl leggibili dalla directory indicata.
    Ritorna una lista di dict come parse_scl_file.
    """
    try:
        if not os.path.isdir(dir_path):
            return []
    except Exception:
        return []
    out: List[dict] = []
    try:
        for fn in os.listdir(dir_path):
            if not fn.lower().endswith('.scl'):
                continue
            fp = os.path.join(dir_path, fn)
            info = parse_scl_file(fp)
            if info and isinstance(info.get('degrees_cents'), list) and info['degrees_cents']:
                out.append(info)
    except Exception:
        return out
    return out


def match_scala_scales(cluster_centers: List[Tuple[float, int]], dir_path: str = 'scl') -> Tuple[Optional[dict], List[Tuple[int,float,int]]]:
    """Trova la scala .scl più vicina ai centroidi (ratio,count).
    - Converte i centroidi in cents [0,1200) e per ciascuno trova il grado più vicino della scala
      usando distanza circolare modulo 1200.
    - Ritorna (best_info, steps_map) dove best_info include 'name','file','avg_cents_error'.
    - steps_map: [(index, ratio_from_degree, count)].
    Se nessuna scala è disponibile, ritorna (None, []).
    """
    scales = load_scales_from_dir(dir_path)
    if not scales:
        return None, []

    # Prepare centers in cents
    centers_c = []  # (cents, count)
    for item in (cluster_centers or []):
        try:
            r, cnt = item
            if r and r > 0:
                c = (1200.0 * math.log2(float(r))) % 1200.0
                centers_c.append((float(c), int(cnt)))
        except Exception:
            continue
    if not centers_c:
        return None, []

    def circ_diff(a: float, b: float, mod: float = 1200.0) -> float:
        d = abs(float(a) - float(b)) % mod
        return d if d <= (mod * 0.5) else (mod - d)

    best = None  # tuple(err, scale_info, mapping)
    for sc in scales:
        degs = sc.get('degrees_cents') or []
        if not degs:
            continue
        total_werr = 0.0
        total_w = 0
        mapping: List[Tuple[int,float,int]] = []
        for (ci, cnt) in centers_c:
            try:
                idx = min(range(len(degs)), key=lambda k: circ_diff(ci, degs[k], 1200.0))
            except ValueError:
                continue
            err = circ_diff(ci, degs[idx], 1200.0)
            total_werr += float(err) * max(1, int(cnt))
            total_w += max(1, int(cnt))
            ratio = 2.0 ** ((float(degs[idx]) % 1200.0) / 1200.0)
            mapping.append((int(idx), float(ratio), int(cnt)))
        if total_w == 0:
            continue
        avg = total_werr / float(total_w)
        if best is None or avg < best[0]:
            best = (avg, sc, mapping)

    if best is None:
        return None, []

    err, sc, mapping = best
    info = {'name': sc.get('name', ''), 'file': sc.get('file', ''), 'avg_cents_error': float(err)}
    # Normalize mapping to ensure index 0 & ratio 1.0 present
    try:
        has_zero = any(int(s) == 0 for (s, _, __) in mapping)
    except Exception:
        has_zero = False
    steps = list(mapping)
    if has_zero:
        steps = [(0 if int(s)==0 else int(s), (1.0 if int(s)==0 else float(r)), int(c)) for (s,r,c) in steps]
    else:
        steps = [(0, 1.0, 0)] + steps
    steps = sorted(steps, key=lambda t: (int(t[0]), float(t[1])))
    return info, steps


def match_scala_scales_topk(cluster_centers: List[Tuple[float, int]], dir_path: str = 'scl', k: int = 5) -> List[Tuple[dict, List[Tuple[int,float,int]]]]:
    """Ritorna le migliori k scale .scl per i centroidi dati.
    Output: lista di (info_dict, steps_map) ordinata per errore crescente.
    Mantiene la stessa logica di punteggio di match_scala_scales.
    """
    scales = load_scales_from_dir(dir_path)
    if not scales:
        return []
    # Prepara centers in cents
    centers_c = []  # (cents, count)
    for item in (cluster_centers or []):
        try:
            r, cnt = item
            if r and r > 0:
                c = (1200.0 * math.log2(float(r))) % 1200.0
                centers_c.append((float(c), int(cnt)))
        except Exception:
            continue
    if not centers_c:
        return []

    def circ_diff(a: float, b: float, mod: float = 1200.0) -> float:
        d = abs(float(a) - float(b)) % mod
        return d if d <= (mod * 0.5) else (mod - d)

    scored: List[Tuple[float, dict, List[Tuple[int,float,int]]]] = []
    for sc in scales:
        degs = sc.get('degrees_cents') or []
        if not degs:
            continue
        total_werr = 0.0
        total_w = 0
        mapping: List[Tuple[int,float,int]] = []
        for (ci, cnt) in centers_c:
            try:
                idx = min(range(len(degs)), key=lambda k: circ_diff(ci, degs[k], 1200.0))
            except ValueError:
                continue
            err = circ_diff(ci, degs[idx], 1200.0)
            total_werr += float(err) * max(1, int(cnt))
            total_w += max(1, int(cnt))
            ratio = 2.0 ** ((float(degs[idx]) % 1200.0) / 1200.0)
            mapping.append((int(idx), float(ratio), int(cnt)))
        if total_w == 0:
            continue
        avg = total_werr / float(total_w)
        info = {'name': sc.get('name', ''), 'file': sc.get('file', ''), 'avg_cents_error': float(avg)}
        # Normalize mapping to ensure index 0 & ratio 1.0 present
        try:
            has_zero = any(int(s) == 0 for (s, _, __) in mapping)
        except Exception:
            has_zero = False
        steps = list(mapping)
        if has_zero:
            steps = [(0 if int(s)==0 else int(s), (1.0 if int(s)==0 else float(r)), int(c)) for (s,r,c) in steps]
        else:
            steps = [(0, 1.0, 0)] + steps
        steps = sorted(steps, key=lambda t: (int(t[0]), float(t[1])))
        scored.append((float(avg), info, steps))

    scored.sort(key=lambda x: float(x[0]))
    top = scored[:max(0, int(k))] if scored else []
    return [(info, steps) for (_err, info, steps) in top]


def match_scala_scales_within_threshold(cluster_centers: List[Tuple[float, int]], dir_path: str = 'scl', threshold_cents: float = 0.0) -> List[dict]:
    """Restituisce tutte le scale .scl con errore medio <= threshold_cents.
    Output: lista di dict {'name','file','avg_cents_error'} ordinata per errore crescente.
    """
    try:
        thr = float(threshold_cents)
    except Exception:
        return []
    if not math.isfinite(thr) or thr < 0:
        return []
    # Riutilizza il punteggio come in topk
    scales = load_scales_from_dir(dir_path)
    if not scales:
        return []
    centers_c = []
    for item in (cluster_centers or []):
        try:
            r, cnt = item
            if r and r > 0:
                c = (1200.0 * math.log2(float(r))) % 1200.0
                centers_c.append((float(c), int(cnt)))
        except Exception:
            continue
    if not centers_c:
        return []
    def circ_diff(a: float, b: float, mod: float = 1200.0) -> float:
        d = abs(float(a) - float(b)) % mod
        return d if d <= (mod * 0.5) else (mod - d)
    scored: List[Tuple[float, dict]] = []
    for sc in scales:
        degs = sc.get('degrees_cents') or []
        if not degs:
            continue
        total_werr = 0.0
        total_w = 0
        for (ci, cnt) in centers_c:
            try:
                idx = min(range(len(degs)), key=lambda k: circ_diff(ci, degs[k], 1200.0))
            except ValueError:
                continue
            err = circ_diff(ci, degs[idx], 1200.0)
            total_werr += float(err) * max(1, int(cnt))
            total_w += max(1, int(cnt))
        if total_w == 0:
            continue
        avg = total_werr / float(total_w)
        if avg <= thr:
            scored.append((float(avg), {'name': sc.get('name',''), 'file': sc.get('file',''), 'avg_cents_error': float(avg)}))
    scored.sort(key=lambda x: float(x[0]))
    return [info for (_err, info) in scored]


def estimate_diapason_and_ratios(audio_path: str,
                                   base_hint_hz: float,
                                   initial_a4_hz: float = 440.0,
                                   frame_size: int = 1024,
                                   hop_size: int = 512,
                                   use_hq: bool = False,
                                   scala_cent: Optional[float] = None) -> Optional[dict]:
    """Stima il diapason (A4) e raggruppa i rapporti dai tracciati F0.
    - Non assume 12-TET o dizionari storici.
    - Usa YIN/pYIN per estrarre un vettore F0, poi ripiega le frequenze nell'intorno dell'A4.
    - Calcola i rapporti rispetto a base_hint_hz (basenote) e li riduce in [1,2).
    - Raggruppa i rapporti per prossimità relativa (~2%).
    Ritorna dict: { 'diapason_est': float, 'f0_list': [..], 'ratios_reduced': [..], 'ratio_clusters': [(center, count)] }
    """
    try:
        import numpy as _np
        import librosa as _lb
    except (ImportError, ModuleNotFoundError):
        return None

    try:
        y, sr = _lb.load(audio_path, sr=None, mono=True)
    except (FileNotFoundError, ValueError, RuntimeError):
        return None

    # Estrai F0: CREPE se use_hq, altrimenti pYIN/YIN
    try:
        fmin = 50.0
        fmax = max(2000.0, sr / 4.0)
        # Auto-adapt fmin so at least two periods of fmin fit into the frame
        required_min_fmin = (2.0 * sr) / float(frame_size) if frame_size and sr else fmin
        fmin_eff = fmin
        try:
            fmin_eff = max(float(fmin), float(required_min_fmin) + 1e-6)
        except Exception:
            fmin_eff = fmin
        if fmin_eff >= fmax:
            fmin_eff = 0.5 * fmax
        if fmin_eff > fmin:
            print(L(
                f"Adatto automaticamente fmin da {fmin:.3f} a {fmin_eff:.3f} Hz per frame_length={frame_size}",
                f"Auto-adapting fmin from {fmin:.3f} to {fmin_eff:.3f} Hz for frame_length={frame_size}"
            ))
        fmin = fmin_eff
        if use_hq:
            try:
                import crepe as _cr
                y16 = _lb.resample(y, orig_sr=sr, target_sr=16000) if sr != 16000 else y
                _t, f_c, conf_c, _act = _cr.predict(y16, 16000, viterbi=True, step_size=10, model_capacity='full', verbose=0)
                mask = conf_c > 0.5
                f0_vals = _np.asarray(f_c[mask], dtype=float)
            except Exception:
                # Fallback pYIN/YIN
                try:
                    f0_series, _, _ = _lb.pyin(y, fmin=fmin, fmax=fmax, frame_length=frame_size, hop_length=hop_size, sr=sr)
                    f0_vals = _np.asarray([v for v in _np.nan_to_num(f0_series, nan=_np.nan) if _np.isfinite(v) and v > 0])
                except Exception:
                    f0_track = _lb.yin(y, fmin=fmin, fmax=fmax, frame_length=frame_size, hop_length=hop_size, sr=sr)
                    f0_vals = _np.asarray([v for v in f0_track if _np.isfinite(v) and v > 0])
        else:
            try:
                f0_series, _, _ = _lb.pyin(y, fmin=fmin, fmax=fmax, frame_length=frame_size, hop_length=hop_size, sr=sr)
                f0_vals = _np.asarray([v for v in _np.nan_to_num(f0_series, nan=_np.nan) if _np.isfinite(v) and v > 0])
            except Exception:
                f0_track = _lb.yin(y, fmin=fmin, fmax=fmax, frame_length=frame_size, hop_length=hop_size, sr=sr)
                f0_vals = _np.asarray([v for v in f0_track if _np.isfinite(v) and v > 0])
    except Exception:
        f0_vals = _np.array([], dtype=float)

    # Deduplicate F0 list (preserve order, round to 0.1 Hz)
    f0_list = []
    if f0_vals.size > 0:
        seen_keys = set()
        for v in map(float, f0_vals.tolist()):
            key = round(v, 1)
            if key not in seen_keys and v > 0:
                seen_keys.add(key)
                f0_list.append(v)
    if not f0_list:
        return None

    # Ripiega le F0 nell'intorno dell'A4
    def fold_to_band(f: float, low: float, high: float) -> float:
        if f <= 0:
            return f
        while f < low:
            f *= 2.0
        while f >= high:
            f /= 2.0
        return f

    band_low = 220.0
    band_high = 880.0
    folded = [fold_to_band(f, band_low, band_high) for f in f0_list]

    # Privilegia l'intorno 360..520; se vuoto usa tutto il fold
    focus = [f for f in folded if 360.0 <= f <= 520.0]
    cand = focus if focus else folded

    import statistics as _stats
    try:
        a4_est = float(_stats.median(cand))
    except Exception:
        a4_est = float(initial_a4_hz if initial_a4_hz and initial_a4_hz > 0 else 440.0)

    # Calcola rapporti rispetto alla basenote indicata
    base = float(base_hint_hz) if base_hint_hz and base_hint_hz > 0 else a4_est
    ratios = [f / base for f in f0_list if f > 0 and base > 0]

    # Riduci in [1,2)
    def reduce_ratio(r: float) -> float:
        if r <= 0:
            return r
        while r < 1.0:
            r *= 2.0
        while r >= 2.0:
            r /= 2.0
        return r

    ratios_reduced = [reduce_ratio(r) for r in ratios]

    # Cluster per prossimità relativa 2%
    if ratios_reduced:
        rr_sorted = sorted(ratios_reduced)
        clusters = []  # each as (sum, count)
        centers = []
        tol = 0.02
        cur_center = rr_sorted[0]
        cur_sum = rr_sorted[0]
        cur_count = 1
        for val in rr_sorted[1:]:
            if abs(val - cur_center) <= tol * cur_center:
                cur_sum += val
                cur_count += 1
                cur_center = cur_sum / cur_count
            else:
                centers.append((cur_center, cur_count))
                cur_center = val
                cur_sum = val
                cur_count = 1
        centers.append((cur_center, cur_count))
    else:
        centers = []

    # Infer tuning system using comparative fitter (primary + alternative)
    prim_info, prim_steps, comp_info, comp_steps = infer_tuning_system_with_comparative(centers)
    tuning_info, scale_steps = prim_info, prim_steps
    tuning_comp, scale_steps_comp = comp_info, comp_steps
    if tuning_info is None:
        # Fallback: no inference possible
        tuning_info, scale_steps = None, []
    # Comparative may be None/[]

    # Stima basenote: classe 12-TET più frequente dai F0 (rispetto ad A4 stimato)
    bn_midi = None
    bn_name = None
    bn_hz = None
    try:
        mids = []
        if a4_est and a4_est > 0:
            for f in f0_list:
                if f and f > 0:
                    m = int(round(MIDI_A4 + SEMITONES_PER_OCTAVE * math.log2(float(f) / float(a4_est))))
                    m = max(MIDI_MIN, min(MIDI_MAX, m))
                    mids.append(m)
        if mids:
            # istogramma classi mod 12
            counts = [0] * 12
            for m in mids:
                counts[m % 12] += 1
            best_class = int(max(range(12), key=lambda k: counts[k])) if any(counts) else (mids[0] % 12)
            # mediana delle note appartenenti alla classe
            in_class = [m for m in mids if (m % 12) == best_class]
            try:
                import statistics as _stats2
                m_med = int(round(_stats2.median(in_class if in_class else mids)))
            except Exception:
                m_med = int(round(sum(in_class if in_class else mids) / float(len(in_class if in_class else mids))))
            m_med = max(MIDI_MIN, min(MIDI_MAX, m_med))
            bn_midi = m_med
            bn_hz = float(a4_est) * (2.0 ** ((bn_midi - MIDI_A4) / 12.0))
            try:
                bn_name = midi_to_note_name_12tet(bn_midi)
            except Exception:
                bn_name = None
    except Exception:
        pass

    # Scala (.scl) best-match from 'scl' directory based on ratio cluster centers
    try:
        scala_info, scala_steps = match_scala_scales(centers, dir_path='scl')
    except Exception:
        scala_info, scala_steps = (None, [])
    # Also compute top-5 Scala matches
    try:
        _topk = match_scala_scales_topk(centers, dir_path='scl', k=5)
        scala_top_matches = [{'name': info.get('name',''), 'file': info.get('file',''), 'avg_cents_error': float(info.get('avg_cents_error', 0.0))} for (info, _steps) in _topk]
    except Exception:
        scala_top_matches = []
    # Optionally compute all matches within threshold cents
    scala_within_matches = []
    scala_within_threshold = None
    try:
        if isinstance(scala_cent, (int, float)) and math.isfinite(scala_cent) and float(scala_cent) >= 0:
            scala_within_matches = match_scala_scales_within_threshold(centers, dir_path='scl', threshold_cents=float(scala_cent))
            scala_within_threshold = float(scala_cent)
    except Exception:
        scala_within_matches = []
        scala_within_threshold = None

    return {
        'diapason_est': a4_est,
        'f0_list': f0_list,
        'ratios_reduced': ratios_reduced,
        'ratio_clusters': centers,
        'base_hint_hz': base,
        'tuning_inferred': tuning_info,
        'scale_steps': scale_steps,
        'tuning_comparative': tuning_comp,
        'scale_steps_comp': scale_steps_comp,
        'basenote_est_hz': bn_hz,
        'basenote_est_midi': bn_midi,
        'basenote_est_name_12tet': bn_name,
        # Scala match results
        'scala_match_info': scala_info,
        'scala_match_steps': scala_steps,
        'scala_top_matches': scala_top_matches,
        'scala_within_matches': scala_within_matches,
        'scala_within_threshold_cents': scala_within_threshold,
    }


def export_diapason_plot_png(output_base: str, analysis_result: dict, basenote_hz: float, diapason_hz: float) -> Optional[str]:
    """Esporta un grafico PNG dell'andamento della F0 nel tempo con quattro riferimenti:
    1) 12-TET (Hz), 2) Midicents, 3) Pitagorico 12, 4) Pitagorico 7.
    Salva come f"{output_base}_diapason.png". Ritorna il path o None in caso di errore.
    """
    try:
        import matplotlib.pyplot as _plt
        import numpy as _np
    except Exception as e:
        print(L(f"matplotlib non disponibile per l'export PNG: {e}", "matplotlib not available for PNG export: {e}"))
        return None

    if not isinstance(analysis_result, dict):
        return None
    f0_series = analysis_result.get('f0_series') or analysis_result.get('f0_list') or []
    if not f0_series:
        return None
    # Times
    t = analysis_result.get('f0_times') or []
    try:
        t = [float(x) for x in t]
    except Exception:
        t = []
    n = len(f0_series)
    f = _np.asarray([float(x) for x in f0_series], dtype=float)
    if not t or len(t) != n:
        # fallback: assume 10ms hop
        t = ( _np.arange(n, dtype=float) * 0.01 ).tolist()

    # Helper to build reference frequencies
    def tet12_ratios():
        return [2.0 ** (k/12.0) for k in range(12)]
    def pyth12_ratios():
        try:
            from fractions import Fraction as _Fr
            vals = [reduce_to_octave(pow_fraction(_Fr(3,2), k)) for k in range(12)]
            return normalize_ratios(vals, reduce_octave=True)
        except Exception:
            return normalize_ratios([(1.5**k) for k in range(12)], reduce_octave=True)
    def pyth7_ratios():
        try:
            from fractions import Fraction as _Fr
            diat = [_Fr(1,1), _Fr(9,8), _Fr(81,64), _Fr(4,3), _Fr(3,2), _Fr(27,16), _Fr(243,128)]
            diat = [reduce_to_octave(r) for r in diat]
            return sorted([float(r) for r in diat])
        except Exception:
            return sorted([1.0, 9/8, 81/64, 4/3, 3/2, 27/16, 243/128])

    tet_refs = [float(basenote_hz) * r for r in tet12_ratios()]
    py12_refs = [float(basenote_hz) * r for r in pyth12_ratios()]
    py7_refs  = [float(basenote_hz) * r for r in pyth7_ratios()]

    # Prepare figure
    fig, axes = _plt.subplots(4, 1, figsize=(12, 14), sharex=True)

    # 1) 12-TET (Hz)
    ax = axes[0]
    ax.plot(t, f, label='F0 (Hz)', color='C0', linewidth=1.2)
    for yref in tet_refs:
        ax.axhline(y=yref, color='C1', alpha=0.35, linewidth=0.8)
    ax.set_ylabel('Hz (12-TET refs)')
    ax.grid(True, alpha=0.3)
    ax.legend(loc='upper right')

    # 2) Midicents
    ax = axes[1]
    try:
        a4 = float(diapason_hz) if diapason_hz and diapason_hz > 0 else 440.0
    except Exception:
        a4 = 440.0
    mc = 1200.0 * _np.log2(_np.clip(f, 1e-6, None) / a4) + 6900.0
    ax.plot(t, mc, label='F0 (midicents)', color='C0', linewidth=1.2)
    # grid every 100 cents
    ymin, ymax = _np.nanmin(mc), _np.nanmax(mc)
    cmin = 100.0 * (int(ymin // 100.0) - 1)
    cmax = 100.0 * (int(ymax // 100.0) + 1)
    for yref in _np.arange(cmin, cmax + 100.0, 100.0):
        ax.axhline(y=yref, color='gray', alpha=0.2, linewidth=0.6)
    ax.set_ylabel('Midicents')
    ax.grid(True, alpha=0.3)
    ax.legend(loc='upper right')

    # 3) Pitagorico 12 (Hz)
    ax = axes[2]
    ax.plot(t, f, label='F0 (Hz)', color='C0', linewidth=1.2)
    for yref in py12_refs:
        ax.axhline(y=yref, color='C2', alpha=0.35, linewidth=0.8)
    ax.set_ylabel('Hz (Pyth-12 refs)')
    ax.grid(True, alpha=0.3)
    ax.legend(loc='upper right')

    # 4) Pitagorico 7 (Hz)
    ax = axes[3]
    ax.plot(t, f, label='F0 (Hz)', color='C0', linewidth=1.2)
    for yref in py7_refs:
        ax.axhline(y=yref, color='C3', alpha=0.35, linewidth=0.8)
    ax.set_ylabel('Hz (Pyth-7 refs)')
    ax.set_xlabel(L('Tempo (s)', 'Time (s)'))
    ax.grid(True, alpha=0.3)
    ax.legend(loc='upper right')

    _plt.tight_layout()
    out_path = f"{output_base}_diapason.png"
    try:
        fig.savefig(out_path, dpi=150)
        _plt.close(fig)
        print(L(f"Grafico PNG esportato: {out_path}", f"PNG plot exported: {out_path}"))
        return out_path
    except Exception as e:
        print(L(f"Errore salvataggio PNG: {e}", f"Error saving PNG: {e}"))
        try:
            _plt.close(fig)
        except Exception:
            pass
        return None


def render_analysis_to_wav(audio_path: str, analysis_result: dict, output_base: str) -> Optional[str]:
    """Renderizza un segnale sinusoidale controllato in frequenza basato sui risultati dell'analisi.
    - Usa f0_series se disponibile; altrimenti f0_list (diapason); altrimenti f0_hz costante.
    - La durata del rendering combacia con quella dell'audio sorgente.
    Ritorna il percorso del file WAV generato, oppure None in caso di errore.
    """
    try:
        import numpy as _np
        import librosa as _lb
    except Exception as e:
        print(L(f"librosa non disponibile per il render: {e}", f"librosa not available for render: {e}"))
        return None

    try:
        y, sr = _lb.load(audio_path, sr=None, mono=True)
    except Exception as e:
        print(L(f"Errore caricamento audio per render: {e}", f"Audio load error for render: {e}"))
        return None

    n_samples = len(y)
    dur = n_samples / float(sr) if sr and sr > 0 else 0.0
    if n_samples <= 0 or dur <= 0:
        print(L("Audio vuoto: render annullato", "Empty audio: render aborted"))
        return None

    # Estrai traiettoria di frequenza
    f0_traj = None
    if isinstance(analysis_result, dict):
        seq = analysis_result.get('f0_series')
        if isinstance(seq, (list, tuple)) and len(seq) > 0:
            f0_traj = [float(x) if (x is not None and x == x and x > 0) else _np.nan for x in seq]
        elif isinstance(analysis_result.get('f0_list'), (list, tuple)) and len(analysis_result.get('f0_list')) > 0:
            f0_traj = [float(x) if (x is not None and x == x and x > 0) else _np.nan for x in analysis_result.get('f0_list')]
        elif isinstance(analysis_result.get('f0_hz'), (int, float)) and float(analysis_result.get('f0_hz')) > 0:
            f0_traj = [float(analysis_result.get('f0_hz'))]

    if not f0_traj:
        print(L("Nessuna F0 disponibile per il render", "No F0 available for render"))
        return None

    f0_arr = _np.array(f0_traj, dtype=float)
    # Sostituisci NaN con interpolazione semplice
    if _np.isnan(f0_arr).any():
        idx = _np.arange(len(f0_arr))
        good = _np.isfinite(f0_arr)
        if good.any():
            f0_arr = _np.interp(idx, idx[good], f0_arr[good])
        else:
            print(L("Traiettoria F0 non valida", "Invalid F0 trajectory"))
            return None

    # Clamp frequenze a [20, 5000] Hz
    f0_arr = _np.clip(f0_arr, 20.0, 5000.0)

    # Interpola a livello per-sample sull'intervallo [0, dur]
    t_series = _np.linspace(0.0, dur, num=len(f0_arr), endpoint=False)
    t_samples = _np.arange(n_samples, dtype=float) / float(sr)
    f_per_sample = _np.interp(t_samples, t_series, f0_arr)

    # Integrazione di fase: phi[n] = phi[n-1] + 2*pi*f[n]/sr
    phase = _np.cumsum(2.0 * _np.pi * f_per_sample / float(sr))
    out = 0.2 * _np.sin(phase).astype(_np.float32)

    out_path = f"{output_base}_render.wav"

    # Scrittura WAV con fallback
    try:
        try:
            import soundfile as _sf
            _sf.write(out_path, out, sr)
        except Exception:
            from scipy.io import wavfile as _wio
            _wio.write(out_path, int(sr), (out * 32767.0).astype(_np.int16))
        print(L(f"Render audio salvato in {out_path}", f"Rendered audio saved to {out_path}"))
        return out_path
    except Exception as e:
        print(L(f"Errore scrittura WAV: {e}", f"WAV write error: {e}"))
        return None


def estimate_a4_from_analysis(analysis_result: dict, default_a4: float = DEFAULT_DIAPASON) -> Optional[float]:
    """Stima A4 (diapason) dai dati di analisi disponibili.
    Preferisce analysis_result['diapason_est']; altrimenti usa f0_series/f0_list/f0_hz
    ripiegando le frequenze nell'intervallo [220,880) e privilegiando 360–520 Hz.
    """
    try:
        v = analysis_result.get('diapason_est') if isinstance(analysis_result, dict) else None
        if isinstance(v, (int, float)) and float(v) > 0:
            return float(v)
    except Exception:
        pass
    try:
        import numpy as _np
    except Exception:
        return float(default_a4) if default_a4 and default_a4 > 0 else DEFAULT_DIAPASON
    f0s = []
    try:
        if isinstance(analysis_result, dict):
            if isinstance(analysis_result.get('f0_series'), (list, tuple)):
                f0s = [float(x) for x in analysis_result.get('f0_series') if isinstance(x, (int, float)) and float(x) > 0]
            elif isinstance(analysis_result.get('f0_list'), (list, tuple)):
                f0s = [float(x) for x in analysis_result.get('f0_list') if isinstance(x, (int, float)) and float(x) > 0]
            elif isinstance(analysis_result.get('f0_hz'), (int, float)) and float(analysis_result.get('f0_hz')) > 0:
                f0s = [float(analysis_result.get('f0_hz'))]
    except Exception:
        f0s = []
    if not f0s:
        try:
            return float(default_a4) if default_a4 and default_a4 > 0 else DEFAULT_DIAPASON
        except Exception:
            return DEFAULT_DIAPASON
    def _fold(f: float, lo: float, hi: float) -> float:
        if f <= 0:
            return f
        while f < lo:
            f *= 2.0
        while f >= hi:
            f /= 2.0
        return f
    folded = [_fold(float(f), 220.0, 880.0) for f in f0s if f and f > 0]
    focus = [f for f in folded if 360.0 <= f <= 520.0]
    cand = focus if focus else folded
    if not cand:
        try:
            return float(default_a4) if default_a4 and default_a4 > 0 else DEFAULT_DIAPASON
        except Exception:
            return DEFAULT_DIAPASON
    try:
        med = float(_np.median(_np.array(cand, dtype=float)))
        return med
    except Exception:
        try:
            return float(default_a4) if default_a4 and default_a4 > 0 else DEFAULT_DIAPASON
        except Exception:
            return DEFAULT_DIAPASON


def render_constant_tone_wav(output_base: str, freq_hz: float, duration_s: float = 30.0, sr: int = 48000) -> Optional[str]:
    """Renderizza una sinusoide costante alla frequenza data per una durata.
    Salva come f"{output_base}_diapason.wav".
    """
    try:
        if not isinstance(freq_hz, (int, float)) or float(freq_hz) <= 0:
            return None
        if not isinstance(duration_s, (int, float)) or float(duration_s) <= 0:
            duration_s = 30.0
        import numpy as _np
        n_samples = int(max(1, int(sr * float(duration_s))))
        t = _np.arange(n_samples, dtype=float) / float(sr)
        y = (0.2 * _np.sin(2.0 * _np.pi * float(freq_hz) * t)).astype(_np.float32)
        out_path = f"{output_base}_diapason.wav"
        try:
            try:
                import soundfile as _sf
                _sf.write(out_path, y, int(sr))
            except Exception:
                from scipy.io import wavfile as _wio
                _wio.write(out_path, int(sr), (y * 32767.0).astype(_np.int16))
            return out_path
        except Exception as e:
            print(L(f"Errore scrittura WAV: {e}", f"WAV write error: {e}"))
            return None
    except Exception as e:
        print(L(f"Errore render diapason: {e}", f"Diapason render error: {e}"))
        return None


def main():
    """Punto di ingresso principale / Main entry point (IT/EN)."""
    # Imposta lingua (pre-scan argv), rileva tema e applica, poi pulisci schermo e stampa banner
    global _LANG
    _LANG = _detect_lang_from_argv()
    # Theme pre-scan and apply before any styled output
    _theme_cli = _detect_theme_from_argv()
    _theme_eff = _detect_terminal_background() if _theme_cli == 'auto' else _theme_cli
    try:
        apply_theme(_theme_eff)
    except Exception:
        pass
    clear_screen()
    print_banner()

    # Custom formatter that adapts to terminal width and aligns help columns
    class ThinHelpFormatter(argparse.RawTextHelpFormatter):
        def __init__(self, prog: str):
            try:
                cols = _term_cols(100)
            except Exception:
                cols = 100
            cols = max(70, min(140, int(cols)))
            super().__init__(prog, max_help_position=28, width=cols)

    parser = argparse.ArgumentParser(
        formatter_class=ThinHelpFormatter,
        description=(
            "IT: THIN – Generatore e comparatore di sistemi di intonazione (PHI).\n"
            "Supporta 12/24/48-TET, sistemi naturali, Danielou, geometrici;\n "
            "basenote microtonale (+ - ! .), export Csound/.tun/Excel; analisi audio opzionale.\n "
            "EN: THIN – Generator and comparator of tuning systems (PHI).\n "
            "Supports 12/24/48-TET, natural, Danielou, geometric; microtonal basenote (+ - ! .);\n "
            "exports Csound/.tun/Excel; optional audio analysis.\n"
        ),
        epilog=(
            "IT - Note utili:\n"
            "• Notazione microtonale: + = +50c, - = -50c, ! = +25c, . = -25c (es.: C+4, A!3).\n"
            "• Confronto TET: --compare-tet 12|24|48; allineamento con --compare-tet-align.\n"
            "• Analisi audio: --audio-file file.wav --analysis lpc|specenv|sfft|cqt [--hq]; "
            "default: lpc; frame=1024 hop=512 lpc-order=12 window=hamming|hanning.\n"
            "• Excel: in System.xlsx F2 contiene Base_Hz; la colonna Ratio usa formula =D[n]/$F$2. "
            "Modifica Hz per aggiornare i Ratio. --convert FILE.xlsx genera .csd e .tun da Ratio o Hz/Base_Hz.\n"
            "• Delta Hz: nelle tabelle Compare trovi |DeltaHz_Harm|, |DeltaHz_Sub|, |DeltaHz_TET|, "
            "e se presenti anche |DeltaHz_F0| e |DeltaHz_Formant|.\n"
            "• Help paginato: massimo 80 righe; usa Invio per continuare, 'q' per uscire.\n"
            "• Output: out.csd, out_system.txt/.xlsx, out_compare.txt/.xlsx, .tun opzionale.\n"
            "• Render audio: --render per esportare un WAV di sinusoidi controllate in frequenza come feedback dell'analisi.\n"
            "EN - Useful notes:\n"
            "• Microtonal: + = +50c, - = -50c, ! = +25c, . = -25c (e.g., C+4, A!3).\n"
            "• TET comparison: --compare-tet 12|24|48; alignment via --compare-tet-align.\n"
            "• Audio: --audio-file file.wav --analysis lpc|specenv|sfft|cqt [--hq]; default: lpc; "
            "frame=1024 hop=512 lpc-order=12 window=hamming|hanning.\n"
            "• Excel: in System.xlsx F2 holds Base_Hz; Ratio column uses =D[n]/$F$2. Edit Hz to auto-update. "
            "--convert FILE.xlsx creates .csd and .tun from Ratio or Hz/Base_Hz.\n"
            "• Delta Hz: Compare tables include |DeltaHz_Harm|, |DeltaHz_Sub|, |DeltaHz_TET|, and if present |DeltaHz_F0|, |DeltaHz_Formant|.\n"
            "• Help paging: up to 80 rows; press Enter to continue, 'q' to quit.\n"
            "• Outputs: out.csd, out_system.txt/.xlsx, out_compare.txt/.xlsx, optional .tun.\n"
            "• Audio render: use --render to export a WAV of frequency-controlled sinusoids as auditory feedback of the analysis.\n"
            "Esempi / Examples:\n"
            "  thin.py --lang it --et 24 1200 --basenote C+4 out\n"
            "  thin.py --geometric 3/2 12 2/1 --compare-tet 48 --audio-file voce.wav --analysis lpc out\n"
            "  thin.py --danielou-all --diapason 442 --export-tun out\n"
        )
    )

    # Groups for clearer help layout
    grp_base = parser.add_argument_group(L("Base", "Base"))
    grp_tuning = parser.add_argument_group(L("Sistemi di intonazione", "Tuning systems"))
    grp_opts = parser.add_argument_group(L("Opzioni aggiuntive", "Extra options"))
    grp_cmp = parser.add_argument_group(L("Confronto", "Comparison"))
    grp_audio = parser.add_argument_group(L("Analisi audio (librosa)", "Audio analysis (librosa)"))
    grp_out = parser.add_argument_group(L("Output", "Output"))

    # Lingua / Language
    grp_base.add_argument("--lang", choices=["it", "en"], default="it",
                        help="Lingua dell'interfaccia / Interface language")
    grp_base.add_argument("--theme", choices=["auto", "dark", "light"], default="auto",
                        help=("Tema colori terminale: auto (predefinito), dark o light / "
                              "Terminal color theme: auto (default), dark or light"))
    grp_base.add_argument("-v", "--version", action="version",
                        version=f"%(prog)s {__version__}")
    grp_base.add_argument("--diapason", type=float, default=DEFAULT_DIAPASON,
                        help=(f"Diapason in Hz (default: {DEFAULT_DIAPASON}) / "
                              f"A4 reference (diapason) in Hz (default: {DEFAULT_DIAPASON})"))
    grp_base.add_argument("--basekey", type=int, default=DEFAULT_BASEKEY,
                        help=(f"Nota base MIDI (default: {DEFAULT_BASEKEY}) / "
                              f"Base MIDI note (default: {DEFAULT_BASEKEY})"))
    grp_base.add_argument("--basenote", type=note_name_or_frequency, default="C4",
                        help=(
                            "Nota di riferimento o frequenza in Hz; supporta microtoni (+ - ! .). / "
                            "Reference note or frequency in Hz; supports microtones (+ - ! .)."
                        ))

    # Sistemi di intonazione
    grp_tuning.add_argument("--et", nargs=2, type=int_or_fraction, default=(12, 1200),
                        metavar=("INDEX", "INTERVAL"),
                        help=("Temperamento equabile / Equal temperament"))
    grp_tuning.add_argument("--geometric", nargs=3, metavar=("GEN", "STEPS", "INTERVAL"),
                        help=(
                            "Sistema geometrico: INTERVAL intero=cents (es. 700); per rapporto usare 2/1 o 2.0; "
                            "accetta anche suffisso 'c' (es. 200c). / "
                            "Geometric system: INTERVAL integer=cents (e.g., 700); for ratios use 2/1 or 2.0; "
                            "also accepts 'c' suffix (e.g., 200c)."
                        ))
    grp_tuning.add_argument("--natural", nargs=2, type=int, metavar=("A_MAX", "B_MAX"),
                        help="Sistema naturale 4:5:6 / Natural system 4:5:6")
    # Nota: per esponenti negativi usare virgolette, es.: --danielou "-1,2,0"
    grp_tuning.add_argument("--danielou", action="append", type=parse_danielou_tuple,
                        default=None, help=("Sistema Danielou manuale (usa virgolette per esponenti negativi, es.: \"-1,2,0\") / "
                                           "Manual Danielou system (quote negative exponents, e.g.: \"-1,2,0\")"))
    grp_tuning.add_argument("--danielou-all", action="store_true",
                        help="Griglia completa Danielou / Full Danielou grid")

    # Opzioni aggiuntive
    grp_opts.add_argument("--no-reduce", action="store_true",
                        help="Non ridurre all'ottava / Do not apply octave reduction")
    grp_opts.add_argument("--span", "--ambitus", dest="span", type=int, default=1,
                        help="Ripetizioni dell'intervallo / Interval repetitions (span)")
    grp_opts.add_argument("--interval-zero", action="store_true",
                        help="Imposta interval=0 in cpstun / Set interval=0 in cpstun")
    grp_opts.add_argument("--export-tun", action="store_true",
                        help="Esporta file .tun / Export .tun file")
    grp_opts.add_argument("--tun-integer", action="store_true",
                        help=".tun: arrotonda i cents al valore intero più vicino (default: due decimali) / .tun: round cents to nearest integer (default: two decimals)")
    grp_opts.add_argument("--convert", metavar="FILE.xlsx", dest="convert", default=None,
                        help="Converte un foglio Excel (System/Compare) in .csd e .tun usando la colonna 'Ratio' o Hz/Base_Hz / Convert an Excel sheet to .csd and .tun using 'Ratio' or Hz/Base_Hz")
    grp_opts.add_argument("--import-tun", metavar="FILE.tun", dest="import_tun", default=None,
                        help="Importa un file .tun e salva un .txt con mappa MIDI->ratio / Import a .tun and save a .txt mapping MIDI->ratio")

    # Confronto
    grp_cmp.add_argument("--compare-fund", nargs="?", type=note_name_or_frequency,
                        default="basenote", const="basenote",
                        help=("Fondamentale per confronto / Fundamental for comparison"))
    grp_cmp.add_argument("--compare-tet", type=int, choices=[12, 24, 48], default=12,
                        help=("Divisioni del TET per confronto (12, 24 o 48). / TET divisions for comparison (12, 24 or 48)."))
    grp_cmp.add_argument("--compare-tet-align", choices=["same", "nearest"],
                        default="same", help=("Allineamento TET / TET alignment"))
    grp_cmp.add_argument("--subharm-fund", type=note_name_or_frequency,
                        default="A5", help=("Fondamentale subarmoniche / Subharmonic fundamental"))
    grp_cmp.add_argument("--midi-truncate", action="store_true",
                        help=("Forza troncamento MIDI / Force MIDI truncation"))

    # Analisi audio (librosa) / Audio analysis (librosa)
    grp_audio.add_argument("--audio-file", dest="audio_file", default=None,
                        help=("Percorso del file WAV da analizzare / Path to WAV file to analyze"))
    grp_audio.add_argument("--analysis", choices=["lpc", "specenv", "sfft", "cqt"], default="lpc",
                        help=("Metodo analisi formanti: lpc (consigliato per voce), specenv (media STFT), sfft (singola FFT), cqt (Constant-Q). / Formant analysis method: lpc (recommended for voice), specenv (STFT average), sfft (single FFT), cqt (Constant-Q)."))
    grp_audio.add_argument("--frame-size", type=int, default=1024,
                        help=("Frame size (default: 1024)"))
    grp_audio.add_argument("--hop-size", type=int, default=512,
                        help=("Hop size (default: 512)"))
    grp_audio.add_argument("--lpc-order", type=int, default=12,
                        help=("Ordine LPC (default: 12) / LPC order (default: 12)"))
    grp_audio.add_argument("--window", choices=["hamming", "hanning"], default="hamming",
                        help=("Finestra: hamming (default) o hanning / Window: hamming (default) or hanning"))
    grp_audio.add_argument("--hq", action="store_true",
                        help=("Usa CREPE per F0 ad alta qualità (monofonico). / Use CREPE for high-quality F0 (monophonic)."))
    grp_audio.add_argument("--diapason-analysis", action="store_true",
                        help=("Se usato con --audio-file: stima il diapason (A4) dall'audio e tenta l'individuazione del sistema (ratios); risultati in un foglio Excel dedicato. / If used with --audio-file: estimates diapason (A4) from audio and attempts to infer the system (ratios); results exported to a dedicated Excel sheet."))
    grp_audio.add_argument("--render", action="store_true",
                        help=("Se usato con --audio-file: esporta un WAV di sinusoidi controllate in frequenza per feedback uditivo dell'analisi. / If used with --audio-file: exports a WAV of frequency-controlled sinusoids for auditory feedback of the analysis."))
    grp_audio.add_argument("--scala-cent", type=float, default=None,
                        help=("Se impostato: elenca tutte le scale .scl con errore medio <= soglia (cents). / If set: list all .scl scales with avg error <= threshold (cents)."))

    # Output
    grp_out.add_argument("output_file", nargs="?", default=None,
                        help="File di output (default: out) / Output file (default: out)")

    # Mostra help se nessun argomento o se richiesto con -h/--help: usa pager "more"
    if len(sys.argv) == 1 or any(a in ("-h", "--help") for a in sys.argv[1:]):
        help_text = parser.format_help()
        help_text = stylize_help(help_text)
        page_lines(help_text.splitlines())
        return

    args = parser.parse_args()

    # Imposta lingua globale / Set global language
    _LANG = args.lang

    # Import .tun se richiesto
    if getattr(args, 'import_tun', None):
        import_tun_file(args.import_tun, basekey=int(args.basekey))
        return

    # Validazione base
    if isinstance(args.basenote, (int, float)) and args.basenote < 0:
        print(L("Nota di riferimento non valida", "Invalid reference note"))
        return

    if args.span is None or args.span < 1:
        args.span = 1

    # Calcola frequenza base (supporto microtoni) / Compute base frequency (microtonal support)
    if isinstance(args.basenote, float):
        basenote = args.basenote
    else:
        try:
            midi_base, cents_off = parse_note_with_microtones(str(args.basenote))
            basenote = apply_cents(convert_midi_to_hz(midi_base, args.diapason), cents_off)
        except ValueError as e:
            print(L(f"Errore conversione nota: {e}", f"Note conversion error: {e}"))
            return

    # Se richiesta conversione Excel -> .csd/.tun, esegui e termina
    if getattr(args, 'convert', None):
        ok = convert_excel_to_outputs(
            excel_path=args.convert,
            output_base=args.output_file,
            default_basekey=args.basekey,
            default_base_hz=basenote,
            diapason_hz=args.diapason,
            midi_truncate=args.midi_truncate,
            tun_integer=args.tun_integer,
        )
        return

    # Fondamentali per confronto
    def parse_fund_hz(value, default):
        if isinstance(value, str) and value.lower() == "basenote":
            return basenote
        elif isinstance(value, float):
            return value
        else:
            try:
                midi_f, cents_f = parse_note_with_microtones(str(value))
                return apply_cents(convert_midi_to_hz(midi_f, args.diapason), cents_f)
            except ValueError:
                return default

    compare_fund_hz = parse_fund_hz(args.compare_fund, basenote)
    subharm_fund_hz = parse_fund_hz(args.subharm_fund, args.diapason)


    # Processa sistema di intonazione
    result = process_tuning_system(args, basenote)
    if result is None:
        print(L("Nessun sistema di intonazione valido specificato", "No valid tuning system specified"))
        return

    ratios, interval = result

    # Applica span
    ratios_spanned = repeat_ratios(ratios, args.span, interval)
    ratios_eff, basekey_eff = ensure_midi_fit(ratios_spanned, args.basekey,
                                              args.midi_truncate)

    # Riepilogo esecuzione / Run summary
    output_base = args.output_file or "out"
    yes_label = L("sì", "yes")
    no_label = L("no", "no")

    # Descrizione sistema
    sys_desc = ""
    if args.natural:
        sys_desc = L("Sistema: Naturale 4:5:6", "System: Natural 4:5:6")
    elif args.danielou_all:
        sys_desc = L("Sistema: Danielou (griglia completa)", "System: Danielou (full grid)")
    elif args.danielou is not None:
        sys_desc = L("Sistema: Danielou (manuale)", "System: Danielou (manual)")
    elif args.geometric:
        sys_desc = L("Sistema: Geometrico", "System: Geometric")
    elif args.et:
        try:
            et_idx, et_int = args.et
            et_cents = fraction_to_cents(et_int) if is_fraction_type(et_int) else float(et_int)
        except Exception:
            et_idx = args.et[0] if args.et else 12
            et_cents = 1200
        sys_desc = L(f"Sistema: ET index={et_idx}, intervallo={et_cents}c",
                     f"System: ET index={et_idx}, interval={et_cents}c")

    # Analisi audio: linea localizzata
    if getattr(args, 'audio_file', None):
        analysis_line = L(f"Analisi audio: {yes_label} ({args.analysis}) file={args.audio_file}",
                          f"Audio analysis: {yes_label} ({args.analysis}) file={args.audio_file}")
    else:
        analysis_line = L(f"Analisi audio: {no_label}", f"Audio analysis: {no_label}")

    summary_lines = [
        L("— Riepilogo esecuzione —", "— Run summary —"),
        sys_desc,
        L(f"Nota di base (Hz): {basenote:.2f}", f"Basenote (Hz): {basenote:.2f}"),
        L(f"Diapason (A4): {args.diapason:.2f} Hz", f"Diapason (A4): {args.diapason:.2f} Hz"),
        L(f"Basekey MIDI: {basekey_eff}", f"MIDI basekey: {basekey_eff}"),
        L(f"Ripetizioni (span): {args.span}", f"Span (repetitions): {args.span}"),
        L(f"Riduzione all'ottava: {yes_label if not args.no_reduce else no_label}",
          f"Octave reduction: {yes_label if not args.no_reduce else no_label}"),
        L(f"Troncamento MIDI: {yes_label if args.midi_truncate else no_label}",
          f"MIDI truncation: {yes_label if args.midi_truncate else no_label}"),
        L(f"Confronto – fondamentale: {compare_fund_hz:.2f} Hz",
          f"Compare – fundamental: {compare_fund_hz:.2f} Hz"),
        L(f"TET: {args.compare_tet} ({args.compare_tet_align})",
          f"TET: {args.compare_tet} ({args.compare_tet_align})"),
        L(f"Subarmonica – fondamentale: {subharm_fund_hz:.2f} Hz",
          f"Subharmonic – fundamental: {subharm_fund_hz:.2f} Hz"),
        L(f"Base output: {output_base}", f"Output base: {output_base}"),
        analysis_line,
        ""
    ]
    page_lines(summary_lines)

    # Stampa tabella
    print_step_hz_table(sorted(ratios_eff), basenote)

    # Prepara dati per CSD
    if args.interval_zero:
        csd_input = ratios_spanned
        csd_interval = 0.0
    else:
        csd_input = ratios
        csd_interval = interval

    csd_ratios, csd_basekey = ensure_midi_fit(csd_input, args.basekey,
                                              args.midi_truncate)
    output_base = args.output_file or "out"
    fnum, existed = write_cpstun_table(output_base, csd_ratios,
                                       csd_basekey, basenote, csd_interval)

    # Export
    export_base = (output_base if not existed
                   else f"{output_base}_{fnum}")

    export_system_tables(export_base, ratios_eff, basekey_eff, basenote)

    # Analisi audio opzionale / Optional audio analysis
    analysis_data = None
    if getattr(args, 'audio_file', None):
        # Live spinner during audio analysis (robust, TTY-aware)
        base_msg_plain = L("Analisi audio in corso", "Audio analysis in progress")
        base_msg = f"{Style.FG_CYAN}{Style.BOLD}{base_msg_plain}{Style.RESET}" if _supports_ansi() else base_msg_plain
        try:
            cols = _term_cols(80)
        except Exception:
            cols = 80
        _stop_evt = threading.Event()

        def _spinner_worker():
            try:
                frames = list("⠋⠙⠹⠸⠼⠴⠦⠧⠇⠏") if _supports_ansi() else ['|','/','-','\\']
            except Exception:
                frames = ['|','/','-','\\']
            i = 0
            last = time.monotonic()
            while not _stop_evt.is_set():
                now = time.monotonic()
                if now - last >= 0.1:
                    last = now
                    ch = frames[i % len(frames)]
                    i += 1
                    try:
                        if sys.stdout and sys.stdout.writable():
                            msg = f"{base_msg} {ch}"
                            if len(msg) > cols - 1:
                                msg = msg[:max(0, cols - 1)]
                            print("\r" + msg + " " * max(0, cols - len(msg) - 1), end="", flush=True)
                    except Exception:
                        # degrade to sleep-only to avoid blocking
                        pass
                try:
                    time.sleep(0.02)
                except Exception:
                    break

        # print initial line
        try:
            if sys.stdout and sys.stdout.writable():
                print(base_msg + " ", end="", flush=True)
        except Exception:
            pass

        _th = threading.Thread(target=_spinner_worker, daemon=True)
        _th.start()
        try:
            analysis_data = analyze_audio(
                audio_path=args.audio_file,
                method=args.analysis,
                frame_size=args.frame_size,
                hop_size=args.hop_size,
                lpc_order=args.lpc_order,
                window_type=args.window,
                use_hq=getattr(args, 'hq', False),
            )
        finally:
            _stop_evt.set()
            try:
                _th.join(timeout=1.0)
            except Exception:
                pass
            # clear spinner line
            try:
                _clear_line(cols)
            except Exception:
                pass
            print("")  # newline after spinner
        if analysis_data is None:
            msg = L("Analisi audio non disponibile o fallita: le tabelle saranno generate senza risultati di analisi.",
                    "Audio analysis unavailable or failed: tables will be generated without analysis results.")
            styled = f"{Style.FG_YELLOW}{msg}{Style.RESET}" if _supports_ansi() else msg
            print(styled)
        else:
            msg = L("Analisi audio completata. Generazione delle tabelle di confronto...",
                    "Audio analysis completed. Generating comparison tables...")
            styled = f"{Style.FG_GREEN}{Style.BOLD}{msg}{Style.RESET}" if _supports_ansi() else msg
            print(styled)
            # Diapason analysis (optional): estimate A4 and ratios from audio without 12-TET assumptions
            if getattr(args, 'diapason_analysis', False):
                est = estimate_diapason_and_ratios(
                    audio_path=args.audio_file,
                    base_hint_hz=basenote,
                    initial_a4_hz=args.diapason,
                    frame_size=args.frame_size,
                    hop_size=args.hop_size,
                    use_hq=getattr(args, 'hq', False),
                    scala_cent=getattr(args, 'scala_cent', None),
                )
                if est:
                    if analysis_data is None:
                        analysis_data = {}
                    analysis_data.update(est)
                    try:
                        a4v = float(est.get('diapason_est')) if est.get('diapason_est') is not None else None
                    except Exception:
                        a4v = None
                    if a4v:
                        print(L(f"Diapason stimato dall'audio (A4): {a4v:.2f} Hz",
                                f"Estimated diapason from audio (A4): {a4v:.2f} Hz"))
                        # Suggested basenote from analysis (nearest 12-TET)
                        try:
                            bn_hz = est.get('basenote_est_hz')
                            bn_midi = est.get('basenote_est_midi')
                            bn_name = est.get('basenote_est_name_12tet')
                        except Exception:
                            bn_hz = bn_midi = bn_name = None
                        if (bn_hz is not None) or (bn_midi is not None) or (bn_name):
                            try:
                                hz_txt = f"{float(bn_hz):.2f} Hz" if isinstance(bn_hz, (int, float)) else ""
                            except Exception:
                                hz_txt = ""
                            it_line = f"Basenote suggerita dall'analisi: {bn_name or ''} (MIDI {bn_midi}) {hz_txt}".strip()
                            en_line = f"Suggested basenote from analysis: {bn_name or ''} (MIDI {bn_midi}) {hz_txt}".strip()
                            print(L(it_line, en_line))
                        # Print matched Scala scale if available
                        try:
                            _sc_info = est.get('scala_match_info') if isinstance(est, dict) else None
                        except Exception:
                            _sc_info = None
                        if _sc_info:
                            it_line2 = f"Scala corrispondente (scl):\n {_sc_info.get('name','')} [{_sc_info.get('file','')}] – err medio: {float(_sc_info.get('avg_cents_error', 0.0)):.2f} cents"
                            en_line2 = f"Matched Scala scale:\n {_sc_info.get('name','')} [{_sc_info.get('file','')}] – avg err: {float(_sc_info.get('avg_cents_error', 0.0)):.2f} cents"
                            print(L(it_line2, en_line2))
                            # Print Top-5 Scala matches
                            try:
                                _sc_top = est.get('scala_top_matches') if isinstance(est, dict) else None
                            except Exception:
                                _sc_top = None
                            if _sc_top:
                                # Build aligned table for top-5 (console)
                                rows = []
                                for i, item in enumerate(_sc_top[:5], start=1):
                                    try:
                                        nm = str(item.get('name',''))
                                        fl = str(item.get('file',''))
                                        er = f"{float(item.get('avg_cents_error', 0.0)):.2f}"
                                    except Exception:
                                        nm, fl, er = "", "", ""
                                    rows.append([str(i), nm, fl, er])
                                hdr = [L("Pos", "Rank"), L("Nome", "Name"), "File", L("Err (c)", "Err (c)")]
                                widths = [len(hdr[c]) for c in range(len(hdr))]
                                for r in rows:
                                    for c, val in enumerate(r):
                                        if len(val) > widths[c]:
                                            widths[c] = len(val)
                                def fmt(cols):
                                    return "  ".join(str(cols[i]).ljust(widths[i]) for i in range(len(cols)))
                                title = L("Top 5 corrispondenze Scala (.scl):", "Top 5 Scala matches (.scl):")
                                print(title)
                                if _supports_ansi():
                                    print(Style.BOLD + fmt(hdr) + Style.RESET)
                                else:
                                    print(fmt(hdr))
                                for r in rows:
                                    print(fmt(r))
                                # Print within-threshold Scala matches if requested
                                try:
                                    _sc_within = est.get('scala_within_matches') if isinstance(est, dict) else None
                                    _sc_thr = est.get('scala_within_threshold_cents') if isinstance(est, dict) else None
                                except Exception:
                                    _sc_within, _sc_thr = None, None
                                if _sc_within:
                                    rows2 = []
                                    for i, item in enumerate(_sc_within, start=1):
                                        try:
                                            nm = str(item.get('name',''))
                                            fl = str(item.get('file',''))
                                            er = f"{float(item.get('avg_cents_error', 0.0)):.2f}"
                                        except Exception:
                                            nm, fl, er = "", "", ""
                                        rows2.append([str(i), nm, fl, er])
                                    hdr2 = [L("Pos", "Rank"), L("Nome", "Name"), "File", L("Err (c)", "Err (c)")]
                                    widths2 = [len(hdr2[c]) for c in range(len(hdr2))]
                                    for r2 in rows2:
                                        for c, val in enumerate(r2):
                                            if len(val) > widths2[c]:
                                                widths2[c] = len(val)
                                    def fmt2(cols):
                                        return "  ".join(str(cols[i]).ljust(widths2[i]) for i in range(len(cols)))
                                    thr_str = (f"{float(_sc_thr):.2f}" if isinstance(_sc_thr, (int,float)) else "")
                                    title2 = L(f"Scala entro <= {thr_str} cents (\u2211={len(rows2)})", f"Scales within <= {thr_str} cents (N={len(rows2)})")
                                    print(title2)
                                    if _supports_ansi():
                                        print(Style.BOLD + fmt2(hdr2) + Style.RESET)
                                    else:
                                        print(fmt2(hdr2))
                                    for r2 in rows2:
                                        print(fmt2(r2))
    else:
        _noaud = L("Nessuna analisi audio: genero subito le tabelle di confronto.",
                   "No audio analysis: generating comparison tables immediately.")
        _noaud = f"{Style.DIM}{_noaud}{Style.RESET}" if _supports_ansi() else _noaud
        print(_noaud)

    export_comparison_tables(export_base, ratios_eff, basekey_eff, basenote,
                             args.diapason, compare_fund_hz,
                             args.compare_tet_align, subharm_fund_hz,
                             tet_divisions=args.compare_tet,
                             analysis_result=analysis_data,
                             delta_threshold_hz=getattr(args, 'delta_threshold_hz', 0.0))

    # PNG plot export for diapason analysis
    if getattr(args, 'diapason_analysis', False) and analysis_data:
        _png_path = export_diapason_plot_png(export_base, analysis_data, basenote, args.diapason)
        if _png_path:
            print(L(f"Grafico PNG diapason salvato: {_png_path}", f"Diapason PNG plot saved: {_png_path}"))
        else:
            print(L("Export PNG diapason non disponibile o fallito", "Diapason PNG export unavailable or failed"))

    # Render audio (frequency-controlled sinusoid) if requested
    if getattr(args, 'render', False) and getattr(args, 'audio_file', None) and analysis_data:
        _render_path = render_analysis_to_wav(args.audio_file, analysis_data, export_base)
        if _render_path:
            print(L(f"Render audio salvato: {_render_path}", f"Audio render saved: {_render_path}"))
        else:
            print(L("Render audio fallito o non disponibile", "Audio render failed or unavailable"))
        # Additionally render constant diapason tone for 30s
        try:
            a4_freq = estimate_a4_from_analysis(analysis_data, default_a4=float(args.diapason))
        except Exception:
            a4_freq = float(args.diapason) if isinstance(args.diapason, (int, float)) else DEFAULT_DIAPASON
        _dia_path = None
        if isinstance(a4_freq, (int, float)) and float(a4_freq) > 0:
            _dia_path = render_constant_tone_wav(export_base, float(a4_freq), duration_s=30.0, sr=48000)
        if _dia_path:
            print(L(f"Diapason render (sinusoide 30s) salvato: {_dia_path}",
                    f"Diapason render (30s sine) saved: {_dia_path}"))
        else:
            print(L("Render diapason non disponibile o fallito",
                    "Diapason render unavailable or failed"))

    if args.export_tun:
        write_tun_file(export_base, args.diapason, ratios_eff, basekey_eff, basenote, args.tun_integer)


if __name__ == "__main__":
    main()


# --- Scala (.scl) parsing and matching utilities ---
# IT: Parser minimale per file .scl (Scala) e matching con i centroidi.
# EN: Minimal parser for .scl (Scala) files and matching with cluster centers.

def parse_scl_file(path: str) -> Optional[dict]:
    """Parsa un file .scl secondo il formato Huygens-Fokker di base.
    Ritorna un dict: {'name': str, 'degrees_cents': List[float], 'file': str}
    dove degrees_cents sono ridotti a [0,1200).
    Se il file non è leggibile/parsabile, ritorna None.
    """
    try:
        with open(path, 'r', encoding='utf-8', errors='ignore') as f:
            lines = f.readlines()
    except Exception:
        return None

    # Rimuovi commenti e righe vuote; in .scl i commenti iniziano con '!'
    cleaned: List[str] = []
    for ln in lines:
        s = ln.strip()
        if not s:
            continue
        # drop lines that start with '!' entirely
        if s.startswith('!'):
            # but keep the first comment line as possible filename/description if name missing
            cleaned.append(s)
            continue
        # remove trailing inline comments starting with '!' if any
        if '!' in s:
            s = s.split('!', 1)[0].strip()
        if s:
            cleaned.append(s)

    if not cleaned:
        return None

    # The format: optional leading comment lines starting with '!'
    # Next non-comment line is the description (name). Following line has the number of notes.
    idx = 0
    # Skip leading comments for description, but remember them if needed
    while idx < len(cleaned) and cleaned[idx].startswith('!'):
        idx += 1
    if idx >= len(cleaned):
        return None
    name_line = cleaned[idx]
    idx += 1
    # Some files may keep description empty; fallback to filename later
    desc = name_line.strip()

    # Read number of notes
    if idx >= len(cleaned):
        return None
    try:
        n_deg = int(str(cleaned[idx]).strip().split()[0])
    except Exception:
        return None
    idx += 1

    degrees_cents: List[float] = []
    # Read next n_deg lines as degrees (can include comments after value that we already stripped)
    for _ in range(n_deg):
        if idx >= len(cleaned):
            break
        val = cleaned[idx].strip()
        idx += 1
        if not val:
            continue
        token = val.split()[0]
        cents_val: Optional[float] = None
        # ratio like 3/2 or 1.5/1.0 is allowed
        if '/' in token:
            try:
                num_str, den_str = token.split('/', 1)
                num = float(Fraction(num_str.strip()))
                den = float(Fraction(den_str.strip()))
                r = num / den if den != 0 else None
                if r and r > 0:
                    cents_val = 1200.0 * math.log2(float(r))
            except Exception:
                cents_val = None
        else:
            # plain number in cents
            try:
                cents_val = float(token)
            except Exception:
                cents_val = None
        if isinstance(cents_val, (int, float)) and math.isfinite(cents_val):
            # Reduce to [0,1200)
            c = float(cents_val) % 1200.0
            degrees_cents.append(c)
    if not degrees_cents:
        return None

    # Ensure unique sorted degrees
    degrees_sorted: List[float] = []
    for c in sorted(degrees_cents):
        if not degrees_sorted or abs(c - degrees_sorted[-1]) > 1e-6:
            degrees_sorted.append(c)

    return {
        'name': desc if desc else os.path.basename(path),
        'degrees_cents': degrees_sorted,
        'file': os.path.basename(path),
    }


def load_scales_from_dir(dir_path: str = 'scl') -> List[dict]:
    """Carica tutti i file .scl leggibili dalla directory indicata.
    Ritorna una lista di dict come parse_scl_file.
    """
    try:
        if not os.path.isdir(dir_path):
            return []
    except Exception:
        return []
    out: List[dict] = []
    try:
        for fn in os.listdir(dir_path):
            if not fn.lower().endswith('.scl'):
                continue
            fp = os.path.join(dir_path, fn)
            info = parse_scl_file(fp)
            if info and isinstance(info.get('degrees_cents'), list) and info['degrees_cents']:
                out.append(info)
    except Exception:
        return out
    return out


def match_scala_scales(cluster_centers: List[Tuple[float, int]], dir_path: str = 'scl') -> Tuple[Optional[dict], List[Tuple[int,float,int]]]:
    """Trova la scala .scl più vicina ai centroidi (ratio,count).
    - Converte i centroidi in cents [0,1200) e per ciascuno trova il grado più vicino della scala
      usando distanza circolare modulo 1200.
    - Ritorna (best_info, steps_map) dove best_info include 'name','file','avg_cents_error'.
    - steps_map: [(index, ratio_from_degree, count)].
    Se nessuna scala è disponibile, ritorna (None, []).
    """
    scales = load_scales_from_dir(dir_path)
    if not scales:
        return None, []

    # Prepare centers in cents
    centers_c = []  # (cents, count)
    for item in (cluster_centers or []):
        try:
            r, cnt = item
            if r and r > 0:
                c = (1200.0 * math.log2(float(r))) % 1200.0
                centers_c.append((float(c), int(cnt)))
        except Exception:
            continue
    if not centers_c:
        return None, []

    def circ_diff(a: float, b: float, mod: float = 1200.0) -> float:
        d = abs(float(a) - float(b)) % mod
        return d if d <= (mod * 0.5) else (mod - d)

    best = None  # tuple(err, scale_info, mapping)
    for sc in scales:
        degs = sc.get('degrees_cents') or []
        if not degs:
            continue
        total_werr = 0.0
        total_w = 0
        mapping: List[Tuple[int,float,int]] = []
        for (ci, cnt) in centers_c:
            try:
                idx = min(range(len(degs)), key=lambda k: circ_diff(ci, degs[k], 1200.0))
            except ValueError:
                continue
            err = circ_diff(ci, degs[idx], 1200.0)
            total_werr += float(err) * max(1, int(cnt))
            total_w += max(1, int(cnt))
            ratio = 2.0 ** ((float(degs[idx]) % 1200.0) / 1200.0)
            mapping.append((int(idx), float(ratio), int(cnt)))
        if total_w == 0:
            continue
        avg = total_werr / float(total_w)
        if best is None or avg < best[0]:
            best = (avg, sc, mapping)

    if best is None:
        return None, []

    err, sc, mapping = best
    info = {'name': sc.get('name', ''), 'file': sc.get('file', ''), 'avg_cents_error': float(err)}
    # Normalize mapping to ensure index 0 & ratio 1.0 present
    try:
        has_zero = any(int(s) == 0 for (s, _, __) in mapping)
    except Exception:
        has_zero = False
    steps = list(mapping)
    if has_zero:
        steps = [(0 if int(s)==0 else int(s), (1.0 if int(s)==0 else float(r)), int(c)) for (s,r,c) in steps]
    else:
        steps = [(0, 1.0, 0)] + steps
    steps = sorted(steps, key=lambda t: (int(t[0]), float(t[1])))
    return info, steps
