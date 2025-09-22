"""Table generation and export functionality for musical tuning systems.

This module handles the creation of comprehensive comparison tables and Excel workbooks
for THIN and SIM applications, providing detailed analysis and visualization of
musical tuning systems and their relationships.

Core Table Generation:
- Multi-system comparison tables with Custom, Harmonic, Subharmonic, and TET references
- Side-by-side frequency and cents analysis across different tuning approaches
- Automatic column formatting with conditional styling and zebra striping
- Mathematical precision preservation in tabular format

Excel Export Features:
- Rich formatting with color-coded columns for different tuning system types
- Integrated diapason analysis worksheets with F0 tracking results
- Audio analysis integration showing formant data and pitch stability
- Scala scale matching results with Top-5 rankings and confidence scores
- 12-TET mapping tables for practical instrument implementation

Audio Analysis Integration:
- Diapason estimation results with uncertainty quantification
- F0 and formant frequency tables with statistical analysis
- Scale inference tables showing detected patterns and matches
- Cross-validation between different analysis methods (CQT vs F0)

Data Presentation:
- Intelligent number formatting with configurable precision
- Ratio simplification and exact fraction display where appropriate
- Cents deviation calculations from standard references
- Harmonic series analysis with overtone relationships

Error Handling and Fallbacks:
- Graceful degradation when Excel dependencies are unavailable
- Text-based fallback exports for all table formats
- Comprehensive error reporting with multilingual messages
- Safe type conversion with default value handling

Dependencies:
- openpyxl for Excel export functionality (optional)
- Integration with utils module for mathematical operations
- Coordination with audio_analysis for data integration

Note: All scale references are based on Scala files and internal reference systems.
Historical DaMuSc integration has been removed for simplified maintenance.
"""

from typing import List, Optional, Tuple, Dict
import consts
import utils
import math


def _safe_get_analysis_field(analysis_result: Optional[Dict], field: str, default=None):
    """Safely extract field from analysis_result with isinstance check."""
    if isinstance(analysis_result, dict):
        return analysis_result.get(field, default)
    return default



def _safe_float_conversion(value, default: float = 0.0) -> float:
    """Safely convert value to float with default fallback."""
    try:
        return float(value) if value is not None else default
    except (ValueError, TypeError):
        return default


def _setup_openpyxl_imports():
    """Setup openpyxl imports with error handling."""
    openpyxl, styles = utils.lazy_import_openpyxl()
    return openpyxl, styles


def _setup_excel_color_fills():
    """Setup Excel color fills and return fill objects."""
    fills = utils.get_excel_color_fills()
    return {
        'custom': fills.get('custom'),
        'harmonic': fills.get('harmonic'),
        'subharmonic': fills.get('subharmonic'),
        'tet': fills.get('tet')
    }




def _calculate_12tet_mapping(hz_freq: float, a4_freq: float) -> Tuple[str, float]:
    """Calculate 12-TET mapping for a frequency. Returns (note_name, delta_cents)."""
    try:
        midi = int(round(consts.MIDI_A4 + consts.SEMITONES_PER_OCTAVE * math.log2(hz_freq / a4_freq)))
        midi = max(consts.MIDI_MIN, min(consts.MIDI_MAX, midi))
        name = utils.midi_to_note_name_12tet(midi)
        f_et = a4_freq * (2.0 ** ((midi - consts.MIDI_A4) / consts.SEMITONES_PER_OCTAVE))
        dc = 1200.0 * math.log2(hz_freq / f_et) if f_et > 0 else 0.0
        return name, dc
    except (ValueError, TypeError, AttributeError, ZeroDivisionError):
        return "", 0.0


def _add_detected_degrees_table(ws, detected_degrees: list, title: str, limit: int = 30):
    """Add detected degrees table to worksheet."""
    ws.append([title])
    ws.append(["Degree", "Theoretical (c)", "Detected (c)", "Error (c)", "Count"])

    for deg in detected_degrees[:limit]:
        ws.append([
            deg.get('degree', 0),
            f"{deg.get('theoretical_cents', 0.0):.1f}",
            f"{deg.get('detected_cents', 0.0):.1f}",
            f"{deg.get('cents_error', 0.0):+.1f}",
            deg.get('count', 0)
        ])

    if len(detected_degrees) > limit:
        ws.append([f"... and {len(detected_degrees) - limit} more detected degrees"])
    ws.append([])


def _add_missing_degrees_summary(ws, missing_degrees: list, limit: int = 20):
    """Add missing degrees summary to worksheet."""
    if missing_degrees:
        missing_count = len(missing_degrees)
        missing_list = [str(deg.get('degree', 0)) for deg in missing_degrees[:limit]]
        ws.append([f"Missing degrees ({missing_count}): " + ", ".join(missing_list)])
        if missing_count > limit:
            ws.append([f"... and {missing_count - limit} more missing degrees"])
        ws.append([])


def export_system_tables(output_base: str, ratios: List[float], basekey: int, basenote_hz: float) -> None:
    """Exports tables for the generated system."""
    computed, headers, txt_path = _compute_and_sort(basekey, basenote_hz, output_base, ratios)

    try:
        rows, widths = _compute_column_widths(basekey, computed, headers)
        with open(txt_path, "w", encoding="utf-8") as f:
            f.write(_format_table_row(headers, widths) + "\n")
            for row in rows:
                f.write(_format_table_row(row, widths) + "\n")
        utils.log_export_success(txt_path)
    except IOError as e:
        utils.log_export_error(txt_path, e)

    # Export Excel (opzionale)
    try:
        _export_excel_system(output_base, basekey, basenote_hz, computed, headers)
    except ImportError:
        _handle_openpyxl_error(f"{output_base}_system.xlsx")
    except (AttributeError, OSError, PermissionError, KeyError) as e:
        utils.log_export_error(f"{output_base}_system.xlsx", e)


def export_comparison_tables(output_base: str, ratios: List[float], basekey: int,
                             basenote_hz: float, diapason_hz: float,
                             compare_fund_hz: Optional[float] = None,
                             subharm_fund_hz: Optional[float] = None,
                             tet_divisions: int = 12,
                             analysis_result: Optional[Dict] = None,
                             delta_threshold_hz: float = 0.0) -> None:
    """Exports comparison tables with TET, harmonics/subharmonics and audio analysis."""

    base_cmp = compare_fund_hz if compare_fund_hz is not None else basenote_hz
    sub_base = subharm_fund_hz if subharm_fund_hz is not None else diapason_hz

    # Prepara dati ordinati
    computed = _compute_sorted_system(ratios, basekey, basenote_hz)
    custom_hz_list = [c[0] for c in computed]

    # Serie armoniche e subarmoniche
    harm_vals = _generate_harmonic_series(base_cmp)
    sub_vals = _generate_subharmonic_series(sub_base, custom_hz_list)

    # Allineamento sequenze
    harm_aligned = _align_sequence(harm_vals, custom_hz_list)
    sub_aligned = _align_sequence(sub_vals, custom_hz_list)

    # Export testo
    _export_comparison_text(output_base, computed, basekey, basenote_hz, diapason_hz,
                           base_cmp, sub_base, tet_divisions, harm_aligned, sub_aligned,
                           analysis_result, delta_threshold_hz)

    # Export Excel (opzionale)
    try:
        _export_comparison_excel(output_base, computed, basekey, basenote_hz, diapason_hz,
                                base_cmp, sub_base, tet_divisions, harm_aligned, sub_aligned,
                                analysis_result, delta_threshold_hz)
    except ImportError:
        _handle_openpyxl_error(f"{output_base}_compare.xlsx")
        if analysis_result:
            utils.generate_diapason_text_fallback(output_base, diapason_hz, basenote_hz, analysis_result, None)
    except AttributeError:
        _handle_openpyxl_error(f"{output_base}_compare.xlsx")
        if analysis_result:
            utils.generate_diapason_text_fallback(output_base, diapason_hz, basenote_hz, analysis_result, None)
    except (OSError, PermissionError, KeyError) as e:
        _handle_excel_export_error(f"{output_base}_compare.xlsx", e, analysis_result,
                                  (output_base, diapason_hz, basenote_hz, analysis_result, None))


def print_step_hz_table(ratios: List[float], basenote_hz: float) -> None:
    """Prints multi-column table with Step/Hz."""
    utils.print_step_hz_table(ratios, basenote_hz)


# --- Private helper functions ---

def _compute_and_sort(basekey: int, basenote_hz: float, output_base: str, ratios: List[float]) -> Tuple[List[Tuple[float, int, int, float]], List[str], str]:
    """Compute and sort system data."""
    computed = [(basenote_hz * float(r), i, basekey + i, float(r)) for i, r in enumerate(ratios)]
    computed.sort(key=lambda t: t[0])

    txt_path = f"{output_base}_system.txt"
    headers = utils.get_standard_system_headers()
    return computed, headers, txt_path


def _compute_column_widths(basekey: int, computed: List[Tuple[float, int, int, float]], headers: List[str]) -> Tuple[List[List[str]], List[int]]:
    """Compute column widths for table formatting."""
    rows = [[str(idx), str(basekey + idx), f"{ratio:.10f}", f"{hz:.6f}"] for idx, (hz, _, _, ratio) in enumerate(computed)]

    widths = [len(h) for h in headers]
    for row in rows:
        for col_idx, val in enumerate(row):
            widths[col_idx] = max(widths[col_idx], len(val))
    return rows, widths


def _format_table_row(vals: List[str], widths: List[int]) -> str:
    """Format table row with proper alignment."""
    try:
        return "  ".join(str(vals[i]).ljust(widths[i]) for i in range(len(vals)))
    except (TypeError, ValueError) as e:
        return "  ".join(str(x) for x in vals)
    except IndexError:
        return "  ".join(str(x) for x in vals)


def _export_excel_system(output_base: str, basekey: int, basenote_hz: float, computed: List[Tuple], headers: List[str]) -> None:
    """Export system to Excel format."""
    openpyxl, _ = utils.lazy_import_openpyxl()
    if not openpyxl:
        raise ImportError("openpyxl not available")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "System"

    utils.setup_excel_worksheet_formatting(ws, headers)

    # Base Hz anchor for formulas
    ws.cell(row=1, column=6, value="Base_Hz")
    ws.cell(row=2, column=6, value=float(basenote_hz))

    # Populate rows
    for step_idx, (hz, _, _, _ratio) in enumerate(computed):
        row_idx = step_idx + 2
        ws.cell(row=row_idx, column=1, value=step_idx)
        ws.cell(row=row_idx, column=2, value=basekey + step_idx)
        ws.cell(row=row_idx, column=3, value=f"=IFERROR(D{row_idx}/$F$2,\"\")")
        ws.cell(row=row_idx, column=4, value=float(hz))

    xlsx_path = f"{output_base}_system.xlsx"
    wb.save(xlsx_path)
    utils.log_export_success(xlsx_path)


def _compute_sorted_system(ratios: List[float], basekey: int, basenote_hz: float) -> List[Tuple[float, int, int, float]]:
    """Compute sorted system data for comparison tables."""
    # Reuse the compute_and_sort logic
    computed, _, _ = _compute_and_sort(basekey, basenote_hz, "", ratios)
    return computed


def _generate_harmonic_series(base_freq: float) -> List[float]:
    """Generate harmonic series up to MAX_HARMONIC_HZ."""
    harm_vals = []
    harm_n = 1
    while True:
        harm_freq = base_freq * harm_n
        if harm_freq > consts.MAX_HARMONIC_HZ:
            break
        harm_vals.append(harm_freq)
        harm_n += 1
    return harm_vals


def _generate_subharmonic_series(base_freq: float, custom_hz_list: List[float]) -> List[float]:
    """Generate subharmonic series with filtering."""
    sub_desc = []
    m = 1
    while True:
        sub_freq = base_freq / m
        if sub_freq < consts.MIN_SUBHARMONIC_HZ:
            break
        sub_desc.append(sub_freq)
        m += 1

    sub_vals = list(reversed(sub_desc))

    # Filter based on custom range
    min_custom = custom_hz_list[0] if custom_hz_list else base_freq
    cutoff_sub = max(consts.MIN_SUBHARMONIC_HZ, min_custom)
    return [v for v in sub_vals if v >= cutoff_sub]


def _align_sequence(seq: List[float], customs: List[float]) -> List[Optional[float]]:
    """Align sequence values to custom frequency points."""
    out: List[Optional[float]] = [None] * len(customs)
    p = 0
    for j in range(len(customs)):
        low = customs[j]
        high = customs[j + 1] if j + 1 < len(customs) else float('inf')
        if p < len(seq) and low <= seq[p] < high:
            out[j] = seq[p]
            p += 1
    return out


def _freq_to_note_name(freq: float, a4_hz: float) -> str:
    """Convert frequency to 12-TET note name."""
    try:
        if not (freq > 0 and a4_hz > 0):
            return ""
        midi = int(round(consts.MIDI_A4 + consts.SEMITONES_PER_OCTAVE * math.log2(freq / a4_hz)))
    except (ValueError, OverflowError):
        return ""

    midi = max(consts.MIDI_MIN, min(consts.MIDI_MAX, midi))
    names = ["C", "C#", "D", "D#", "E", "F", "F#", "G", "G#", "A", "A#", "B"]
    name = names[midi % consts.SEMITONES_PER_OCTAVE]
    octave = (midi // consts.SEMITONES_PER_OCTAVE) - 1
    return f"{name}{octave}"


def _tet_step_index(freq: float, base_freq: float, divs: int) -> int:
    """Return nearest TET step index from base_freq."""
    if freq <= 0 or base_freq <= 0:
        return 0
    return int(round(divs * math.log2(freq / base_freq)))


def _export_comparison_text(output_base: str, computed: List[Tuple], basekey: int, basenote_hz: float,
                           diapason_hz: float, base_cmp: float, sub_base: float, tet_divisions: int,
                           harm_aligned: List[Optional[float]], sub_aligned: List[Optional[float]],
                           analysis_result: Optional[Dict], delta_threshold_hz: float) -> None:
    """Export comparison table to text format."""
    txt_path = f"{output_base}_compare.txt"

    try:
        # Prepare audio analysis alignment
        audio_f0_idx, audio_f0_val, audio_formant_map = _prepare_audio_alignment(analysis_result, [c[0] for c in computed])

        headers = utils.get_comparison_table_headers()

        rows = []
        for row_i, (custom_hz, _step_idx, _midi, r) in enumerate(computed):
            harm_val = harm_aligned[row_i]
            sub_val = sub_aligned[row_i]

            # TET calculation
            if custom_hz > 0 and base_cmp > 0:
                tet_val = base_cmp * (2.0 ** (_tet_step_index(custom_hz, base_cmp, tet_divisions) / tet_divisions))
            else:
                tet_val = base_cmp

            # Format values with proximity indicators
            harm_str = f"{harm_val:.6f}" if harm_val is not None else ""
            d_har_str = f"{(custom_hz - harm_val):.6f}" if harm_val is not None else ""
            approx = "≈" if harm_val is not None and abs(custom_hz - harm_val) < consts.PROXIMITY_THRESHOLD_HZ else ""

            sub_str = f"{sub_val:.6f}" if sub_val is not None else ""
            d_sub_str = f"{(custom_hz - sub_val):.6f}" if sub_val is not None else ""

            tet_str = f"{tet_val:.6f}"
            tet_note = _freq_to_note_name(tet_val, diapason_hz)
            d_tet_str = f"{(custom_hz - tet_val):.6f}"

            # Audio alignments
            f0_str = f"{audio_f0_val:.6f}" if (audio_f0_idx is not None and audio_f0_idx == row_i and audio_f0_val) else ""
            af = audio_formant_map.get(row_i)
            a_formant_str = f"{af[0]:.6f}" if af else ""
            a_amp_str = f"{af[1]:.3f}" if af else ""
            d_f0_str = f"{(custom_hz - float(audio_f0_val)):.6f}" if f0_str else ""
            d_form_str = f"{(custom_hz - af[0]):.6f}" if af else ""

            # Apply threshold filtering
            if delta_threshold_hz > 0:
                harm_str, d_har_str = _apply_threshold_filter(harm_str, d_har_str, custom_hz, harm_val, delta_threshold_hz)
                sub_str, d_sub_str = _apply_threshold_filter(sub_str, d_sub_str, custom_hz, sub_val, delta_threshold_hz)
                tet_str, d_tet_str = _apply_threshold_filter(tet_str, d_tet_str, custom_hz, tet_val, delta_threshold_hz)
                if f0_str and abs(custom_hz - float(audio_f0_val)) < delta_threshold_hz:
                    f0_str = d_f0_str = ""
                if af and abs(custom_hz - af[0]) < delta_threshold_hz:
                    a_formant_str = a_amp_str = d_form_str = ""

            rows.append([
                str(row_i), str(basekey + row_i), f"{r:.10f}",
                f"{custom_hz:.6f}{approx}",
                f"{harm_str}{approx if harm_str else ''}",
                d_har_str, sub_str, d_sub_str,
                tet_str, tet_note, d_tet_str,
                f0_str, a_formant_str, a_amp_str,
                d_f0_str, d_form_str
            ])

        # Write with legend
        widths = [len(h) for h in headers]
        for row in rows:
            for col_idx, val in enumerate(row):
                widths[col_idx] = max(widths[col_idx], len(val))

        with open(txt_path, "w", encoding="utf-8") as f:
            _write_comparison_legend(f, basenote_hz, diapason_hz, base_cmp, sub_base, tet_divisions)
            f.write(_format_table_row(headers, widths) + "\n")
            for row in rows:
                f.write(_format_table_row(row, widths) + "\n")

        utils.log_export_success(txt_path)

    except IOError as e:
        utils.log_export_error(txt_path, e)


def _export_comparison_excel(output_base: str, computed: List[Tuple], basekey: int, basenote_hz: float,
                            diapason_hz: float, base_cmp: float, sub_base: float, tet_divisions: int,
                            harm_aligned: List[Optional[float]], sub_aligned: List[Optional[float]],
                            analysis_result: Optional[Dict], delta_threshold_hz: float) -> None:
    """Export comparison table to Excel format."""
    openpyxl, _ = utils.lazy_import_openpyxl()
    if not openpyxl:
        raise ImportError("openpyxl not available")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Compare"

    # Setup headers and formatting
    headers_xl = utils.get_comparison_table_headers_excel()
    utils.setup_excel_worksheet_formatting(ws, headers_xl)

    # Add reference information in header area (since legend sheet was removed)
    _add_reference_info_to_worksheet(ws, basenote_hz, diapason_hz, base_cmp, sub_base, tet_divisions)

    # Populate data rows
    _populate_comparison_excel_data(ws, computed, basekey, basenote_hz, base_cmp, tet_divisions, diapason_hz,
                                   harm_aligned, sub_aligned, analysis_result, delta_threshold_hz)

    # Diapason analysis is now exported separately to _diapason.xlsx to avoid corruption
    # Remove diapason sheet creation from compare file

    xlsx_path = f"{output_base}_compare.xlsx"
    wb.save(xlsx_path)
    utils.log_export_success(xlsx_path)


def export_diapason_excel(output_base: str, analysis_result: Optional[Dict], basenote_hz: float, diapason_hz: float) -> None:
    """Export diapason analysis to separate _diapason.xlsx file with complete info from diapason.txt."""
    if not analysis_result:
        return

    import math

    try:
        openpyxl, _ = utils.lazy_import_openpyxl()
        if not openpyxl:
            _handle_openpyxl_error(f"{output_base}_diapason.xlsx", "Diapason Excel export")
            return

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Diapason Analysis"

        # Font formatting for titles
        font_cls = _get_font_class()

        def add_title(title_text):
            _add_title_row(ws, title_text, font_cls)

        def add_section_title(title_text):
            ws.append([""])  # Empty row for spacing
            add_title(title_text)

        # Main diapason info
        add_title("DIAPASON – Reference systems")
        ws.append([""])

        # Use the corrected diapason logic
        user_set = bool(analysis_result.get('user_diapason_set', False))
        if user_set:
            a4_user_display = float(diapason_hz)
            bn_user_display = float(basenote_hz)
        else:
            a4_user_display = 440.0
            bn_user_display = float(basenote_hz) * (440.0 / float(diapason_hz))

        # Basic diapason info
        ws.append([f"A4_utente (Hz):", f"{a4_user_display:.6f}"])
        a4_est = _safe_get_analysis_field(analysis_result, 'diapason_est')
        ws.append([f"A4_estimated (Hz):", f"{float(a4_est):.6f}" if a4_est else "N/A"])

        if isinstance(a4_est, (int, float)):
            delta_hz = float(a4_est) - 440.0
            delta_pct = (delta_hz / 440.0) * 100.0
            ws.append([f"A4_delta_from_440Hz:", f"{delta_hz:.3f} Hz ({delta_pct:.2f}%)"])

        # Additional diapason fields if available
        sample_count = _safe_get_analysis_field(analysis_result, 'diapason_sample_count')
        if sample_count:
            ws.append([f"A4_sample_count:", str(sample_count)])

        conf_std = _safe_get_analysis_field(analysis_result, 'diapason_confidence_std')
        if conf_std:
            ws.append([f"A4_confidence_std:", f"±{float(conf_std):.3f} Hz"])

        conf_ci = _safe_get_analysis_field(analysis_result, 'diapason_confidence_interval')
        if conf_ci and isinstance(conf_ci, (list, tuple)) and len(conf_ci) >= 2:
            ws.append([f"A4_confidence_95CI:", f"[{float(conf_ci[0]):.3f}, {float(conf_ci[1]):.3f}] Hz"])

        ws.append([f"Basenote_Hz (utente):", f"{bn_user_display:.6f}"])
        bn_est = _safe_get_analysis_field(analysis_result, 'basenote_est_hz')
        ws.append([f"Basenote_Hz (estimated):", f"{float(bn_est):.6f}" if bn_est else "N/A"])

        bn_name = _safe_get_analysis_field(analysis_result, 'basenote_est_name_12tet')
        bn_midi = _safe_get_analysis_field(analysis_result, 'basenote_est_midi')
        if bn_name and bn_midi:
            ws.append([f"Basenote_12TET (estimated):", f"{bn_name} (MIDI {bn_midi})"])

        # Scala match info
        scala_info = _safe_get_analysis_field(analysis_result, 'scala_match_info')
        if scala_info and isinstance(scala_info, dict):
            add_section_title("Scala_match:")
            ws.append([f"  name:", scala_info.get('name', 'N/A')])
            ws.append([f"  file:", scala_info.get('file', 'N/A')])
            avg_err = scala_info.get('avg_error_cents')
            if isinstance(avg_err, (int, float)):
                ws.append([f"  avg_error_cents:", f"{float(avg_err):.2f}"])

        # Scala_map section (same format as diapason.txt)
        sc_steps = _safe_get_analysis_field(analysis_result, 'scala_match_steps')
        if sc_steps and isinstance(sc_steps, list):
            add_section_title("Scala_map (index, Ratio, Hz_from_Base(user), Hz_from_Base(estimated), Cents, Count, 12TET_user, DeltaCents_user, 12TET_est, DeltaCents_est):")
            ws.append(["Idx", "Ratio", "Hz_user", "Hz_est", "Cents", "Count", "12TET_user", "DeltaCents_user", "12TET_est", "DeltaCents_est"])

            # Show detected steps with count > 0
            detected_steps = [(idx, ratio, cnt) for (idx, ratio, cnt) in sc_steps if cnt > 0]
            for idx, ratio, count in detected_steps:
                if isinstance(ratio, (int, float)) and ratio > 0:
                    hz_user = bn_user_display * ratio
                    hz_est = float(bn_est) * ratio if bn_est else 0
                    cents_val = 1200.0 * math.log2(ratio) if ratio > 0 else 0
                    ws.append([idx, f"{ratio:.10f}", f"{hz_user:.6f}", f"{hz_est:.6f}", f"{cents_val:.2f}", count, "", "", "", ""])

        # Tuning inferred info
        tuning_inferred = _safe_get_analysis_field(analysis_result, 'tuning_inferred')
        if tuning_inferred and isinstance(tuning_inferred, dict):
            add_section_title("Tuning_inferred:")
            ws.append([f"  name:", tuning_inferred.get('name', 'N/A')])
            avg_err = tuning_inferred.get('avg_error_cents')
            if isinstance(avg_err, (int, float)):
                ws.append([f"  avg_error_cents:", f"{float(avg_err):.2f}"])

        # Inferred steps
        scale_steps = _safe_get_analysis_field(analysis_result, 'scale_steps')
        if scale_steps and isinstance(scale_steps, list):
            add_section_title("Inferred steps (index, Ratio, Hz_from_Base(user), Hz_from_Base(estimated), Cents, Count, 12TET_user, DeltaCents_user, 12TET_est, DeltaCents_est):")
            ws.append(["Idx", "Ratio", "Hz_user", "Hz_est", "Cents", "Count", "12TET_user", "DeltaCents_user", "12TET_est", "DeltaCents_est"])

            detected_steps = [(idx, ratio, cnt) for (idx, ratio, cnt) in scale_steps if cnt > 0]
            for idx, ratio, count in detected_steps:
                if isinstance(ratio, (int, float)) and ratio > 0:
                    hz_user = bn_user_display * ratio
                    hz_est = float(bn_est) * ratio if bn_est else 0
                    cents_val = 1200.0 * math.log2(ratio) if ratio > 0 else 0
                    ws.append([idx, f"{ratio:.10f}", f"{hz_user:.6f}", f"{hz_est:.6f}", f"{cents_val:.2f}", count, "", "", "", ""])

        # Tuning comparative info
        tuning_comparative = _safe_get_analysis_field(analysis_result, 'tuning_comparative')
        if tuning_comparative and isinstance(tuning_comparative, dict):
            add_section_title("Tuning_comparative:")
            ws.append([f"  name:", tuning_comparative.get('name', 'N/A')])
            avg_err = tuning_comparative.get('avg_error_cents')
            if isinstance(avg_err, (int, float)):
                ws.append([f"  avg_error_cents:", f"{float(avg_err):.2f}"])

        # Comparative steps
        scale_steps_comp = _safe_get_analysis_field(analysis_result, 'scale_steps_comp')
        if scale_steps_comp and isinstance(scale_steps_comp, list):
            add_section_title("Comparative steps (index, Ratio, Hz_from_Base(user), Hz_from_Base(estimated), Cents, Count, 12TET_user, DeltaCents_user, 12TET_est, DeltaCents_est):")
            ws.append(["Idx", "Ratio", "Hz_user", "Hz_est", "Cents", "Count", "12TET_user", "DeltaCents_user", "12TET_est", "DeltaCents_est"])

            detected_steps = [(idx, ratio, cnt) for (idx, ratio, cnt) in scale_steps_comp if cnt > 0]
            for idx, ratio, count in detected_steps:
                if isinstance(ratio, (int, float)) and ratio > 0:
                    hz_user = bn_user_display * ratio
                    hz_est = float(bn_est) * ratio if bn_est else 0
                    cents_val = 1200.0 * math.log2(ratio) if ratio > 0 else 0
                    ws.append([idx, f"{ratio:.10f}", f"{hz_user:.6f}", f"{hz_est:.6f}", f"{cents_val:.2f}", count, "", "", "", ""])

        # Reference systems (12-TET)
        add_section_title(f"12-TET (utente A4={a4_user_display:.6f} Hz)")
        ws.append(["Step", "Ratio", "Hz", "Cents"])

        for step in range(12):
            ratio = 2.0 ** (step / 12.0)
            hz_val = bn_user_display * ratio
            cents_val = step * 100.0
            ws.append([step, f"{ratio:.10f}", f"{hz_val:.6f}", f"{cents_val:.2f}"])

        # Save diapason file
        xlsx_path = f"{output_base}_diapason.xlsx"
        wb.save(xlsx_path)
        utils.log_export_success(xlsx_path)

    except ImportError:
        _handle_openpyxl_error(f"{output_base}_diapason.xlsx", "Diapason Excel export")
    except Exception as e:
        utils.log_export_error(f"{output_base}_diapason.xlsx", e)


def _prepare_audio_alignment(analysis_result: Optional[Dict], custom_hz_list: List[float]) -> Tuple[Optional[int], Optional[float], Dict[int, Tuple[float, float]]]:
    """Prepare audio analysis alignment data."""
    audio_f0_idx = None
    audio_f0_val = None
    audio_formant_map = {}

    if analysis_result is not None:
        f0_val = _safe_get_analysis_field(analysis_result, 'f0_hz')
        formants = _safe_get_analysis_field(analysis_result, 'formants')

        if f0_val and f0_val > 0 and custom_hz_list:
            audio_f0_idx = int(min(range(len(custom_hz_list)), key=lambda k: abs(custom_hz_list[k] - f0_val)))
            audio_f0_val = float(f0_val)

        if formants and custom_hz_list:
            _ff_list = formants
            if f0_val and f0_val > 0:
                _ff_list = [(ff, amp) for (ff, amp) in formants if ff >= f0_val]
            for (ff, amp) in sorted(_ff_list, key=lambda x: x[0]):
                idx = int(min(range(len(custom_hz_list)), key=lambda k: abs(custom_hz_list[k] - ff)))
                if idx not in audio_formant_map or amp > audio_formant_map[idx][1]:
                    audio_formant_map[idx] = (float(ff), float(max(0.0, min(1.0, amp))))

    return audio_f0_idx, audio_f0_val, audio_formant_map


def _apply_threshold_filter(str_val: str, delta_str: str, custom_hz: float, ref_val: Optional[float], threshold: float) -> Tuple[str, str]:
    """Apply delta threshold filtering to string values."""
    if ref_val is not None and abs(custom_hz - ref_val) < threshold:
        return "", ""
    return str_val, delta_str


def _write_comparison_legend(f, basenote_hz: float, diapason_hz: float, base_cmp: float, sub_base: float, tet_divisions: int) -> None:
    """Write comparison legend to file."""
    legend_lines = [
        "— Legenda confronto —",
        f"Basenote={basenote_hz:.6f} Hz; Diapason(A4)={diapason_hz:.2f} Hz; Fond. confronto={base_cmp:.6f} Hz; Fond. sub={sub_base:.6f} Hz; TET={tet_divisions}",
        "• Step/MIDI/Ratio: indice di scala, numero MIDI relativo, rapporto rispetto alla Basenote (ridotto a [1,2)).",
        "• Custom_Hz: frequency of the generated system (Basenote*Ratio).",
        "• Harmonic_Hz/Subharm_Hz: armonica/subarmonica più vicina; DeltaHz_* = Custom_Hz − Ref_Hz (Hz; segno indica sopra/sotto).",
        "• TET_Hz/TET_Note: passo del TET scelto; DeltaHz_TET = scostamento in Hz.",
        "• AudioF0_Hz/AudioFormant_Hz/Formant_RelAmp: data from audio analysis (if present); DeltaHz_F0/DeltaHz_Formant = deviation in Hz.",
        "• Simbolo ≈ indica prossimità visiva (< soglia).",
        ""
    ]
    for line in legend_lines:
        f.write(line + "\n")


def _get_font_class():
    """Get openpyxl Font class for formatting, returns None if not available."""
    try:
        from openpyxl.styles import Font as font_cls
        return font_cls
    except ImportError:
        return None


def _add_bold_cell(ws, row: int, col: int, text: str, font_cls=None) -> None:
    """Add a cell with bold text formatting."""
    ws.cell(row=row, column=col, value=text)
    if font_cls is not None:
        ws.cell(row=row, column=col).font = font_cls(bold=True)


def _add_title_row(ws, title: str, font_cls=None) -> None:
    """Add a title row with bold formatting."""
    ws.append([title])
    if font_cls is not None:
        ws.cell(row=ws.max_row, column=1).font = font_cls(bold=True)


def _handle_openpyxl_error(output_path: str, operation: str = "Excel export") -> None:
    """Handle openpyxl import/availability errors with consistent messaging."""
    print(f"openpyxl not available for {output_path}: {operation} skipped")


def _handle_excel_export_error(output_path: str, error: Exception, analysis_result: Optional[Dict] = None,
                              fallback_args: Optional[Tuple] = None) -> None:
    """Handle Excel export errors with fallback to text generation."""
    utils.log_export_error(output_path, error)
    if analysis_result and fallback_args:
        utils.generate_diapason_text_fallback(*fallback_args)


def _add_scale_steps_section(ws, section_title: str, steps_data: List, bn_user_display: float,
                           bn_est: Optional[float], add_section_title_func) -> None:
    """Add a scale steps section to Excel worksheet with consistent formatting."""
    if not steps_data or not isinstance(steps_data, list):
        return

    add_section_title_func(section_title)
    ws.append(["Idx", "Ratio", "Hz_user", "Hz_est", "Cents", "Count", "12TET_user", "DeltaCents_user", "12TET_est", "DeltaCents_est"])

    detected_steps = [(idx, ratio, cnt) for (idx, ratio, cnt) in steps_data if cnt > 0]
    for idx, ratio, count in detected_steps:
        if isinstance(ratio, (int, float)) and ratio > 0:
            hz_user = bn_user_display * ratio
            hz_est = float(bn_est) * ratio if bn_est else 0
            cents_val = 1200.0 * math.log2(ratio) if ratio > 0 else 0
            ws.append([idx, f"{ratio:.10f}", f"{hz_user:.6f}", f"{hz_est:.6f}", f"{cents_val:.2f}", count, "", "", "", ""])


def _add_reference_info_to_worksheet(ws, basenote_hz: float, diapason_hz: float, base_cmp: float, sub_base: float, tet_divisions: int) -> None:
    """Add reference information to the worksheet header area."""
    openpyxl, styles = _setup_openpyxl_imports()
    if not openpyxl:
        return

    try:
        from openpyxl.utils import get_column_letter
    except ImportError:
        return

    # Add reference info in columns beyond the main data (starting from column P)
    info_col = 16  # Column P

    # Headers and values
    ref_info = [
        ("Reference Values", ""),
        ("Basenote (Hz)", f"{basenote_hz:.6f}"),
        ("Diapason A4 (Hz)", f"{diapason_hz:.2f}"),
        ("Comparison Fund (Hz)", f"{base_cmp:.6f}"),
        ("Subharmonic Fund (Hz)", f"{sub_base:.6f}"),
        ("TET Divisions", str(tet_divisions))
    ]

    for row_idx, (label, value) in enumerate(ref_info, start=1):
        # Label in column P, value in column Q
        label_cell = ws.cell(row=row_idx, column=info_col, value=label)
        if value:  # Only set value if not empty
            ws.cell(row=row_idx, column=info_col + 1, value=value)

        # Bold formatting for labels
        if styles and label:
            label_cell.font = styles.Font(bold=True)

    # Set column widths
    ws.column_dimensions[get_column_letter(info_col)].width = 20
    ws.column_dimensions[get_column_letter(info_col + 1)].width = 15


def _create_legend_sheet(wb, basenote_hz: float, diapason_hz: float, base_cmp: float, sub_base: float, tet_divisions: int) -> None:
    """Create legend sheet in Excel workbook."""
    openpyxl, styles = _setup_openpyxl_imports()
    if not styles:
        return

    ws_leg = wb.create_sheet(title="Legend", index=0)
    legend_it = [
        "— Legenda confronto —",
        f"Basenote={basenote_hz:.6f} Hz; Diapason(A4)={diapason_hz:.2f} Hz; Fond. confronto={base_cmp:.6f} Hz; Fond. sub={sub_base:.6f} Hz; TET={tet_divisions}",
        "• Step/MIDI/Ratio: indice di scala, numero MIDI relativo, rapporto rispetto alla Basenote (ridotto a [1,2)).",
        "• Custom_Hz: frequency of the generated system (Basenote*Ratio).",
        "• Harmonic_Hz/Subharm_Hz: armonica/subarmonica più vicina; |DeltaHz_*| = scostamento in Hz (valore assoluto).",
        "• TET_Hz/TET_Note: passo del TET scelto; |DeltaHz_TET| = scostamento in Hz.",
        "• AudioF0_Hz/AudioFormant_Hz/Formant_RelAmp: from audio analysis (if present); |DeltaHz_F0|/|DeltaHz_Formant| = deviation in Hz.",
    ]

    for i, ln in enumerate(legend_it, start=1):
        ws_leg.append([str(ln)])
        ws_leg[f"A{i}"].alignment = styles.Alignment(wrap_text=True)
        if i == 1:
            ws_leg[f"A{i}"].font = styles.Font(bold=True)
    ws_leg.column_dimensions['A'].width = 120


def _populate_comparison_excel_data(ws, computed: List[Tuple], basekey: int, basenote_hz: float, base_cmp: float,
                                   tet_divisions: int, diapason_hz: float, harm_aligned: List[Optional[float]],
                                   sub_aligned: List[Optional[float]], analysis_result: Optional[Dict],
                                   delta_threshold_hz: float) -> None:
    """Populate Excel comparison data rows.

    Args:
        basenote_hz: User-specified base note frequency (used for validation and metadata)
        delta_threshold_hz: Threshold in Hz for filtering significant deltas (highlights values above threshold)

    Note:
        Delta threshold filtering is applied to highlight cells with significant deviations.
        Base note frequency is used for reference frequency validation.
    """
    openpyxl, _ = utils.lazy_import_openpyxl()
    if not openpyxl:
        return

    # Validate base note frequency
    if not isinstance(basenote_hz, (int, float)) or basenote_hz <= 0:
        print(f"Warning: Invalid base note frequency {basenote_hz} Hz in comparison table generation")
        basenote_hz = 261.626  # Fallback to C4

    # Color fills for different series
    fills = _setup_excel_color_fills()
    fill_custom = fills['custom']
    fill_harm = fills['harmonic']
    fill_sub = fills['subharmonic']
    fill_tet = fills['tet']

    custom_hz_list = [c[0] for c in computed]
    audio_f0_idx, audio_f0_val, audio_formant_map = _prepare_audio_alignment(analysis_result, custom_hz_list)

    for row_i, (custom_hz, _step_idx, _midi, r) in enumerate(computed):
        harm_val = harm_aligned[row_i]
        sub_val = sub_aligned[row_i]

        if custom_hz > 0 and base_cmp > 0:
            tet_val = base_cmp * (2.0 ** (_tet_step_index(custom_hz, base_cmp, tet_divisions) / tet_divisions))
        else:
            tet_val = base_cmp

        # Generate Excel formulas for values and deltas (removed unused idx variable)

        harm_cell = harm_val if harm_val is not None else None
        sub_cell = sub_val if sub_val is not None else None
        tet_note = _freq_to_note_name(tet_val, diapason_hz)

        # Audio values for this row
        f0_cell_val = None
        f0_hz_val = _safe_get_analysis_field(analysis_result, 'f0_hz')
        if (audio_f0_idx is not None and audio_f0_idx == row_i and
            isinstance(f0_hz_val, (int, float))):
            f0_cell_val = float(f0_hz_val)

        a_formant = audio_formant_map.get(row_i)
        a_formant_hz = a_formant[0] if a_formant else None
        a_formant_amp = a_formant[1] if a_formant else None

        # Calculate delta values directly (formulas were not being used)
        # Apply delta threshold filtering if specified
        if delta_threshold_hz > 0:
            # Filter out values below threshold
            if harm_cell is not None and abs(custom_hz - harm_cell) < delta_threshold_hz:
                harm_cell = None
            if sub_cell is not None and abs(custom_hz - sub_cell) < delta_threshold_hz:
                sub_cell = None
            if f0_cell_val is not None and abs(custom_hz - f0_cell_val) < delta_threshold_hz:
                f0_cell_val = None
            if a_formant_hz is not None and abs(custom_hz - a_formant_hz) < delta_threshold_hz:
                a_formant_hz = None
                a_formant_amp = None  # Clear amplitude if frequency is filtered

        # Ensure all numeric values are properly converted and valid
        safe_row_data = [
            int(row_i),
            int(basekey + row_i),
            utils.safe_numeric_value(r, 0.0),
            utils.safe_numeric_value(custom_hz, 0.0),
            utils.safe_numeric_value(harm_cell, ""),
            utils.safe_delta_calculation(custom_hz, harm_cell),
            utils.safe_numeric_value(sub_cell, ""),
            utils.safe_delta_calculation(custom_hz, sub_cell),
            utils.safe_numeric_value(tet_val, 0.0),
            str(tet_note) if tet_note else "",
            utils.safe_delta_calculation(custom_hz, tet_val),
            utils.safe_numeric_value(f0_cell_val, ""),
            utils.safe_numeric_value(a_formant_hz, ""),
            utils.safe_numeric_value(a_formant_amp, ""),
            utils.safe_delta_calculation(custom_hz, f0_cell_val),
            utils.safe_delta_calculation(custom_hz, a_formant_hz)
        ]

        ws.append(safe_row_data)

        # Apply color formatting
        row_num = ws.max_row
        if fill_custom:
            ws.cell(row=row_num, column=4).fill = fill_custom
        if fill_harm:
            ws.cell(row=row_num, column=5).fill = fill_harm
        if fill_sub:
            ws.cell(row=row_num, column=7).fill = fill_sub
        if fill_tet:
            ws.cell(row=row_num, column=9).fill = fill_tet
            ws.cell(row=row_num, column=10).fill = fill_tet


def _create_diapason_analysis_sheet(wb, analysis_result: Dict, basenote_hz: float, diapason_hz: float) -> None:
    """Create diapason analysis sheet in Excel workbook."""
    if not (_safe_get_analysis_field(analysis_result, 'diapason_est') or
            _safe_get_analysis_field(analysis_result, 'f0_list') or
            _safe_get_analysis_field(analysis_result, 'ratio_clusters')):
        return

    openpyxl, styles = _setup_openpyxl_imports()
    if not openpyxl:
        return

    ws2 = wb.create_sheet(title="Diapason")
    ws2.freeze_panes = "A2"

    # Summary section - match text format exactly
    diapason_est = _safe_get_analysis_field(analysis_result, 'diapason_est')
    a4_est_val = _safe_float_conversion(diapason_est)

    # Determine user A4 (display): if user explicitly set --diapason use it; otherwise use 440 Hz default
    try:
        user_set = bool(_safe_get_analysis_field(analysis_result, 'user_diapason_set', False))
    except (ValueError, TypeError):
        user_set = False
    if user_set:
        a4_user_eff = _safe_float_conversion(_safe_get_analysis_field(analysis_result, 'user_diapason_value', diapason_hz))
    else:
        # For --diapason-analysis, user A4 should be 440 Hz (default) unless explicitly specified
        a4_user_eff = 440.0  # Use standard A4=440 Hz as reference
    if not (isinstance(a4_user_eff, (int, float)) and a4_user_eff > 0):
        a4_user_eff = 440.0

    basenote_est = _safe_get_analysis_field(analysis_result, 'basenote_est_hz')
    bn_est_hz = _safe_float_conversion(basenote_est)

    # Calculate basenote_user based on A4=440 Hz reference
    # This is the basenote frequency that corresponds to A4=440 Hz
    try:
        # If user specified a diapason, basenote_hz is already based on that
        # If not, we need to calculate what the basenote would be at A4=440 Hz
        if user_set:
            bn_user_hz = float(basenote_hz)  # Use actual basenote from user's diapason
        else:
            # Calculate basenote that corresponds to A4=440 Hz
            bn_user_hz = float(basenote_hz) * (440.0 / float(diapason_hz))
    except (ValueError, TypeError, ZeroDivisionError):
        bn_user_hz = float(basenote_hz)

    # Calculate derived basenote_estimated if not available
    if (not isinstance(bn_est_hz, (int, float))) or not (bn_est_hz and bn_est_hz > 0):
        try:
            if isinstance(a4_est_val, (int, float)) and a4_est_val and bn_user_hz:
                # Scale basenote from A4=440 Hz to estimated A4
                bn_est_hz = bn_user_hz * (a4_est_val / 440.0)
            else:
                bn_est_hz = None
        except (ValueError, TypeError, AttributeError):
            bn_est_hz = None

    # Basenote 12-TET (estimated)
    bn12_name = ""
    bn12_midi = None
    if isinstance(bn_est_hz, (int, float)) and bn_est_hz and isinstance(a4_est_val, (int, float)) and a4_est_val:
        try:
            midi = int(round(consts.MIDI_A4 + consts.SEMITONES_PER_OCTAVE * math.log2(bn_est_hz / a4_est_val)))
            midi = max(consts.MIDI_MIN, min(consts.MIDI_MAX, midi))
            bn12_name = utils.midi_to_note_name_12tet(midi)
            bn12_midi = midi
        except (ValueError, TypeError, AttributeError):
            bn12_name = ""
            bn12_midi = None

    # Add summary matching text format
    ws2.append(["DIAPASON – Reference systems"])
    ws2.append([])
    ws2.append(["A4_utente (Hz):", f"{float(a4_user_eff):.6f}"])
    ws2.append(["A4_estimated (Hz):", f"{a4_est_val:.6f}" if isinstance(a4_est_val, (int, float)) else ""])
    ws2.append(["Basenote_Hz (utente):", f"{float(bn_user_hz):.6f}"])
    ws2.append(["Basenote_Hz (estimated):", f"{bn_est_hz:.6f}" if isinstance(bn_est_hz, (int, float)) else ""])
    if bn12_name and bn12_midi is not None:
        ws2.append(["Basenote_12TET (estimated):", f"{bn12_name} (MIDI {bn12_midi})"])
    else:
        ws2.append(["Basenote_12TET (estimated):", ""])
    ws2.append([])

    # Define bold title helper for this function
    font_cls = _get_font_class()

    def _add_bold_title(title: str) -> None:
        _add_title_row(ws2, title, font_cls)

    # Add Scala map section if available
    sc_steps = _safe_get_analysis_field(analysis_result, 'scala_match_steps')
    if sc_steps and isinstance(sc_steps, list) and sc_steps:
        _add_bold_title("Scala_map (index, Ratio, Hz_from_Base(user), Hz_from_Base(estimated), Cents, Count, 12TET_user, DeltaCents_user, 12TET_est, DeltaCents_est)")
        ws2.append(["Idx", "Ratio", "Hz_user", "Hz_est", "Cents", "Count", "12TET_user", "DeltaCents_user", "12TET_est", "DeltaCents_est"])
        for (idx, ratio, cnt) in sc_steps:
            _add_scala_map_row(ws2, idx, ratio, cnt, float(bn_user_hz), bn_est_hz, float(a4_user_eff), a4_est_val)
        ws2.append([])

    # Add inferred and comparative tuning sections
    _add_tuning_sections(ws2, analysis_result, float(bn_user_hz), bn_est_hz, float(a4_user_eff), a4_est_val, _add_bold_title)

    # Add reference systems sections
    _add_reference_systems(ws2, bn_user_hz, a4_user_eff, a4_est_val)

    # Add Scala sections derived from analysis (if available)

    # Scala matches
    try:
        sc_info = _safe_get_analysis_field(analysis_result, 'scala_match_info')
        # sc_top = _safe_get_analysis_field(analysis_result, 'scala_top_matches')
        sc_within = _safe_get_analysis_field(analysis_result, 'scala_within_matches')
        sc_thr = _safe_get_analysis_field(analysis_result, 'scala_within_threshold_cents')

        if sc_info:
            _add_bold_title("Scala – migliore corrispondenza (.scl)")
            ws2.append(["Pos", "Nome", "File", "Err (c)"])
            rows_sc = utils.process_scale_items_to_rows([sc_info], start_index=1)
            for r in rows_sc:
                ws2.append(r)
            ws2.append([])


        if sc_within:
            try:
                thr_str = f" (<= {float(sc_thr):.2f} cents)" if isinstance(sc_thr, (int, float)) else ""
            except (ValueError, TypeError):
                thr_str = ""
            _add_bold_title(f"Scala – entro soglia{thr_str}")
            ws2.append(["Pos", "Nome", "File", "Err (c)"])
            rows_sc3 = utils.process_scale_items_to_rows(sc_within, start_index=1)
            for r in rows_sc3:
                ws2.append(r)
            ws2.append([])
    except (ValueError, TypeError, AttributeError, KeyError):
        pass


    # NEW: Add fragment analysis section
    try:
        _add_fragment_analysis_section(ws2, analysis_result, _add_bold_title)
    except (ValueError, TypeError, AttributeError, KeyError):
        pass

    # Apply formatting
    _apply_diapason_sheet_formatting(ws2, styles)


def _add_scala_map_row(ws2, idx: int, ratio: float, cnt: int, bn_user: float, bn_est_hz, a4_user: float, a4_est):
    """Add a single Scala map row to the Excel worksheet."""
    try:
        rr = float(ratio)
        hz_user = bn_user * rr
        hz_est = (bn_est_hz * rr) if isinstance(bn_est_hz, (int, float)) else None
        cents = utils.ratio_to_cents(rr) % 1200.0

        # User 12-TET mapping
        name_u, dc_u = _calculate_12tet_mapping(hz_user, a4_user)

        # Estimated 12-TET mapping
        if isinstance(hz_est, (int, float)) and isinstance(a4_est, (int, float)):
            name_e, dc_e = _calculate_12tet_mapping(hz_est, a4_est)
        else:
            name_e, dc_e = "", 0.0

        ws2.append([
            idx, f"{rr:.10f}", f"{hz_user:.6f}",
            f"{hz_est:.6f}" if isinstance(hz_est, (int, float)) else "",
            f"{cents:.2f}", int(cnt), name_u, f"{dc_u:.2f}",
            name_e, f"{dc_e:.2f}"
        ])
    except (ValueError, TypeError, AttributeError):
        ws2.append([idx, f"{float(ratio):.10f}", "", "", "", int(cnt), "", "", "", ""])


def _add_tuning_sections(ws2, analysis_result: Dict, bn_user: float, bn_est_hz, a4_user: float, a4_est, add_bold_title):
    """Add inferred and comparative tuning sections to the Excel worksheet."""

    # Inferred tuning (primary)
    tinf = _safe_get_analysis_field(analysis_result, 'tuning_inferred')
    steps_inf = _safe_get_analysis_field(analysis_result, 'scale_steps')
    if tinf and isinstance(tinf, dict):
        nm = str(tinf.get('name', '')).strip()
        try:
            er = float(tinf.get('avg_cents_error', 0.0))
        except (ValueError, TypeError):
            er = 0.0
        add_bold_title("Tuning_inferred:")
        ws2.append(["name:", nm])
        ws2.append(["avg_error_cents:", f"{er:.2f}"])
        ws2.append([])

    if steps_inf and isinstance(steps_inf, list) and steps_inf:
        add_bold_title("Inferred steps (index, Ratio, Hz_from_Base(user), Hz_from_Base(estimated), Cents, Count, 12TET_user, DeltaCents_user, 12TET_est, DeltaCents_est)")
        ws2.append(["Idx", "Ratio", "Hz_user", "Hz_est", "Cents", "Count", "12TET_user", "DeltaCents_user", "12TET_est", "DeltaCents_est"])
        for (idx, ratio, cnt) in steps_inf:
            _add_scala_map_row(ws2, idx, ratio, cnt, bn_user, bn_est_hz, a4_user, a4_est)
        ws2.append([])

    # Comparative tuning
    tcmp = _safe_get_analysis_field(analysis_result, 'tuning_comparative')
    steps_cmp = _safe_get_analysis_field(analysis_result, 'scale_steps_comp')
    if tcmp and isinstance(tcmp, dict):
        nm = str(tcmp.get('name', '')).strip()
        try:
            er = float(tcmp.get('avg_cents_error', 0.0))
        except (ValueError, TypeError):
            er = 0.0
        add_bold_title("Tuning_comparative:")
        ws2.append(["name:", nm])
        ws2.append(["avg_error_cents:", f"{er:.2f}"])
        ws2.append([])

    if steps_cmp and isinstance(steps_cmp, list) and steps_cmp:
        add_bold_title("Comparative steps (index, Ratio, Hz_from_Base(user), Hz_from_Base(estimated), Cents, Count, 12TET_user, DeltaCents_user, 12TET_est, DeltaCents_est)")
        ws2.append(["Idx", "Ratio", "Hz_user", "Hz_est", "Cents", "Count", "12TET_user", "DeltaCents_user", "12TET_est", "DeltaCents_est"])
        for (idx, ratio, cnt) in steps_cmp:
            _add_scala_map_row(ws2, idx, ratio, cnt, bn_user, bn_est_hz, a4_user, a4_est)
        ws2.append([])


def _add_reference_systems(ws2, basenote_hz: float, diapason_hz: float, a4_est_val) -> None:
    """Add reference systems sections to diapason sheet."""

    # Font for formatting
    font_cls = _get_font_class()

    def add_section(title: str, rows: List) -> None:
        _add_title_row(ws2, title, font_cls)
        ws2.append(utils.get_reference_system_headers())
        for r in rows:
            ws2.append(r)
        ws2.append([])

    # User diapason sections
    rows_user = utils.build_reference_system_rows(float(basenote_hz), 12)
    add_section(f"12-TET (utente A4={float(diapason_hz):.2f} Hz)", rows_user["tet"])
    add_section("Pitagorico 12 (utente)", rows_user["py12"])
    add_section("Pitagorico 7 (utente)", rows_user["py7"])

    # Estimated sections if available
    if isinstance(a4_est_val, (int, float)) and a4_est_val > 0:
        base_est_hz = float(basenote_hz) * (a4_est_val / float(diapason_hz))
        rows_est = utils.build_reference_system_rows(base_est_hz, 12)
        add_section(f"12-TET (estimated A4={a4_est_val:.2f} Hz)", rows_est["tet"])
        add_section("Pythagorean 12 (estimated)", rows_est["py12"])
        add_section("Pythagorean 7 (estimated)", rows_est["py7"])


def _apply_diapason_sheet_formatting(ws2, styles) -> None:
    """Apply formatting to diapason analysis sheet."""
    try:
        from openpyxl.utils import get_column_letter
    except ImportError:
        return

    # Get styles for formatting (not currently used but available for future enhancements)
    # fills = utils.get_excel_color_fills()
    # borders = utils.get_excel_borders()

    # Bold for summary labels (column A)
    for r in range(1, min(20, ws2.max_row + 1)):
        cell = ws2.cell(row=r, column=1)
        if cell.value not in (None, ""):
            cell.font = styles.Font(bold=True)

    # Auto-width columns (limit max 40)
    for c in range(1, min(ws2.max_column + 1, 10)):
        max_len = 0
        for r in range(1, ws2.max_row + 1):
            v = ws2.cell(row=r, column=c).value
            if v is not None:
                max_len = max(max_len, len(str(v)))
        width = max(10, min(40, int(max_len * 1.2) + 2))
        ws2.column_dimensions[get_column_letter(c)].width = width


def _add_fragment_analysis_section(ws2, analysis_result: Dict, add_bold_title) -> None:
    """Add scale fragment analysis section to diapason worksheet."""

    fragment_analysis = _safe_get_analysis_field(analysis_result, 'fragment_analysis')
    if not fragment_analysis or not isinstance(fragment_analysis, dict):
        return

    add_bold_title("=== SCALE FRAGMENT ANALYSIS ===")
    ws2.append([])

    # Inferred system fragment
    inferred_frag = fragment_analysis.get('inferred_system')
    if inferred_frag and isinstance(inferred_frag, dict):
        system_name = str(inferred_frag.get('system_name', 'Unknown'))
        total_degrees = inferred_frag.get('total_degrees', 0)
        detected_count = inferred_frag.get('detected_count', 0)
        coverage_percent = inferred_frag.get('coverage_percent', 0.0)

        add_bold_title(f"Inferred System Fragment: {system_name}")
        ws2.append([f"Coverage: {detected_count}/{total_degrees} degrees ({coverage_percent:.1f}%)"])
        ws2.append([])

        # Detected degrees table
        detected = inferred_frag.get('detected_degrees', [])
        if detected:
            _add_detected_degrees_table(ws2, detected, "Detected Degrees (Inferred System)")

        # Missing degrees summary
        missing = inferred_frag.get('missing_degrees', [])
        _add_missing_degrees_summary(ws2, missing)

    # Scala fragment
    scala_frag = fragment_analysis.get('scala_match')
    if scala_frag and isinstance(scala_frag, dict):
        scala_name = str(scala_frag.get('scala_name', 'Unknown'))
        scala_file = str(scala_frag.get('scala_file', 'Unknown'))
        total_degrees = scala_frag.get('total_degrees', 0)
        detected_count = scala_frag.get('detected_count', 0)
        coverage_percent = scala_frag.get('coverage_percent', 0.0)

        add_bold_title(f"Scala Match Fragment: {scala_name}")
        ws2.append([f"File: {scala_file}"])
        ws2.append([f"Coverage: {detected_count}/{total_degrees} degrees ({coverage_percent:.1f}%)"])
        ws2.append([])

        # Detected degrees table
        detected = scala_frag.get('detected_degrees', [])
        if detected:
            _add_detected_degrees_table(ws2, detected, "Detected Degrees (Scala Match)")

        # Missing degrees summary
        missing = scala_frag.get('missing_degrees', [])
        _add_missing_degrees_summary(ws2, missing)